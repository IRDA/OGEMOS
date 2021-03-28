from openpyxl import load_workbook
import requests
from GUI.fonction_utilitaire import is_valid_string


def tests():
    APPORT_DEFAUT = 10.0
    CULTURE_SECONDAIRE_DEFAUT = 1.0
    TRAVAIL_DU_SOL_DEFAUT = "Travail réduit"

    wb = load_workbook(
        filename='C:\\Users\\Samuel\\Documents\\Stage IRDA\\Donnees__validation_Ogemos_Logiag_2021-03-06_SC_v2.xlsx')
    donnees_champs = wb['Données champs']
    donnees_regies = wb['Données régies']
    donnees_amendements = wb['Données amendements']

    row_index = 2
    column_index = 1
    while donnees_amendements.cell(row=row_index, column=column_index).value is not None:
        id_amendement = donnees_amendements.cell(row=row_index, column=column_index).value
        column_index += 1
        pourcentage_humidite = donnees_amendements.cell(row=row_index, column=column_index).value
        column_index += 1
        carbon_total = donnees_amendements.cell(row=row_index, column=column_index).value
        column_index = 1
        row_index += 1
        if str(id_amendement).isalnum() and str(pourcentage_humidite).isalnum() and str(carbon_total).isalnum():
            amendement_dict = {"amendement": id_amendement,
                               "pourcentage_humidite": str(pourcentage_humidite),
                               "carbon_total": str(carbon_total)}
            response = requests.post('http://localhost:5000/api/ajout-amendement', json=amendement_dict)
    simulation = {
        "simulations": [{"nom_simulation": "Test OGEMOS", "duree_projection": 13, "annee_initiale_projection": 2008,
                         "annee_finale_projection": 2020,
                         "entreprise_agricole": {"nom": "Test OGEMOS", "champs": []}}]}
    champ_dict = {}
    row_index = 4
    column_index = 1
    while donnees_champs.cell(row=row_index, column=column_index).value is not None:
        id_champ = str(donnees_champs.cell(row=row_index, column=column_index).value)
        column_index += 1
        centroide_champ_long = donnees_champs.cell(row=row_index, column=column_index).value
        column_index += 1
        centroide_champ_lat = donnees_champs.cell(row=row_index, column=column_index).value
        column_index += 1
        taux_mos_initial = donnees_champs.cell(row=row_index, column=column_index).value
        column_index += 1
        annee_taux_mos_initial = donnees_champs.cell(row=row_index, column=column_index).value
        column_index += 1
        taux_mos_final = donnees_champs.cell(row=row_index, column=column_index).value
        column_index += 1
        annee_taux_mos_final = donnees_champs.cell(row=row_index, column=column_index).value
        column_index += 1
        municipalite = donnees_champs.cell(row=row_index, column=column_index).value
        column_index += 1
        code_postal = donnees_champs.cell(row=row_index, column=column_index).value
        column_index += 1
        serie_sol = donnees_champs.cell(row=row_index, column=column_index).value
        column_index += 1
        groupe_textural = donnees_champs.cell(row=row_index, column=column_index).value
        column_index += 1
        classe_de_drainage = donnees_champs.cell(row=row_index, column=column_index).value
        column_index += 1
        masse_volumique_apparente = donnees_champs.cell(row=row_index, column=column_index).value
        column_index += 1
        profondeur = donnees_champs.cell(row=row_index, column=column_index).value
        column_index = 1
        row_index += 1
        if (is_valid_string(str(id_champ)) and is_valid_string(str(centroide_champ_long)) and is_valid_string(
                str(centroide_champ_lat)) and is_valid_string(
                str(taux_mos_initial)) and is_valid_string(str(annee_taux_mos_initial)) and
                is_valid_string(str(taux_mos_final)) and is_valid_string(str(annee_taux_mos_final)) and is_valid_string(
                    str(municipalite)) and is_valid_string(str(code_postal)) and
                is_valid_string(str(serie_sol)) and is_valid_string(str(groupe_textural)) and is_valid_string(
                    str(classe_de_drainage)) and is_valid_string(str(masse_volumique_apparente)) and
                is_valid_string(str(profondeur))):
            if profondeur == -1:
                #profondeur = 20
                profondeur = None
            if groupe_textural == -1:
                groupe_textural = "Groupe 2 (texture moyenne)"
            if classe_de_drainage == -1:
                classe_de_drainage = "Modérément bien drainé"
            if masse_volumique_apparente == -1:
                #masse_volumique_apparente = 1.35
                masse_volumique_apparente = None

            champ = {"nom": id_champ,
                     "zones_de_gestion": [
                         {"taux_matiere_organique": float(taux_mos_initial), "municipalite": municipalite,
                          "groupe_textural": groupe_textural, "classe_de_drainage": classe_de_drainage,
                          "masse_volumique_apparente": masse_volumique_apparente,
                          "profondeur": profondeur,
                          "superficie_de_la_zone": 1, "regies_sol_et_culture_projection": [],
                          "regies_sol_et_culture_historique": []}]}
            champ_dict[id_champ] = champ
    row_index = 4
    column_index = 1
    while donnees_regies.cell(row=row_index, column=column_index).value is not None:
        id_champ = str(donnees_regies.cell(row=row_index, column=column_index).value)
        column_index += 1
        annee_culture = donnees_regies.cell(row=row_index, column=column_index).value
        column_index += 1
        culture_principale = donnees_regies.cell(row=row_index, column=column_index).value
        column_index += 1
        rendement_culture_principale = donnees_regies.cell(row=row_index, column=column_index).value
        column_index += 1
        pourcentage_tige_paille = donnees_regies.cell(row=row_index, column=column_index).value
        column_index += 1
        production_recoltee = donnees_regies.cell(row=row_index, column=column_index).value
        column_index += 1
        travail_du_sol = donnees_regies.cell(row=row_index, column=column_index).value
        column_index += 1
        culture_secondaire = donnees_regies.cell(row=row_index, column=column_index).value
        column_index += 1
        rendement_culture_secondaire = donnees_regies.cell(row=row_index, column=column_index).value
        column_index += 1
        id_amendement = donnees_regies.cell(row=row_index, column=column_index).value
        column_index += 1
        apport = donnees_regies.cell(row=row_index, column=column_index).value
        column_index = 1
        row_index += 1
        if (is_valid_string(str(id_champ)) and is_valid_string(str(annee_culture)) and is_valid_string(
                str(culture_principale)) and
                is_valid_string(str(pourcentage_tige_paille)) and
                is_valid_string(str(production_recoltee)) and is_valid_string(str(travail_du_sol)) and is_valid_string(
                    str(culture_secondaire)) and
                is_valid_string(str(rendement_culture_secondaire)) and is_valid_string(str(id_champ)) and
                is_valid_string(str(apport))):
            if pourcentage_tige_paille == -1:
                pourcentage_tige_paille = None
            else:
                pourcentage_tige_paille = float(pourcentage_tige_paille)
            if travail_du_sol == -1:
                travail_du_sol = TRAVAIL_DU_SOL_DEFAUT
            if rendement_culture_principale == -1:
                rendement_culture_principale = None
            if production_recoltee == 1:
                production_recoltee = True
            else:
                production_recoltee = False
            if culture_secondaire != -1 and rendement_culture_secondaire != -1:
                rendement_culture_secondaire = float(rendement_culture_secondaire)
            elif culture_secondaire != -1 and rendement_culture_secondaire == -1:
                rendement_culture_secondaire = CULTURE_SECONDAIRE_DEFAUT
            else:
                culture_secondaire = None
                rendement_culture_secondaire = None
            amendement_list = []
            if id_amendement != -1 and apport != -1:
                amendements = id_amendement.split(",")
                if "," in str(apport):
                    apports = apport.split(",")
                    index = 0
                    for amendement in amendements:
                        amendement_list.append({"amendement": amendement,
                                                "apport": float(apports[index])})
                        index += 1
                else:
                    amendement_list.append({"amendement": id_amendement,
                                            "apport": float(apport)})

            elif id_amendement != -1 and apport == -1:
                amendements = id_amendement.split(",")
                for amendement in amendements:
                    amendement_list.append({"amendement": amendement,
                                            "apport": float(APPORT_DEFAUT)})
            else:
                amendement_list.append(({"amendement": None,
                                         "apport": None}))
            regie = {
                "culture_principale": {"culture_principale": culture_principale,
                                       "rendement": rendement_culture_principale,
                                       "pourcentage_tige_exporte": pourcentage_tige_paille,
                                       "produit_recolte": production_recoltee,
                                       "pourcentage_humidite": None},
                "culture_secondaire": {"culture_secondaire": culture_secondaire,
                                       "rendement": rendement_culture_secondaire},
                "amendements": amendement_list,
                "travail_du_sol": {"travail_du_sol": travail_du_sol}}
            champ_dict[str(id_champ)]["zones_de_gestion"][0]["regies_sol_et_culture_projection"].append(regie)
    for champ in champ_dict:
        simulation["simulations"][0]["entreprise_agricole"]["champs"].append(champ_dict[champ])
    test_response = requests.post("http://localhost:5000/api/icbm-bilan", json=simulation)
    association_champ_teneur_finale = {}
    test_response_json = test_response.json()
    bilan_simulations_list = test_response_json["bilans_des_simulations"]
    simulation = bilan_simulations_list[0]
    bilan_des_champs_list = simulation["bilans_des_champs"]
    for champ in bilan_des_champs_list:
        nom_du_champ = champ["nom_du_champs"]
        blian_zones_list = champ["bilans_des_zones"]
        teneur_finale_projetee_mos = blian_zones_list[0]["teneur_finale_projetee"]
        association_champ_teneur_finale[str(nom_du_champ)] = teneur_finale_projetee_mos
    row_index = 4
    column_index = 1
    while str(donnees_champs.cell(row=row_index, column=column_index).value) != "None":
        donnees_champs.cell(row=row_index, column=15, value=association_champ_teneur_finale[str(donnees_champs.cell(row=row_index, column=column_index).value)])
        row_index += 1

    wb.save("C:\\Users\\Samuel\\Documents\\Stage IRDA\\Donnees__validation_Ogemos_Logiag_2021-03-06_SC_v2.xlsx")


if __name__ == "__main__":
    tests()
