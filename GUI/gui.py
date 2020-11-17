import copy
import json
import subprocess
import tkinter as tk
import os
import sys
from tkinter import filedialog
from tkinter import messagebox
from tkinter import ttk

import psutil
import requests
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment

import GUI.fonction_utilitaire as util

os.chdir(sys._MEIPASS)
sp = subprocess.Popen(
    "API_OGEMOS.exe",
    shell=True)


def set_globals():
    try:
        global municipalites_supportees
        municipalites_supportees_temporaire = requests.get("http://localhost:5000/api/get-municipalite")
        municipalites_supportees = municipalites_supportees_temporaire.json()["municipalites_supportees"]
        municipalites_supportees.sort()

        global groupes_texturaux_supportees
        groupes_texturaux_supportees_temporaire = requests.get("http://localhost:5000/api/get-groupe_textural")
        groupes_texturaux_supportees = groupes_texturaux_supportees_temporaire.json()["groupes_texturaux_supportees"]
        groupes_texturaux_supportees.sort()

        global classes_de_drainage_supportees
        classe_de_drainage_supportees_temporaire = requests.get("http://localhost:5000/api/get-classe_de_drainage")
        classes_de_drainage_supportees = classe_de_drainage_supportees_temporaire.json()[
            "classes_de_drainage_supportees"]
        classes_de_drainage_supportees.sort()

        global cultures_principales_supportees
        culture_principale_supportees_temporaire = requests.get("http://localhost:5000/api/get-culture_principale")
        cultures_principales_supportees = culture_principale_supportees_temporaire.json()[
            "cultures_principales_supportees"]
        cultures_principales_supportees.sort()

        global types_travail_du_sol_supportes
        travail_du_sol_supportees_temporaire = requests.get("http://localhost:5000/api/get-travail_du_sol")
        types_travail_du_sol_supportes = travail_du_sol_supportees_temporaire.json()["types_travail_du_sol_supportes"]
        types_travail_du_sol_supportes.sort()

        global cultures_secondaires_supportees
        culture_secondaire_supportees_temporaire = requests.get("http://localhost:5000/api/get-culture_secondaire")
        cultures_secondaires_supportees = culture_secondaire_supportees_temporaire.json()[
            "cultures_secondaires_supportees"]
        cultures_secondaires_supportees.sort()

        global amendements_supportees
        amendements_supportees_temporaire = requests.get("http://localhost:5000/api/get-amendement")
        amendements_supportees = amendements_supportees_temporaire.json()["amendements_supportees"]
        amendements_supportees.sort()
    except requests.ConnectionError:
        print("Connection Error occured!")
        set_globals()


def initialize_globals():
    set_globals()

    def initialize_run_gui():

        global furthest_left_tab_index_simulation
        furthest_left_tab_index_simulation = 0
        global furthest_right_tab_index_simulation
        furthest_right_tab_index_simulation = 0

        global min_index_simulation
        min_index_simulation = 0
        global max_index_simulation
        max_index_simulation = 0

        global furthest_left_tab_index_champs
        furthest_left_tab_index_champs = 0
        global furthest_right_tab_index_champs
        furthest_right_tab_index_champs = 0

        global min_index_champs
        min_index_champs = 0
        global max_index_champs
        max_index_champs = 0

        global furthest_left_tab_index_zone
        furthest_left_tab_index_zone = 0
        global furthest_right_tab_index_zone
        furthest_right_tab_index_zone = 0

        global min_index_zone
        min_index_zone = 0
        global max_index_zone
        max_index_zone = 0

        global nombre_simulations
        nombre_simulations = 0

        global duree_simulation
        duree_simulation = []

        global simulations_chargees
        simulations_chargees = False

        global filename
        filename = None

        global plan_gestion_filename
        plan_gestion_filename = None

        def kill(proc_pid):
            process = psutil.Process(proc_pid)
            for proc in process.children(recursive=True):
                proc.kill()
            process.kill()

        root = tk.Tk()
        root.title("OGEMOS")
        root.resizable(0, 0)
        mainframe = ttk.Frame(root)
        mainframe.grid(row=0, column=0, ipadx=5, ipady=5)

        def on_mousewheel(event):
            event_coordinates = root.winfo_pointerxy()
            try:
                canvas_path = str(root.winfo_containing(event_coordinates[0], event_coordinates[1]))
                if "canvas" in canvas_path:
                    canvas_found = False
                    widget_list = canvas_path.split(".")
                    real_canvas_path = ""
                    for widget in reversed(widget_list):
                        if canvas_found:
                            if widget != "":
                                real_canvas_path = "." + widget + real_canvas_path
                        else:
                            if "canvas" in widget:
                                real_canvas_path = "." + widget
                                canvas_found = True
                    canvas = root.nametowidget(real_canvas_path)
                    canvas.yview_scroll(int(-1 * (event.delta / 120)), "units")
            except KeyError:
                pass

        root.bind_all("<MouseWheel>", on_mousewheel)
        root.bind_all("<Button-4>", on_mousewheel)
        root.bind_all("<Button-5>", on_mousewheel)

        root.unbind_class("TCombobox", "<MouseWheel>")
        root.unbind_class("TCombobox", "<ButtonPress-4>")
        root.unbind_class("TCombobox", "<ButtonPress-5>")

        def filter_combobox_values(combobox, values):
            current_string_in_combobox = combobox.get()
            if current_string_in_combobox == "":
                combobox.configure(values=values)
            else:
                new_values = []
                for item in values:
                    if current_string_in_combobox.lower() in item.lower():
                        new_values.append(item)
                combobox.configure(values=new_values)

        def closing_root_protocol():
            kill(sp.pid)
            root.destroy()

        root.protocol("WM_DELETE_WINDOW", closing_root_protocol)

        def run_gui(frame):
            def show_entreprise_set_up(frame_entreprise):
                def get_information_entreprise():
                    global nom_entreprise
                    nom_entreprise = nom_entreprise_entry.get()
                    nombre_de_champs_string = nombre_de_champs_entry.get()
                    if nombre_de_champs_string.isdigit() and int(nombre_de_champs_string) > 0:
                        global nombre_de_champs
                        nombre_de_champs = int(nombre_de_champs_string)
                        for widget in frame_entreprise.winfo_children():
                            widget.destroy()
                        show_creation_champs(frame)
                    else:
                        messagebox.showwarning("Warning",
                                               "L'entrée \"Nombre de champs\" est invalide.Veuillez entrer un nombre naturel plus grand que 0.")

                nom_entreprise_label = ttk.Label(frame_entreprise, text="Nom de l'entreprise: ")
                nom_entreprise_entry = ttk.Entry(frame_entreprise)
                nombre_de_champs_label = ttk.Label(frame_entreprise, text="Nombre de champs: ")
                nombre_de_champs_entry = ttk.Entry(frame_entreprise)
                nombre_de_champs_entry.insert(0, "1")

                nom_entreprise_label.grid(row=0, column=0, sticky="E")
                nom_entreprise_entry.grid(row=0, column=1, sticky="W")
                nombre_de_champs_label.grid(row=1, column=0, sticky="E")
                nombre_de_champs_entry.grid(row=1, column=1, sticky="W")
                frame_entreprise.grid_columnconfigure(0, pad=10)
                frame_entreprise.grid_columnconfigure(1, pad=10)
                frame_entreprise.grid_rowconfigure(0, minsize=30)
                frame_entreprise.grid_rowconfigure(1, minsize=30)
                frame_entreprise.grid_rowconfigure(2, minsize=30)

                creer_bouton = ttk.Button(frame_entreprise, text="Créer", command=get_information_entreprise)
                creer_bouton.grid(row=2, columnspan=2, column=0, pady=3)
                retour_menu_principal_button = ttk.Button(frame_entreprise, text="Retour au menu principal",
                                                          command=retour_au_menu_principal)
                retour_menu_principal_button.grid(columnspan=2, row=3, column=0, pady=3)

            def show_creation_champs(frame_champs_list):
                def get_information_champs(scrollable_frame):
                    global information_champs
                    information_champs = []
                    champs_valides = True
                    for scrollable_frame_widget in scrollable_frame.winfo_children():
                        if isinstance(scrollable_frame_widget, ttk.LabelFrame):
                            grid_slave0_1 = scrollable_frame_widget.grid_slaves(row=0, column=1)
                            for entry in grid_slave0_1:
                                nom_du_champs = entry.get()
                            grid_slave1_1 = scrollable_frame_widget.grid_slaves(row=1, column=1)
                            for entry in grid_slave1_1:
                                nombre_de_zone_de_gestion = entry.get()
                            if nombre_de_zone_de_gestion.isdigit() and int(nombre_de_zone_de_gestion) > 0:
                                if len(nom_du_champs) < 13:
                                    information_champs.append({"nom_du_champs": nom_du_champs,
                                                               "nombre_de_zone_de_gestion": nombre_de_zone_de_gestion,
                                                               "information_zone_de_gestion": []})
                                else:
                                    champs_valides = False
                                    information_champs = []
                                    messagebox.showwarning("Warning",
                                                           "Le nom du champ devrait être composé de 12 caractères ou moins.")
                                    break

                            else:
                                champs_valides = False
                                information_champs = []
                                messagebox.showwarning("Warning",
                                                       "Une entrée \"Nombre de zones gestion\" est invalide.Veuillez entrer un nombre naturel plus grand que 0.")
                                break

                    if champs_valides:
                        for widget in frame_champs_list.winfo_children():
                            widget.destroy()

                        show_creation_zone_de_gestion(frame_champs_list)

                canvas = tk.Canvas(frame_champs_list)
                scrollbar = ttk.Scrollbar(frame_champs_list, orient="vertical", command=canvas.yview)
                scrollable_frame = ttk.Frame(canvas)
                scrollable_frame.bind("<Configure>", lambda e: canvas.configure(scrollregion=canvas.bbox("all")))
                canvas.create_window((0, 0), window=scrollable_frame, anchor="nw")
                canvas.configure(yscrollcommand=scrollbar.set)

                for index_champ in range(int(nombre_de_champs)):
                    champs_frame = ttk.LabelFrame(scrollable_frame, text="Champs " + str(index_champ + 1))
                    nom_du_champs_label = ttk.Label(champs_frame, text="Nom du champ: ")
                    nom_du_champs_entry = ttk.Entry(champs_frame)
                    nombre_de_zone_de_gestion_label = ttk.Label(champs_frame, text="Nombre de zones gestion: ")
                    nombre_de_zone_de_gestion_entry = ttk.Entry(champs_frame)
                    nombre_de_zone_de_gestion_entry.insert(0, "1")
                    nom_du_champs_label.grid(row=0, column=0, sticky="w", pady=3, padx=5)
                    nom_du_champs_entry.grid(row=0, column=1, sticky="w", pady=3, padx=5)
                    nombre_de_zone_de_gestion_label.grid(row=1, column=0, sticky="w", pady=3, padx=5)
                    nombre_de_zone_de_gestion_entry.grid(row=1, column=1, sticky="w", pady=3, padx=5)

                    champs_frame.pack(fill="both", padx=55, pady=5)

                canvas.pack(side="left", ipadx=10)
                scrollbar.pack(side="right", fill="y")
                frame_bouton = ttk.Frame(scrollable_frame)
                creation_champs_bouton = ttk.Button(frame_bouton, text="Créer",
                                                    command=lambda: get_information_champs(scrollable_frame))
                creation_champs_bouton.grid(row=0, column=0, pady=3)
                retour_menu_principal_button = ttk.Button(frame_bouton, text="Retour au menu principal",
                                                          command=retour_au_menu_principal)
                retour_menu_principal_button.grid(row=1, column=0, pady=3)
                frame_bouton.pack()

            def show_creation_zone_de_gestion(zone_de_gestion_mainframe, fill_fields=False):
                def get_information_zone_de_gestion(scrollable_frame):
                    entree_invalide_liste = []
                    global information_champs
                    info_champs_temporaire = copy.deepcopy(information_champs)
                    index = 0
                    for scrollable_frame_widget in scrollable_frame.winfo_children():
                        if isinstance(scrollable_frame_widget, ttk.LabelFrame):
                            index_zone = 1
                            for champs_frame_widget in scrollable_frame_widget.winfo_children():
                                if isinstance(champs_frame_widget, ttk.LabelFrame):
                                    grid_slave0_1 = champs_frame_widget.grid_slaves(row=0, column=1)
                                    for entry in grid_slave0_1:
                                        taux_matiere_organique = entry.get()
                                        if not util.is_decimal_number(taux_matiere_organique) or float(
                                                taux_matiere_organique) < 0 or float(taux_matiere_organique) > 100:
                                            entree_invalide_liste.append((information_champs[index]["nom_du_champs"],
                                                                          "Zone gestion " + str(index_zone),
                                                                          "\"Taux de matière organique\" doit être un réel positif dans l'intervalle [0,100]"))
                                        else:
                                            taux_matiere_organique = float(taux_matiere_organique)
                                    grid_slave1_1 = champs_frame_widget.grid_slaves(row=1, column=1)
                                    for entry in grid_slave1_1:
                                        municipalite = entry.get()
                                        global municipalites_supportees
                                        if municipalite not in municipalites_supportees:
                                            entree_invalide_liste.append(
                                                (information_champs[index]["nom_du_champs"],
                                                 "Zone gestion " + str(index_zone),
                                                 "\"Municipalité\" doit être parmis les choix disponibles"))
                                    grid_slave2_1 = champs_frame_widget.grid_slaves(row=2, column=1)
                                    for entry in grid_slave2_1:
                                        groupe_textural = entry.get()
                                        global groupes_texturaux_supportees
                                        if groupe_textural not in groupes_texturaux_supportees:
                                            entree_invalide_liste.append(
                                                (information_champs[index]["nom_du_champs"],
                                                 "Zone gestion " + str(index_zone),
                                                 "\"Groupe textural\" doit être parmis les choix disponibles"))
                                    grid_slave3_1 = champs_frame_widget.grid_slaves(row=3, column=1)
                                    for entry in grid_slave3_1:
                                        classe_de_drainage = entry.get()
                                        global classes_de_drainage_supportees
                                        if classe_de_drainage not in classes_de_drainage_supportees:
                                            entree_invalide_liste.append(
                                                (information_champs[index]["nom_du_champs"],
                                                 "Zone gestion " + str(index_zone),
                                                 "\"Classe de drainage\" doit être parmis les choix disponibles"))
                                    grid_slave4_1 = champs_frame_widget.grid_slaves(row=4, column=1)
                                    for entry in grid_slave4_1:
                                        masse_volumique_apparente = entry.get()
                                        if (not util.is_decimal_number(
                                                masse_volumique_apparente) and masse_volumique_apparente != "") or (
                                                util.is_decimal_number(masse_volumique_apparente) and float(
                                            masse_volumique_apparente) < 0):
                                            entree_invalide_liste.append(
                                                (information_champs[index]["nom_du_champs"],
                                                 "Zone gestion " + str(
                                                     index_zone),
                                                 "\"Masse volumique apparente\" doit être un réel positif ou laissé vide pour aller chercher la valeur par défaut"))
                                        else:
                                            if masse_volumique_apparente == "":
                                                masse_volumique_apparente = None
                                            else:
                                                masse_volumique_apparente = float(masse_volumique_apparente)
                                    grid_slave5_1 = champs_frame_widget.grid_slaves(row=5, column=1)
                                    for entry in grid_slave5_1:
                                        profondeur = entry.get()
                                        if not util.is_decimal_number(profondeur) or float(
                                                profondeur) < 0:
                                            entree_invalide_liste.append((information_champs[index]["nom_du_champs"],
                                                                          "Zone gestion " + str(index_zone),
                                                                          "\"Profondeur de la couche arable\" doit être un réel positif"))
                                        else:
                                            profondeur = float(profondeur)
                                    grid_slave6_1 = champs_frame_widget.grid_slaves(row=6, column=1)
                                    for entry in grid_slave6_1:
                                        superficie_de_la_zone = entry.get()
                                        if not util.is_decimal_number(superficie_de_la_zone) or float(
                                                superficie_de_la_zone) < 0:
                                            entree_invalide_liste.append((information_champs[index]["nom_du_champs"],
                                                                          "Zone gestion " + str(index_zone),
                                                                          "\"Superficie de la zone\" doit être un réel positif"))
                                        else:
                                            superficie_de_la_zone = float(superficie_de_la_zone)
                                    information_champs[index]["information_zone_de_gestion"].append(
                                        {"taux_matiere_organique": taux_matiere_organique,
                                         "municipalite": municipalite,
                                         "groupe_textural": groupe_textural,
                                         "classe_de_drainage": classe_de_drainage,
                                         "masse_volumique_apparente": masse_volumique_apparente,
                                         "profondeur": profondeur,
                                         "superficie_de_la_zone": superficie_de_la_zone})
                                    index_zone += 1
                        index += 1
                    if len(entree_invalide_liste) == 0:
                        sauvegarde_reussi = sauvegarder_attributs_entreprise_apres_creation()
                        for widget in zone_de_gestion_mainframe.winfo_children():
                            widget.destroy()

                        if sauvegarde_reussi:
                            show_creation_des_regies(zone_de_gestion_mainframe)
                        else:
                            show_creation_zone_de_gestion(zone_de_gestion_mainframe, fill_fields=True)
                    else:
                        information_champs = info_champs_temporaire
                        message = ""
                        for entree_invalide in entree_invalide_liste:
                            message = message + "Dans le " + entree_invalide[0] + " et la " + entree_invalide[
                                1] + " l'entrée " + entree_invalide[2] + "\n"
                        messagebox.showwarning("Warning", message)

                canvas = tk.Canvas(zone_de_gestion_mainframe)
                scrollbar = ttk.Scrollbar(zone_de_gestion_mainframe, orient="vertical", command=canvas.yview)
                scrollable_frame = ttk.Frame(canvas)
                scrollable_frame.bind("<Configure>", lambda e: canvas.configure(scrollregion=canvas.bbox("all")))
                canvas.create_window((0, 0), window=scrollable_frame, anchor="nw")
                canvas.configure(yscrollcommand=scrollbar.set)

                def zone_frame_generation(zone_gestion_frame):
                    taux_matiere_organique_label = ttk.Label(zone_gestion_frame,
                                                             text="Taux matière organique (en %): ")
                    taux_matiere_organique_entry = ttk.Entry(zone_gestion_frame)
                    municipalite_label = ttk.Label(zone_de_gestion_frame, text="Municipalité: ")
                    global municipalites_supportees
                    municipalite_combobox = ttk.Combobox(zone_gestion_frame, values=municipalites_supportees,
                                                         postcommand=lambda: filter_combobox_values(
                                                             municipalite_combobox,
                                                             municipalites_supportees))
                    groupe_textural_label = ttk.Label(zone_de_gestion_frame, text="Groupe textural: ")
                    global groupes_texturaux_supportees
                    groupe_textural_combobox = ttk.Combobox(zone_gestion_frame, values=groupes_texturaux_supportees,
                                                            postcommand=lambda: filter_combobox_values(
                                                                groupe_textural_combobox,
                                                                groupes_texturaux_supportees))
                    classe_de_drainage_label = ttk.Label(zone_de_gestion_frame, text="Classe de drainage: ")
                    global classes_de_drainage_supportees
                    classe_de_drainage_combobox = ttk.Combobox(zone_gestion_frame,
                                                               values=classes_de_drainage_supportees,
                                                               postcommand=lambda: filter_combobox_values(
                                                                   classe_de_drainage_combobox,
                                                                   classes_de_drainage_supportees))
                    masse_volumique_apparente_label = ttk.Label(zone_gestion_frame,
                                                                text="Masse volumique apparente (g/cm3): ")
                    masse_volumique_apparente_entry = ttk.Entry(zone_gestion_frame)
                    profondeur_label = ttk.Label(zone_de_gestion_frame, text="Profondeur de la couche arable (cm): ")
                    profondeur_entry = ttk.Entry(zone_de_gestion_frame)
                    superficie_de_la_zone_label = ttk.Label(zone_de_gestion_frame, text="Superficie de la zone (ha): ")
                    superficie_de_la_zone_entry = ttk.Entry(zone_de_gestion_frame)

                    if fill_fields:
                        taux_matiere_organique_entry.insert(0,
                                                            information_champs[index_champs][
                                                                "information_zone_de_gestion"][
                                                                index_zone_de_gestion]["taux_matiere_organique"])
                        municipalite_combobox.insert(0, information_champs[index_champs]["information_zone_de_gestion"][
                            index_zone_de_gestion]["municipalite"])
                        groupe_textural_combobox.insert(0,
                                                        information_champs[index_champs]["information_zone_de_gestion"][
                                                            index_zone_de_gestion]["groupe_textural"])
                        classe_de_drainage_combobox.insert(0,
                                                           information_champs[index_champs][
                                                               "information_zone_de_gestion"][
                                                               index_zone_de_gestion]["classe_de_drainage"])
                        masse_volumique_apparente_entry.insert(0, information_champs[index_champs][
                            "information_zone_de_gestion"][index_zone_de_gestion]["masse_volumique_apparente"])
                        profondeur_entry.insert(0, information_champs[index_champs]["information_zone_de_gestion"][
                            index_zone_de_gestion]["profondeur"])
                        superficie_de_la_zone_entry.insert(0,
                                                           information_champs[index_champs][
                                                               "information_zone_de_gestion"][
                                                               index_zone_de_gestion]["superficie_de_la_zone"])

                    taux_matiere_organique_label.grid(row=0, column=0, sticky="w", pady=3, padx=5)
                    taux_matiere_organique_entry.grid(row=0, column=1, sticky="w", pady=3, padx=5)
                    municipalite_label.grid(row=1, column=0, sticky="w", pady=3, padx=5)
                    municipalite_combobox.grid(row=1, column=1, sticky="w", pady=3, padx=5)
                    groupe_textural_label.grid(row=2, column=0, sticky="w", pady=3, padx=5)
                    groupe_textural_combobox.grid(row=2, column=1, sticky="w", pady=3, padx=5)
                    classe_de_drainage_label.grid(row=3, column=0, sticky="w", pady=3, padx=5)
                    classe_de_drainage_combobox.grid(row=3, column=1, sticky="w", pady=3, padx=5)
                    masse_volumique_apparente_label.grid(row=4, column=0, sticky="w", pady=3, padx=5)
                    masse_volumique_apparente_entry.grid(row=4, column=1, sticky="w", pady=3, padx=5)
                    profondeur_label.grid(row=5, column=0, sticky="w", pady=3, padx=5)
                    profondeur_entry.grid(row=5, column=1, sticky="w", pady=3, padx=5)
                    superficie_de_la_zone_label.grid(row=6, column=0, sticky="w", pady=3, padx=5)
                    superficie_de_la_zone_entry.grid(row=6, column=1, sticky="w", pady=3, padx=5)

                    zone_de_gestion_frame.pack()

                for index_champs in range(int(nombre_de_champs)):
                    global information_champs
                    champs_frame = ttk.LabelFrame(scrollable_frame,
                                                  text=information_champs[index_champs]["nom_du_champs"],
                                                  name="champ" + str(index_champs))
                    for index_zone_de_gestion in range(
                            int(information_champs[index_champs]["nombre_de_zone_de_gestion"])):
                        zone_de_gestion_frame = ttk.LabelFrame(champs_frame,
                                                               text="Zone  gestion " + str(index_zone_de_gestion + 1),
                                                               name="zonegestion" + str(index_zone_de_gestion))
                        zone_frame_generation(zone_de_gestion_frame)
                    champs_frame.pack(fill="both", padx=10, pady=5, ipadx=10, ipady=5)

                canvas.pack(side="left", ipadx=10)
                scrollbar.pack(side="right", fill="y")

                creation_zone_de_gestion_bouton = ttk.Button(scrollable_frame, text="Créer",
                                                             command=lambda: get_information_zone_de_gestion(
                                                                 scrollable_frame))
                creation_zone_de_gestion_bouton.pack()

            def show_creation_des_regies(parent_frame_tabs, simulations=None):
                def add_new_simulation_tab(event):
                    clicked_tab = simulation_notebook.tk.call(simulation_notebook._w, "identify", "tab", event.x,
                                                              event.y)
                    if clicked_tab == simulation_notebook.index("end") - 1:
                        index_clicked_tab = simulation_notebook.index(clicked_tab)
                        simulation_notebook.winfo_children()[index_clicked_tab].destroy()
                        global nombre_simulations
                        nombre_simulations += 1

                        def duree_de_la_simulation():
                            def readd_simulation():
                                root.deiconify()
                                tab = ttk.Frame(simulation_notebook)
                                simulation_notebook.add(tab, text="+")
                                global nombre_simulations
                                nombre_simulations -= 1
                                duree_simulation_window.destroy()

                            def get_duree_de_la_simulation():
                                root.deiconify()
                                nom_simulation = nom_simulation_entry.get()
                                annee_projection_initiale = annee_projection_initiale_entry.get()
                                duree_projection = duree_projection_entry.get()
                                numero_simulation_copie = numero_simulation_copie_entry.get()
                                if numero_simulation_copie == "":
                                    numero_simulation_copie = None
                                global nombre_simulations
                                annee_projection_initiale_valide = True
                                if annee_projection_initiale.isdigit() and duree_projection.isdigit() \
                                        and int(duree_projection) > 0 \
                                        and annee_projection_initiale_valide and \
                                        (numero_simulation_copie is None or (numero_simulation_copie.isdigit() and
                                                                             (nombre_simulations > int(
                                                                                 numero_simulation_copie) > 0))) and \
                                        len(nom_simulation) < 13:
                                    global duree_simulation
                                    duree_simulation.append({"annee_projection_initiale": annee_projection_initiale,
                                                             "duree_projection": duree_projection,
                                                             "nom_simulation": nom_simulation})
                                    if numero_simulation_copie is not None:
                                        simulation = simulation_notebook.winfo_children()[
                                            int(numero_simulation_copie) - 1]
                                        simulation_copie = get_information_simulation(simulation,
                                                                                      int(numero_simulation_copie) - 1,
                                                                                      simulation_unique=True)
                                        if simulation_copie[0] is not None:
                                            simulation_copie = simulation_copie[0]
                                            duree_simulation_window.destroy()
                                            set_up_simulation(simulation_notebook, simulation_copie, False)
                                        else:
                                            message = ""
                                            for entree_invalide in simulation_copie[1]:
                                                message = message + "Dans la " + entree_invalide[3] + ", le " + \
                                                          entree_invalide[
                                                              0] + " et la " + \
                                                          entree_invalide[
                                                              1] + " l'entrée " + entree_invalide[2] + "\n"
                                            messagebox.showwarning("Warning", message)
                                            readd_simulation()
                                    else:
                                        simulation_copie = None
                                        duree_simulation_window.destroy()
                                        if len(simulation_notebook.winfo_children()) == 0:
                                            set_up_simulation(simulation_notebook, simulation_copie, True)
                                        else:
                                            set_up_simulation(simulation_notebook, simulation_copie, False)
                                else:
                                    entree_invalide_liste = []
                                    message = ""
                                    if not annee_projection_initiale.isdigit() or not annee_projection_initiale_valide:
                                        entree_invalide_liste.append(
                                            "L'entrée \"Année de projection initiale\" est invalide.")
                                    if not duree_projection.isdigit() or int(duree_projection) <= 0:
                                        entree_invalide_liste.append(
                                            "L'entrée \"Durée de la projection\" est invalide. Elle doit être un nombre naturel plus grand que 0.")
                                    if (
                                            numero_simulation_copie is not None and not numero_simulation_copie.isdigit()) or (
                                            numero_simulation_copie is not None and (int(
                                        numero_simulation_copie) >= nombre_simulations or int(
                                        numero_simulation_copie) < 1)):
                                        entree_invalide_liste.append(
                                            "L'entrée \"Numéro de la simulation à copier\" est invalide. Elle doit être un nombre naturel plus grand que 0 et parmis les numéros de simulations existantes.")
                                    if len(nom_simulation) > 12:
                                        entree_invalide_liste.append(
                                            "Le nom de la simulation doit être composé de 12 caractères ou moins.")

                                    for entree_invalide in entree_invalide_liste:
                                        message = message + entree_invalide
                                    messagebox.showwarning("Warning", message)
                                    readd_simulation()

                            root.withdraw()
                            duree_simulation_window = tk.Toplevel()
                            duree_simulation_window.resizable(0, 0)
                            duree_simulation_window.focus()
                            duree_simulation_window.protocol("WM_DELETE_WINDOW", readd_simulation)
                            duree_simulation_frame = ttk.Frame(duree_simulation_window)
                            nom_simulation_label = ttk.Label(duree_simulation_frame, text="Nom simulation: ")
                            nom_simulation_entry = ttk.Entry(duree_simulation_frame)
                            annee_projection_initiale_label = ttk.Label(duree_simulation_frame,
                                                                        text="Année de projection initiale: ")
                            annee_projection_initiale_entry = ttk.Entry(duree_simulation_frame)
                            duree_projection_label = ttk.Label(duree_simulation_frame,
                                                               text="Durée de la projection: ")
                            duree_projection_entry = ttk.Entry(duree_simulation_frame)
                            duree_projection_entry.insert(0, "30")

                            numero_simulation_copie_label = ttk.Label(duree_simulation_frame,
                                                                      text="Numéro de la simulation à copier: ")
                            numero_simulation_copie_entry = ttk.Entry(duree_simulation_frame)

                            global nombre_simulations
                            if nombre_simulations < 2:
                                numero_simulation_copie_entry.configure(state="disabled")

                            nom_simulation_label.grid(row=0, column=0, sticky="w", pady=3, padx=5)
                            nom_simulation_entry.grid(row=0, column=1, sticky="w", pady=3, padx=5)
                            annee_projection_initiale_label.grid(row=1, column=0, sticky="w", pady=3, padx=5)
                            annee_projection_initiale_entry.grid(row=1, column=1, sticky="w", pady=3, padx=5)
                            duree_projection_label.grid(row=2, column=0, sticky="w", pady=3, padx=5)
                            duree_projection_entry.grid(row=2, column=1, sticky="w", pady=3, padx=5)
                            numero_simulation_copie_label.grid(row=3, column=0, sticky="w", pady=3, padx=5)
                            numero_simulation_copie_entry.grid(row=3, column=1, sticky="w", pady=3, padx=5)

                            creer_simulation_bouton = ttk.Button(duree_simulation_frame, text="Créer",
                                                                 command=get_duree_de_la_simulation)
                            creer_simulation_bouton.grid(row=5, column=0, columnspan=2, pady=3)
                            duree_simulation_frame.pack()

                        duree_de_la_simulation()

                def delete_simulation_tab(event):
                    clicked_tab = simulation_notebook.tk.call(simulation_notebook._w, "identify", "tab", event.x,
                                                              event.y)
                    index_clicked_tab = simulation_notebook.index(clicked_tab)
                    if index_clicked_tab != simulation_notebook.index("end") - 1:
                        simulation_notebook.winfo_children()[index_clicked_tab].destroy()
                        global max_index_simulation
                        global furthest_left_tab_index_simulation
                        global furthest_right_tab_index_simulation
                        if furthest_left_tab_index_simulation == min_index_simulation and len(
                                simulation_notebook.winfo_children()) > 5:
                            simulation_notebook.tab(furthest_right_tab_index_simulation, state="normal")
                        elif furthest_left_tab_index_simulation == min_index_simulation and len(
                                simulation_notebook.winfo_children()) == 5:
                            for tab_index in range(len(simulation_notebook.winfo_children())):
                                simulation_notebook.tab(tab_index, state="normal")
                            scroll_right_button_simulation.configure(state="disabled")
                        elif furthest_right_tab_index_simulation == max_index_simulation and len(
                                simulation_notebook.winfo_children()) == 5:
                            furthest_right_tab_index_simulation -= 1
                            furthest_left_tab_index_simulation -= 1
                            simulation_notebook.tab(furthest_left_tab_index_simulation, state="normal")
                            scroll_left_button_simulation.configure(state="disabled")
                        elif furthest_left_tab_index_simulation == min_index_simulation and len(
                                simulation_notebook.winfo_children()) <= 4:
                            furthest_right_tab_index_simulation -= 1
                        elif furthest_right_tab_index_simulation == max_index_simulation and len(
                                simulation_notebook.winfo_children()) > 5:
                            furthest_left_tab_index_simulation -= 1
                            furthest_right_tab_index_simulation -= 1
                            simulation_notebook.tab(furthest_left_tab_index_simulation, state="normal")
                        else:
                            simulation_notebook.tab(furthest_right_tab_index_simulation, state="normal")
                            if max_index_simulation - 1 == furthest_right_tab_index_simulation:
                                scroll_right_button_simulation.configure(state="disabled")
                        global nombre_simulations
                        nombre_simulations -= 1
                        global duree_simulation
                        duree_simulation.pop(index_clicked_tab)
                        max_index_simulation -= 1

                def set_up_simulation(simulation_notebook, simulation_copie, simulation_en_tete_de_liste):
                    global nombre_simulations
                    global furthest_left_tab_index_simulation
                    global furthest_right_tab_index_simulation
                    global max_index_simulation
                    global duree_simulation
                    global information_champs
                    global nombre_de_champs
                    global simulations_chargees
                    global min_index_champs
                    if (4 < nombre_simulations and not simulations_chargees) or (
                            simulations_chargees and len(simulation_notebook.winfo_children()) > 3):
                        simulation_notebook.tab(furthest_left_tab_index_simulation, state="hidden")
                        scroll_left_button_simulation.configure(state="normal")
                        furthest_left_tab_index_simulation += 1
                    furthest_right_tab_index_simulation += 1
                    max_index_simulation += 1
                    tab = ttk.Frame(simulation_notebook)

                    if simulations_chargees:
                        simulation_notebook.add(tab, text=simulation_copie["nom_simulation"])
                    else:
                        simulation_notebook.add(tab, text=duree_simulation[nombre_simulations - 1]["nom_simulation"])
                    new_tab = ttk.Frame(simulation_notebook)
                    simulation_notebook.add(new_tab, text="+")
                    champs_notebook = ttk.Notebook(tab)
                    champs_notebook.bind("<<NotebookTabChanged>>", zone_tab_management)
                    simulation_notebook.bind("<<NotebookTabChanged>>", zone_tab_management)

                    def add_new_champs_tab(event):
                        clicked_tab = champs_notebook.tk.call(champs_notebook._w, "identify", "tab", event.x, event.y)
                        if clicked_tab == champs_notebook.index("end") - 1:
                            global nombre_de_champs
                            set_up_new_champs()

                    def delete_champs_tab(event):
                        clicked_tab = champs_notebook.tk.call(champs_notebook._w, "identify", "tab", event.x, event.y)
                        index_clicked_tab = champs_notebook.index(clicked_tab)
                        if index_clicked_tab != champs_notebook.index("end") - 1:
                            global information_champs
                            information_champs.pop(index_clicked_tab)
                            global nombre_de_champs
                            nombre_de_champs -= 1
                            sauvegarder_attributs_entreprise_apres_modification()
                            donnees_de_rechauffement_label_frame.winfo_children()[0].winfo_children()[
                                0].winfo_children()[
                                index_clicked_tab].destroy()
                            champs_index = index_clicked_tab
                            while champs_index < len(
                                    donnees_de_rechauffement_label_frame.winfo_children()[0].winfo_children()[
                                        0].winfo_children()):
                                donnees_de_rechauffement_label_frame.winfo_children()[0].winfo_children()[
                                    0].winfo_children()[
                                    champs_index].configure(text=information_champs[champs_index]["nom_du_champs"])
                                champs_index += 1
                            for simulation_frame in simulation_notebook.winfo_children():
                                if len(simulation_frame.winfo_children()) == 0:
                                    pass
                                else:
                                    notebook = simulation_frame.winfo_children()[0]
                                    notebook.winfo_children()[index_clicked_tab].destroy()
                                    global furthest_left_tab_index_champs
                                    global furthest_right_tab_index_champs
                                    global max_index_champs
                                    if furthest_left_tab_index_champs == min_index_champs and len(
                                            notebook.winfo_children()) > 5:
                                        notebook.tab(furthest_right_tab_index_champs, state="normal")
                                    elif furthest_left_tab_index_champs == min_index_champs and len(
                                            notebook.winfo_children()) == 5:
                                        for tab_index in range(len(notebook.winfo_children())):
                                            notebook.tab(tab_index, state="normal")
                                    elif furthest_right_tab_index_champs == max_index_champs and len(
                                            notebook.winfo_children()) == 5:
                                        notebook.tab(furthest_left_tab_index_champs - 1, state="normal")
                                    elif furthest_left_tab_index_champs == min_index_champs and len(
                                            notebook.winfo_children()) <= 4:
                                        pass
                                    elif furthest_right_tab_index_champs == max_index_champs and len(
                                            notebook.winfo_children()) > 5:
                                        notebook.tab(furthest_left_tab_index_champs, state="normal")
                                    else:
                                        notebook.tab(furthest_right_tab_index_champs, state="normal")

                            notebook = simulation_notebook.winfo_children()[0].winfo_children()[0]
                            if furthest_left_tab_index_champs == min_index_champs and len(
                                    notebook.winfo_children()) > 5:
                                pass
                            elif furthest_left_tab_index_champs == min_index_champs and len(
                                    notebook.winfo_children()) == 5:
                                scroll_right_button_champs.configure(state="disabled")
                            elif furthest_right_tab_index_champs == max_index_champs and len(
                                    notebook.winfo_children()) == 5:
                                furthest_left_tab_index_champs -= 1
                                furthest_right_tab_index_champs -= 1
                                scroll_left_button_champs.configure(state="disabled")
                            elif furthest_left_tab_index_champs == min_index_champs and len(
                                    notebook.winfo_children()) <= 4:
                                furthest_right_tab_index_champs -= 1
                            elif furthest_right_tab_index_champs == max_index_champs and len(
                                    notebook.winfo_children()) > 5:
                                furthest_left_tab_index_champs -= 1
                                furthest_right_tab_index_champs -= 1
                            else:
                                if max_index_champs - 1 == furthest_right_tab_index_champs:
                                    scroll_right_button_champs.configure(state="disabled")
                            max_index_champs -= 1

                    champs_notebook.bind("<Button-1>", add_new_champs_tab)
                    champs_notebook.bind("<Button-3>", delete_champs_tab)

                    global max_index_champs
                    global furthest_right_tab_index_champs
                    global furthest_left_tab_index_champs
                    index_champs = 0
                    for champs in information_champs:
                        tab = ttk.Frame(champs_notebook, name="champ" + str(len(champs_notebook.winfo_children())))
                        if simulation_en_tete_de_liste:
                            max_index_champs += 1
                            furthest_right_tab_index_champs += 1
                            if index_champs > 3:
                                champs_notebook.tab(furthest_left_tab_index_champs, state="hidden")
                                furthest_left_tab_index_champs += 1
                        champs_notebook.add(tab, text=champs["nom_du_champs"])
                        if not simulation_en_tete_de_liste and (
                                index_champs < furthest_left_tab_index_champs or index_champs > furthest_right_tab_index_champs):
                            champs_notebook.tab(index_champs, state="hidden")
                        zone_de_gestion_notebook = ttk.Notebook(tab)
                        set_up_champs(zone_de_gestion_notebook, int(champs["nombre_de_zone_de_gestion"]),
                                      champs_notebook,
                                      simulation_copie, index_champs)
                        index_champs += 1

                    global nombre_de_champs
                    if nombre_de_champs > 4:
                        scroll_left_button_champs.configure(state="normal")

                    tab = ttk.Frame(champs_notebook)
                    if index_champs > furthest_right_tab_index_champs:
                        champs_notebook.add(tab, text="+", state="hidden")
                    else:
                        champs_notebook.add(tab, text="+")

                    champs_notebook.pack(anchor="nw")

                def set_up_new_champs():
                    new_champs_window = tk.Toplevel()
                    new_champs_window.resizable(0, 0)
                    global nombre_de_champs_modifie
                    nombre_de_champs_modifie = False
                    global information_champs_modifie
                    information_champs_modifie = False

                    def revenir_a_etat_initial():
                        global nombre_de_champs_modifie
                        if nombre_de_champs_modifie:
                            global nombre_de_champs
                            nombre_de_champs -= 1
                        global information_champs_modifie
                        if information_champs_modifie:
                            global information_champs
                            information_champs.pop(len(information_champs) - 1)
                        new_champs_window.destroy()

                    new_champs_window.protocol("WM_DELETE_WINDOW", revenir_a_etat_initial)

                    def creation_des_zone_de_gestion_du_nouveau_champs():
                        nom_du_champs = nom_du_champs_entry.get()
                        nombre_de_zone_de_gestion = nombre_de_zone_de_gestion_entry.get()

                        def zone_frame_generation(zone_gestion_frame):
                            taux_matiere_organique_label = ttk.Label(zone_gestion_frame,
                                                                     text="Taux matière organique (en %): ")
                            taux_matiere_organique_entry = ttk.Entry(zone_gestion_frame)
                            municipalite_label = ttk.Label(zone_de_gestion_frame, text="Municipalité: ")
                            global municipalites_supportees
                            municipalite_combobox = ttk.Combobox(zone_gestion_frame, values=municipalites_supportees,
                                                                 postcommand=lambda: filter_combobox_values(
                                                                     municipalite_combobox,
                                                                     municipalites_supportees))
                            groupe_textural_label = ttk.Label(zone_de_gestion_frame, text="Groupe textural: ")
                            global groupes_texturaux_supportees
                            groupe_textural_combobox = ttk.Combobox(zone_gestion_frame,
                                                                    values=groupes_texturaux_supportees,
                                                                    postcommand=lambda: filter_combobox_values(
                                                                        groupe_textural_combobox,
                                                                        groupes_texturaux_supportees))
                            classe_de_drainage_label = ttk.Label(zone_de_gestion_frame, text="Classe de drainage: ")
                            global classes_de_drainage_supportees
                            classe_de_drainage_combobox = ttk.Combobox(zone_gestion_frame,
                                                                       values=classes_de_drainage_supportees,
                                                                       postcommand=lambda: filter_combobox_values(
                                                                           classe_de_drainage_combobox,
                                                                           classes_de_drainage_supportees))
                            masse_volumique_apparente_label = ttk.Label(zone_gestion_frame,
                                                                        text="Masse volumique apparente (g/cm3): ")
                            masse_volumique_apparente_entry = ttk.Entry(zone_gestion_frame)
                            profondeur_label = ttk.Label(zone_de_gestion_frame,
                                                         text="Profondeur de la couche arable(cm): ")
                            profondeur_entry = ttk.Entry(zone_de_gestion_frame)
                            superficie_de_la_zone_label = ttk.Label(zone_de_gestion_frame,
                                                                    text="Superficie de la zone (ha): ")
                            superficie_de_la_zone_entry = ttk.Entry(zone_de_gestion_frame)

                            taux_matiere_organique_label.grid(row=0, column=0, sticky="w", pady=3, padx=5)
                            taux_matiere_organique_entry.grid(row=0, column=1, sticky="w", pady=3, padx=5)
                            municipalite_label.grid(row=1, column=0, sticky="w", pady=3, padx=5)
                            municipalite_combobox.grid(row=1, column=1, sticky="w", pady=3, padx=5)
                            groupe_textural_label.grid(row=2, column=0, sticky="w", pady=3, padx=5)
                            groupe_textural_combobox.grid(row=2, column=1, sticky="w", pady=3, padx=5)
                            classe_de_drainage_label.grid(row=3, column=0, sticky="w", pady=3, padx=5)
                            classe_de_drainage_combobox.grid(row=3, column=1, sticky="w", pady=3, padx=5)
                            masse_volumique_apparente_label.grid(row=4, column=0, sticky="w", pady=3, padx=5)
                            masse_volumique_apparente_entry.grid(row=4, column=1, sticky="w", pady=3, padx=5)
                            profondeur_label.grid(row=5, column=0, sticky="w", pady=3, padx=5)
                            profondeur_entry.grid(row=5, column=1, sticky="w", pady=3, padx=5)
                            superficie_de_la_zone_label.grid(row=6, column=0, sticky="w", pady=3, padx=5)
                            superficie_de_la_zone_entry.grid(row=6, column=1, sticky="w", pady=3, padx=5)

                            zone_de_gestion_frame.pack()

                        if nombre_de_zone_de_gestion.isdigit() and int(nombre_de_zone_de_gestion) > 0:
                            if len(nom_du_champs) < 13:
                                global information_champs
                                information_champs.append({"nom_du_champs": nom_du_champs,
                                                           "nombre_de_zone_de_gestion": nombre_de_zone_de_gestion,
                                                           "information_zone_de_gestion": []})
                                global nombre_de_champs
                                nombre_de_champs += 1
                                global nombre_de_champs_modifie
                                nombre_de_champs_modifie = True
                                global information_champs_modifie
                                information_champs_modifie = True

                                for widget in new_champs_window.winfo_children():
                                    widget.destroy()
                                creation_zone_frame = ttk.Frame(new_champs_window)
                                canvas = tk.Canvas(creation_zone_frame)
                                scrollbar = ttk.Scrollbar(creation_zone_frame, orient="vertical", command=canvas.yview)
                                scrollable_frame = ttk.Frame(canvas)
                                scrollable_frame.bind("<Configure>",
                                                      lambda e: canvas.configure(scrollregion=canvas.bbox("all")))
                                canvas.create_window((0, 0), window=scrollable_frame, anchor="nw")
                                canvas.configure(yscrollcommand=scrollbar.set)
                                for index_zone_de_gestion_nouveau_champs in range(
                                        int(nombre_de_zone_de_gestion)):
                                    zone_de_gestion_frame = ttk.LabelFrame(scrollable_frame,
                                                                           text="Zone gestion " + str(
                                                                               index_zone_de_gestion_nouveau_champs + 1),
                                                                           name="zonegestion" + str(
                                                                               index_zone_de_gestion_nouveau_champs))
                                    zone_frame_generation(zone_de_gestion_frame)

                                creation_zone_de_gestion_bouton = ttk.Button(scrollable_frame, text="Créer",
                                                                             command=lambda: add_new_zone_tab())
                                creation_zone_de_gestion_bouton.pack()
                                canvas.pack(side="left", fill="x", expand=True)
                                scrollbar.pack(side="right", fill="y")
                                creation_zone_frame.pack(fill=None, expand=False)
                            else:
                                messagebox.showwarning("Warning",
                                                       "Le nom du champ devrait être composé de 12 caractères ou moins.")
                                new_champs_window.focus()
                        else:
                            message = "L'entrée \"Nombre de zones gestion\" est invalide. Elle doit être un nombre naturel plus grand que 0."
                            messagebox.showwarning("Warning", message)
                            new_champs_window.focus()

                        def add_new_zone_tab():
                            entree_invalide_liste = []
                            global nomde_de_champs
                            global information_champs
                            index_zone = 0
                            for scrollable_frame_widget in scrollable_frame.winfo_children():
                                if isinstance(scrollable_frame_widget, ttk.LabelFrame):
                                    grid_slave0_1 = scrollable_frame_widget.grid_slaves(row=0, column=1)
                                    for entry in grid_slave0_1:
                                        taux_matiere_organique = entry.get()
                                        if not util.is_decimal_number(taux_matiere_organique) or float(
                                                taux_matiere_organique) < 0 or float(taux_matiere_organique) > 100:
                                            entree_invalide_liste.append(
                                                (information_champs[nombre_de_champs - 1]["nom_du_champs"],
                                                 "Zone gestion " + str(index_zone + 1),
                                                 "\"Taux de matière organique\" doit être un réel positif dans l'intervalle [0,100]"))
                                        else:
                                            taux_matiere_organique = float(taux_matiere_organique)
                                    grid_slave1_1 = scrollable_frame_widget.grid_slaves(row=1, column=1)
                                    for entry in grid_slave1_1:
                                        municipalite = entry.get()
                                        global municipalites_supportees
                                        if municipalite not in municipalites_supportees:
                                            entree_invalide_liste.append(
                                                (information_champs[nombre_de_champs - 1]["nom_du_champs"],
                                                 "Zone gestion " + str(index_zone + 1),
                                                 "\"Municipalité\" doit être parmis les choix disponibles"))
                                    grid_slave2_1 = scrollable_frame_widget.grid_slaves(row=2, column=1)
                                    for entry in grid_slave2_1:
                                        groupe_textural = entry.get()
                                        global groupes_texturaux_supportees
                                        if groupe_textural not in groupes_texturaux_supportees:
                                            entree_invalide_liste.append(
                                                (information_champs[nombre_de_champs - 1]["nom_du_champs"],
                                                 "Zone gestion " + str(index_zone + 1),
                                                 "\"Groupe textural\" doit être parmis les choix disponibles"))
                                    grid_slave3_1 = scrollable_frame_widget.grid_slaves(row=3, column=1)
                                    for entry in grid_slave3_1:
                                        classe_de_drainage = entry.get()
                                        global classes_de_drainage_supportees
                                        if classe_de_drainage not in classes_de_drainage_supportees:
                                            entree_invalide_liste.append(
                                                (information_champs[nombre_de_champs - 1]["nom_du_champs"],
                                                 "Zone gestion " + str(index_zone + 1),
                                                 "\"Classe de drainage\" doit être parmis les choix disponibles"))
                                    grid_slave4_1 = scrollable_frame_widget.grid_slaves(row=4, column=1)
                                    for entry in grid_slave4_1:
                                        masse_volumique_apparente = entry.get()
                                        if (not util.is_decimal_number(
                                                masse_volumique_apparente) and masse_volumique_apparente != "") or (
                                                util.is_decimal_number(masse_volumique_apparente) and float(
                                            masse_volumique_apparente) < 0):
                                            entree_invalide_liste.append(
                                                (information_champs[nombre_de_champs - 1]["nom_du_champs"],
                                                 "Zone gestion " + str(
                                                     index_zone + 1),
                                                 "\"Masse volumique apparente\" doit être un réel positif ou laissé vide pour aller chercher la valeur par défaut"))
                                        else:
                                            if masse_volumique_apparente == "":
                                                masse_volumique_apparente = None
                                            else:
                                                masse_volumique_apparente = float(masse_volumique_apparente)
                                    grid_slave5_1 = scrollable_frame_widget.grid_slaves(row=5, column=1)
                                    for entry in grid_slave5_1:
                                        profondeur = entry.get()
                                        if not util.is_decimal_number(profondeur) or float(
                                                profondeur) < 0:
                                            entree_invalide_liste.append(
                                                (information_champs[nombre_de_champs - 1]["nom_du_champs"],
                                                 "Zone gestion " + str(index_zone + 1),
                                                 "\"Profondeur de la couche arable\" doit être un réel positif"))
                                        else:
                                            profondeur = float(profondeur)
                                    grid_slave6_1 = scrollable_frame_widget.grid_slaves(row=6, column=1)
                                    for entry in grid_slave6_1:
                                        superficie_de_la_zone = entry.get()
                                        if not util.is_decimal_number(superficie_de_la_zone) or float(
                                                superficie_de_la_zone) < 0:
                                            entree_invalide_liste.append(
                                                (information_champs[nombre_de_champs - 1]["nom_du_champs"],
                                                 "Zone gestion " + str(index_zone + 1),
                                                 "\"Superficie de la zone\" doit être un réel positif"))
                                        else:
                                            superficie_de_la_zone = float(superficie_de_la_zone)
                                    information_champs[len(information_champs) - 1][
                                        "information_zone_de_gestion"].append(
                                        {"taux_matiere_organique": taux_matiere_organique,
                                         "municipalite": municipalite,
                                         "groupe_textural": groupe_textural,
                                         "classe_de_drainage": classe_de_drainage,
                                         "masse_volumique_apparente": masse_volumique_apparente,
                                         "profondeur": profondeur,
                                         "superficie_de_la_zone": superficie_de_la_zone})
                                index_zone += 1
                            if len(entree_invalide_liste) == 0:
                                sauvegarder_attributs_entreprise_apres_modification()
                                new_champs_window.destroy()
                                global furthest_left_tab_index_champs
                                global max_index_champs
                                global furthest_right_tab_index_champs
                                index_champ = ""
                                for simulation_frame in simulation_notebook.winfo_children():
                                    if len(simulation_frame.winfo_children()) == 0:
                                        pass
                                    else:
                                        notebook = simulation_frame.winfo_children()[0]
                                        for champs_frame in notebook.winfo_children():
                                            if len(champs_frame.winfo_children()) == 0:
                                                champs_frame.destroy()
                                            else:
                                                pass
                                        index_champ = str(len(notebook.winfo_children()))
                                        tab = ttk.Frame(notebook, name="champ" + str(len(notebook.winfo_children())))
                                        notebook.add(tab, text=nom_du_champs)
                                        zone_notebook = ttk.Notebook(tab)
                                        set_up_champs(zone_notebook, nombre_de_zone_de_gestion, notebook)
                                        new_tab = ttk.Frame(notebook)
                                        notebook.add(new_tab, text="+")
                                        if len(notebook.winfo_children()) > 5:
                                            notebook.tab(furthest_left_tab_index_champs, state="hidden")

                                max_index_champs += 1
                                furthest_right_tab_index_champs += 1
                                if len(simulation_notebook.winfo_children()[0].winfo_children()[
                                           0].winfo_children()) > 5:
                                    furthest_left_tab_index_champs += 1
                                    scroll_left_button_champs.configure(state="normal")
                                rechauffement_champs_label_frame = ttk.LabelFrame(
                                    donnees_de_rechauffement_label_frame.winfo_children()[0].winfo_children()[0],
                                    text=nom_du_champs, name="champ" + index_champ)
                                rechauffement_champs_label_frame.pack()
                                index_zone = 0
                                while index_zone < int(nombre_de_zone_de_gestion):
                                    rechauffement_zone_label_frame = ttk.LabelFrame(rechauffement_champs_label_frame,
                                                                                    text="Zone gestion " + str(
                                                                                        index_zone + 1),
                                                                                    name="zonegestion" + str(
                                                                                        index_zone))
                                    rechauffement_zone_label_frame.pack()
                                    ajouter_une_annee_a_la_rotation(rechauffement_zone_label_frame)
                                    index_zone += 1
                            else:
                                information_champs[nombre_de_champs - 1]["information_zone_de_gestion"] = []
                                message = ""
                                for entree_invalide in entree_invalide_liste:
                                    message = message + "Dans le " + entree_invalide[0] + " et la " + entree_invalide[
                                        1] + " l'entrée " + entree_invalide[2] + "\n"
                                messagebox.showwarning("Warning", message)
                                new_champs_window.focus()

                    nouveau_champs_frame = ttk.Frame(new_champs_window)
                    nom_du_champs_label = ttk.Label(nouveau_champs_frame, text="Nom du champ: ")
                    nom_du_champs_entry = ttk.Entry(nouveau_champs_frame)
                    nombre_de_zone_de_gestion_label = ttk.Label(nouveau_champs_frame,
                                                                text="Nombre de zones gestion: ")
                    nombre_de_zone_de_gestion_entry = ttk.Entry(nouveau_champs_frame)
                    nom_du_champs_label.grid(row=0, column=0, sticky="w", pady=3, padx=5)
                    nom_du_champs_entry.grid(row=0, column=1, sticky="w", pady=3, padx=5)
                    nombre_de_zone_de_gestion_label.grid(row=1, sticky="w", column=0, pady=3, padx=5)
                    nombre_de_zone_de_gestion_entry.grid(row=1, sticky="w", column=1, pady=3, padx=5)
                    creer_nouveau_champs_bouton = ttk.Button(nouveau_champs_frame, text="Créer",
                                                             command=creation_des_zone_de_gestion_du_nouveau_champs)
                    creer_nouveau_champs_bouton.grid(row=3, column=0, columnspan=2, pady=3, padx=5)
                    nouveau_champs_frame.pack()

                def set_up_champs(zone_notebook, nombre_de_zone, champs_notebook, simulation_copie=None,
                                  index_champs=None):

                    def add_new_zone_de_gestion_tab(event):
                        clicked_tab = zone_notebook.tk.call(zone_notebook._w, "identify", "tab", event.x, event.y)
                        if clicked_tab == zone_notebook.index("end") - 1:
                            champs_index = champs_notebook.index("current")
                            global information_champs
                            information_champs[champs_index]["nombre_de_zone_de_gestion"] = str(
                                int(information_champs[champs_index]["nombre_de_zone_de_gestion"]) + 1)
                            set_up_new_zone_de_gestion(champs_index)

                    def delete_zone_de_gestion_tab(event):
                        clicked_tab = zone_notebook.tk.call(zone_notebook._w, "identify", "tab", event.x, event.y)
                        index_clicked_tab = zone_notebook.index(clicked_tab)
                        if index_clicked_tab != zone_notebook.index("end") - 1:
                            champs_index = champs_notebook.index("current")
                            global information_champs
                            information_champs[champs_index]["nombre_de_zone_de_gestion"] = str(
                                int(information_champs[champs_index]["nombre_de_zone_de_gestion"]) - 1)
                            information_champs[champs_index]["information_zone_de_gestion"].pop(index_clicked_tab)
                            sauvegarder_attributs_entreprise_apres_modification()
                            donnees_de_rechauffement_label_frame.winfo_children()[0].winfo_children()[
                                0].winfo_children()[
                                champs_index].winfo_children()[index_clicked_tab].destroy()
                            zone_index = index_clicked_tab
                            while zone_index < len(
                                    donnees_de_rechauffement_label_frame.winfo_children()[0].winfo_children()[
                                        0].winfo_children()[
                                        champs_index].winfo_children()):
                                donnees_de_rechauffement_label_frame.winfo_children()[0].winfo_children()[
                                    0].winfo_children()[
                                    champs_index].winfo_children()[zone_index].configure(
                                    text="Zone gestion " + str(zone_index + 1))
                                zone_index += 1
                            for simulation_frame in simulation_notebook.winfo_children():
                                if len(simulation_frame.winfo_children()) == 0:
                                    pass
                                else:
                                    notebook = simulation_frame.winfo_children()[0]
                                    champs_courant_zone_notebook = \
                                        notebook.winfo_children()[champs_index].winfo_children()[0]
                                    champs_courant_zone_notebook.winfo_children()[index_clicked_tab].destroy()
                                    current_index = index_clicked_tab
                                    while current_index < champs_courant_zone_notebook.index("end"):
                                        champs_courant_zone_notebook.tab(current_index,
                                                                         text="Zone gestion " + str(current_index + 1))
                                        if current_index == champs_courant_zone_notebook.index("end") - 1:
                                            champs_courant_zone_notebook.tab(current_index, text="+")
                                        current_index += 1
                            global max_index_zone
                            global furthest_left_tab_index_zone
                            global furthest_right_tab_index_zone
                            champs_courant_notebook = \
                                simulation_notebook.winfo_children()[
                                    simulation_notebook.index("current")].winfo_children()[
                                    0]
                            champs_courant_zone_notebook = \
                                champs_courant_notebook.winfo_children()[
                                    champs_courant_notebook.index("current")].winfo_children()[
                                    0]
                            if furthest_left_tab_index_zone == min_index_zone and len(
                                    champs_courant_zone_notebook.winfo_children()) > 4:
                                champs_courant_zone_notebook.tab(furthest_right_tab_index_zone, state="normal")
                            elif furthest_left_tab_index_zone == min_index_zone and len(
                                    champs_courant_zone_notebook.winfo_children()) == 4:
                                for tab_index in range(len(champs_courant_zone_notebook.winfo_children())):
                                    champs_courant_zone_notebook.tab(tab_index, state="normal")
                                scroll_right_button_zone.configure(state="disabled")
                            elif furthest_right_tab_index_zone == max_index_zone and len(
                                    champs_courant_zone_notebook.winfo_children()) == 4:
                                furthest_right_tab_index_zone -= 1
                                furthest_left_tab_index_zone -= 1
                                champs_courant_zone_notebook.tab(furthest_left_tab_index_zone, state="normal")
                                scroll_left_button_zone.configure(state="disabled")
                            elif furthest_left_tab_index_zone == min_index_zone and len(
                                    champs_courant_zone_notebook.winfo_children()) <= 3:
                                furthest_right_tab_index_zone -= 1
                            elif furthest_right_tab_index_zone == max_index_zone and len(
                                    champs_courant_zone_notebook.winfo_children()) > 4:
                                furthest_left_tab_index_zone -= 1
                                furthest_right_tab_index_zone -= 1
                                champs_courant_zone_notebook.tab(furthest_left_tab_index_zone, state="normal")
                            else:
                                champs_courant_zone_notebook.tab(furthest_right_tab_index_zone, state="normal")
                                if max_index_zone - 1 == furthest_right_tab_index_zone:
                                    scroll_right_button_zone.configure(state="disabled")
                            max_index_zone -= 1

                    def set_up_new_zone_de_gestion(champs_index):
                        global information_champs

                        def remise_a_etat_initial():
                            information_champs[champs_index]["nombre_de_zone_de_gestion"] = str(
                                int(information_champs[champs_index]["nombre_de_zone_de_gestion"]) - 1)
                            new_zone_window.destroy()

                        new_zone_window = tk.Toplevel()
                        new_zone_window.resizable(0, 0)
                        new_zone_window.protocol("WM_DELETE_WINDOW", remise_a_etat_initial)
                        creation_zone_frame = ttk.Frame(new_zone_window)
                        canvas = tk.Canvas(creation_zone_frame)
                        scrollbar = ttk.Scrollbar(creation_zone_frame, orient="vertical", command=canvas.yview)
                        scrollable_frame = ttk.Frame(canvas)
                        scrollable_frame.bind("<Configure>",
                                              lambda e: canvas.configure(scrollregion=canvas.bbox("all")))
                        canvas.create_window((0, 0), window=scrollable_frame, anchor="nw")
                        canvas.configure(yscrollcommand=scrollbar.set)
                        zone_de_gestion_frame = ttk.LabelFrame(scrollable_frame,
                                                               text="Zone gestion " + str(
                                                                   information_champs[champs_index][
                                                                       "nombre_de_zone_de_gestion"]))
                        taux_matiere_organique_label = ttk.Label(zone_de_gestion_frame,
                                                                 text="Taux matière organique (en %): ")
                        taux_matiere_organique_entry = ttk.Entry(zone_de_gestion_frame)
                        municipalite_label = ttk.Label(zone_de_gestion_frame, text="Municipalité: ")
                        global municipalites_supportees
                        municipalite_combobox = ttk.Combobox(zone_de_gestion_frame, values=municipalites_supportees,
                                                             postcommand=lambda: filter_combobox_values(
                                                                 municipalite_combobox,
                                                                 municipalites_supportees))
                        groupe_textural_label = ttk.Label(zone_de_gestion_frame, text="Groupe textural: ")
                        global groupes_texturaux_supportees
                        groupe_textural_combobox = ttk.Combobox(zone_de_gestion_frame,
                                                                values=groupes_texturaux_supportees,
                                                                postcommand=lambda: filter_combobox_values(
                                                                    groupe_textural_combobox,
                                                                    groupes_texturaux_supportees))
                        classe_de_drainage_label = ttk.Label(zone_de_gestion_frame, text="Classe de drainage: ")
                        global classes_de_drainage_supportees
                        classe_de_drainage_combobox = ttk.Combobox(zone_de_gestion_frame,
                                                                   values=classes_de_drainage_supportees,
                                                                   postcommand=lambda: filter_combobox_values(
                                                                       classe_de_drainage_combobox,
                                                                       classes_de_drainage_supportees))
                        masse_volumique_apparente_label = ttk.Label(zone_de_gestion_frame,
                                                                    text="Masse volumique apparente (g/cm3): ")
                        masse_volumique_apparente_entry = ttk.Entry(zone_de_gestion_frame)
                        profondeur_label = ttk.Label(zone_de_gestion_frame,
                                                     text="Profondeur de la couche arable (cm): ")
                        profondeur_entry = ttk.Entry(zone_de_gestion_frame)
                        superficie_de_la_zone_label = ttk.Label(zone_de_gestion_frame,
                                                                text="Superficie de la zone (ha): ")
                        superficie_de_la_zone_entry = ttk.Entry(zone_de_gestion_frame)
                        taux_matiere_organique_label.grid(row=0, column=0, sticky="w", pady=3, padx=5)
                        taux_matiere_organique_entry.grid(row=0, column=1, sticky="w", pady=3, padx=5)
                        municipalite_label.grid(row=1, column=0, sticky="w", pady=3, padx=5)
                        municipalite_combobox.grid(row=1, column=1, sticky="w", pady=3, padx=5)
                        groupe_textural_label.grid(row=2, column=0, sticky="w", pady=3, padx=5)
                        groupe_textural_combobox.grid(row=2, column=1, sticky="w", pady=3, padx=5)
                        classe_de_drainage_label.grid(row=3, column=0, sticky="w", pady=3, padx=5)
                        classe_de_drainage_combobox.grid(row=3, column=1, sticky="w", pady=3, padx=5)
                        masse_volumique_apparente_label.grid(row=4, column=0, sticky="w", pady=3, padx=5)
                        masse_volumique_apparente_entry.grid(row=4, column=1, sticky="w", pady=3, padx=5)
                        profondeur_label.grid(row=5, column=0, sticky="w", pady=3, padx=5)
                        profondeur_entry.grid(row=5, column=1, sticky="w", pady=3, padx=5)
                        superficie_de_la_zone_label.grid(row=6, column=0, sticky="w", pady=3, padx=5)
                        superficie_de_la_zone_entry.grid(row=6, column=1, sticky="w", pady=3, padx=5)

                        zone_de_gestion_frame.pack()

                        creation_zone_de_gestion_bouton = ttk.Button(scrollable_frame, text="Créer",
                                                                     command=lambda: add_new_tab())
                        creation_zone_de_gestion_bouton.pack()
                        canvas.pack(side="left", fill="x", expand=True)
                        scrollbar.pack(side="right", fill="y")
                        creation_zone_frame.pack()

                        def add_new_tab():
                            entree_invalide_liste = []
                            global information_champs
                            index_zone = 0
                            for scrollable_frame_widget in scrollable_frame.winfo_children():
                                if isinstance(scrollable_frame_widget, ttk.LabelFrame):
                                    grid_slave0_1 = scrollable_frame_widget.grid_slaves(row=0, column=1)
                                    for entry in grid_slave0_1:
                                        taux_matiere_organique = entry.get()
                                        if not util.is_decimal_number(taux_matiere_organique) or float(
                                                taux_matiere_organique) < 0 or float(taux_matiere_organique) > 100:
                                            entree_invalide_liste.append(
                                                (information_champs[champs_index]["nom_du_champs"],
                                                 "Zone gestion " + str(
                                                     information_champs[champs_index]["nombre_de_zone_de_gestion"]),
                                                 "\"Taux de matière organique\" doit être un réel positif dans l'intervalle [0-100]"))
                                        else:
                                            taux_matiere_organique = float(taux_matiere_organique)
                                    grid_slave1_1 = scrollable_frame_widget.grid_slaves(row=1, column=1)
                                    for entry in grid_slave1_1:
                                        municipalite = entry.get()
                                        global municipalites_supportees
                                        if municipalite not in municipalites_supportees:
                                            entree_invalide_liste.append(
                                                (information_champs[champs_index]["nom_du_champs"],
                                                 "Zone gestion " + str(
                                                     information_champs[champs_index]["nombre_de_zone_de_gestion"]),
                                                 "\"Municipalité\" doit être parmis les choix disponibles"))
                                    grid_slave2_1 = scrollable_frame_widget.grid_slaves(row=2, column=1)
                                    for entry in grid_slave2_1:
                                        groupe_textural = entry.get()
                                        global groupes_texturaux_supportees
                                        if groupe_textural not in groupes_texturaux_supportees:
                                            entree_invalide_liste.append(
                                                (information_champs[champs_index]["nom_du_champs"],
                                                 "Zone gestion " + str(
                                                     information_champs[champs_index]["nombre_de_zone_de_gestion"]),
                                                 "\"Groupe textural\" doit être parmis les choix disponibles"))
                                    grid_slave3_1 = scrollable_frame_widget.grid_slaves(row=3, column=1)
                                    for entry in grid_slave3_1:
                                        classe_de_drainage = entry.get()
                                        global classes_de_drainage_supportees
                                        if classe_de_drainage not in classes_de_drainage_supportees:
                                            entree_invalide_liste.append(
                                                (information_champs[champs_index]["nom_du_champs"],
                                                 "Zone gestion " + str(
                                                     information_champs[champs_index]["nombre_de_zone_de_gestion"]),
                                                 "\"Classe de drainage\" doit être parmis les choix disponibles"))
                                    grid_slave4_1 = scrollable_frame_widget.grid_slaves(row=4, column=1)
                                    for entry in grid_slave4_1:
                                        masse_volumique_apparente = entry.get()
                                        if (not util.is_decimal_number(
                                                masse_volumique_apparente) and masse_volumique_apparente != "") or (
                                                util.is_decimal_number(masse_volumique_apparente) and float(
                                            masse_volumique_apparente) < 0):
                                            entree_invalide_liste.append(
                                                (information_champs[champs_index]["nom_du_champs"],
                                                 "Zone gestion " + str(
                                                     information_champs[champs_index]["nombre_de_zone_de_gestion"]),
                                                 "\"Masse volumique apparente\" doit être un réel positif ou laissé vide pour aller chercher la valeur par défaut"))
                                        else:
                                            if masse_volumique_apparente == "":
                                                masse_volumique_apparente = None
                                            else:
                                                masse_volumique_apparente = float(masse_volumique_apparente)
                                    grid_slave5_1 = scrollable_frame_widget.grid_slaves(row=5, column=1)
                                    for entry in grid_slave5_1:
                                        profondeur = entry.get()
                                        if not util.is_decimal_number(profondeur) or float(
                                                profondeur) < 0:
                                            entree_invalide_liste.append(
                                                (information_champs[champs_index]["nom_du_champs"],
                                                 "Zone gestion " + str(
                                                     information_champs[champs_index]["nombre_de_zone_de_gestion"]),
                                                 "\"Profondeur de la couche arable\" doit être un réel positif"))
                                        else:
                                            profondeur = float(profondeur)
                                    grid_slave6_1 = scrollable_frame_widget.grid_slaves(row=6, column=1)
                                    for entry in grid_slave6_1:
                                        superficie_de_la_zone = entry.get()
                                        if not util.is_decimal_number(superficie_de_la_zone) or float(
                                                superficie_de_la_zone) < 0:
                                            entree_invalide_liste.append(
                                                (information_champs[champs_index]["nom_du_champs"],
                                                 "Zone gestion " + str(
                                                     information_champs[champs_index]["nombre_de_zone_de_gestion"]),
                                                 "\"Superficie de la zone\" doit être un réel positif"))
                                        else:
                                            superficie_de_la_zone = float(superficie_de_la_zone)
                                    information_champs[champs_index]["information_zone_de_gestion"].append(
                                        {"taux_matiere_organique": taux_matiere_organique,
                                         "municipalite": municipalite,
                                         "groupe_textural": groupe_textural,
                                         "classe_de_drainage": classe_de_drainage,
                                         "masse_volumique_apparente": masse_volumique_apparente,
                                         "profondeur": profondeur,
                                         "superficie_de_la_zone": superficie_de_la_zone})
                                index_zone += 1
                            if len(entree_invalide_liste) == 0:
                                sauvegarder_attributs_entreprise_apres_modification()
                                rechauffement_champs_label_frame = \
                                    donnees_de_rechauffement_label_frame.winfo_children()[0].winfo_children()[
                                        0].winfo_children()[
                                        champs_index]
                                rechauffement_champs_new_zone_label_frame = ttk.LabelFrame(
                                    rechauffement_champs_label_frame,
                                    text="Zone gestion " + str(int(len(
                                        rechauffement_champs_label_frame.winfo_children()) / 2) + 1),
                                    name="zonegestion" + str(len(
                                        rechauffement_champs_label_frame.winfo_children())))
                                rechauffement_champs_new_zone_label_frame.pack()
                                ajouter_une_annee_a_la_rotation(rechauffement_champs_new_zone_label_frame)
                                new_zone_window.destroy()

                                global furthest_left_tab_index_zone
                                global furthest_right_tab_index_zone
                                global max_index_zone
                                index = 0
                                for simulation_frame in simulation_notebook.winfo_children():
                                    if len(simulation_frame.winfo_children()) == 0:
                                        pass
                                    else:
                                        notebook = simulation_frame.winfo_children()[0]
                                        champs_courant_zone_notebook = \
                                            notebook.winfo_children()[champs_index].winfo_children()[0]
                                        champs_courant_zone_notebook.winfo_children()[
                                            len(champs_courant_zone_notebook.winfo_children()) - 1].destroy()
                                        if furthest_right_tab_index_zone >= 3:
                                            champs_courant_zone_notebook.tab(furthest_left_tab_index_zone,
                                                                             state="hidden")
                                        tab = ttk.Frame(champs_courant_zone_notebook, name="zonegestion" + str(
                                            len(champs_courant_zone_notebook.winfo_children())))
                                        champs_courant_zone_notebook.add(tab, text="Zone gestion " + str(
                                            len(champs_courant_zone_notebook.winfo_children())))
                                        set_up_regies_projections(tab)

                                        new_zone_tab = ttk.Frame(champs_courant_zone_notebook)
                                        champs_courant_zone_notebook.add(new_zone_tab, text="+")
                                    index += 1
                                furthest_right_tab_index_zone += 1
                                if furthest_right_tab_index_zone > 3:
                                    furthest_left_tab_index_zone += 1
                                max_index_zone += 1
                                if furthest_left_tab_index_zone != min_index_zone:
                                    scroll_left_button_zone.configure(state="normal")
                            else:
                                information_champs[champs_index]["information_zone_de_gestion"].pop(
                                    len(information_champs[champs_index]["information_zone_de_gestion"]) - 1)
                                message = ""
                                for entree_invalide in entree_invalide_liste:
                                    message = message + "Dans le " + entree_invalide[0] + " et la " + entree_invalide[
                                        1] + " l'entrée " + entree_invalide[2] + "\n"
                                messagebox.showwarning("Warning", message)
                                new_zone_window.focus()

                    zone_notebook.bind("<Button-1>", add_new_zone_de_gestion_tab)
                    zone_notebook.bind("<Button-3>", delete_zone_de_gestion_tab)

                    if simulation_copie is not None and index_champs is not None:
                        champs = simulation_copie["entreprise_agricole"]["champs"][index_champs]
                    else:
                        champs = None

                    global duree_simulation
                    for zone in range(int(nombre_de_zone)):
                        zone_tab = ttk.Frame(zone_notebook, name="zonegestion" + str(zone))
                        zone_notebook.add(zone_tab, text="Zone gestion " + str(zone + 1), state="hidden")
                        set_up_regies_projections(zone_tab, champs, zone)

                    new_tab = ttk.Frame(zone_notebook)
                    zone_notebook.add(new_tab, text="+", state="hidden")
                    zone_notebook.pack(anchor="nw")

                def set_up_regies_projections(zone_tab, champs=None, zone_index=None):
                    if champs is not None:
                        zone = champs["zones_de_gestion"][zone_index]
                    else:
                        zone = None
                    projection_frame = ttk.LabelFrame(zone_tab, text="Régies des sols et cultures projetées")
                    canvas_projection = tk.Canvas(projection_frame)
                    scrollbar_projection = ttk.Scrollbar(projection_frame, orient="vertical",
                                                         command=canvas_projection.yview)
                    scrollable_frame_projection = ttk.Frame(canvas_projection)
                    scrollable_frame_projection.bind("<Configure>", lambda e: canvas_projection.configure(
                        scrollregion=canvas_projection.bbox("all")))
                    canvas_projection.create_window((0, 0), window=scrollable_frame_projection, anchor="nw")
                    canvas_projection.configure(yscrollcommand=scrollbar_projection.set)
                    canvas_projection.pack(side="left", fill="both", expand=True)
                    scrollbar_projection.pack(side="right", fill="y")
                    ajouter_une_annee_a_la_rotation(scrollable_frame_projection, zone)
                    projection_frame.grid(row=1, column=0, columnspan=2, pady=3)
                    get_information_simulation_button = ttk.Button(zone_tab, text="Lancer les simulations",
                                                                   command=get_information_toutes_les_simulations)
                    get_information_simulation_button.grid(row=3, column=0, pady=3)
                    editer_information_entreprise_button = ttk.Button(zone_tab, text="Éditer attributs de l'entreprise",
                                                                      command=editer_caracteristique_physique_entreprise)
                    editer_information_entreprise_button.grid(row=3, column=1, pady=3)
                    appliquer_les_regies_aux_autres_zones_button = ttk.Button(zone_tab,
                                                                              text="Appliquer la rotation aux autres zones du champ",
                                                                              command=lambda: appliquer_les_regies_aux_autres_zones(
                                                                                  zone_tab))
                    appliquer_les_regies_aux_autres_zones_button.grid(row=2, column=0, columnspan=2, pady=3)

                def ajouter_une_annee_a_la_rotation(scrollable_frame_projection, zone=None, zone_rechauffement=None):
                    if len(scrollable_frame_projection.winfo_children()) == 0:
                        if zone is not None:
                            for regie in zone["regies_sol_et_culture_projection"]:
                                index = len(scrollable_frame_projection.winfo_children()) + 1
                                add_regies_projection(scrollable_frame_projection, index, regie)
                        elif zone_rechauffement is not None and len(
                                zone_rechauffement["regies_sol_et_culture_historique"]) > 0:
                            for regie in zone_rechauffement["regies_sol_et_culture_historique"]:
                                index = len(scrollable_frame_projection.winfo_children()) + 1
                                add_regies_projection(scrollable_frame_projection, index, regie)
                        else:
                            index = len(scrollable_frame_projection.winfo_children()) + 1
                            add_regies_projection(scrollable_frame_projection, index)

                        button_frame = ttk.Frame(scrollable_frame_projection)
                        ajouter_une_annee_a_la_rotation_button = ttk.Button(button_frame,
                                                                            text="Ajouter une année de rotation",
                                                                            command=lambda: ajouter_une_annee_a_la_rotation(
                                                                                scrollable_frame_projection))
                        ajouter_une_annee_a_la_rotation_button.grid(row=0, column=0, pady=3)
                        if zone is not None and len(zone["regies_sol_et_culture_projection"]) > 1:
                            enlever_une_annee_a_la_rotation_button = ttk.Button(button_frame,
                                                                                text="Enlever une année de rotation",
                                                                                command=lambda: enlever_une_annee_a_la_rotation(
                                                                                    scrollable_frame_projection))
                        else:
                            enlever_une_annee_a_la_rotation_button = ttk.Button(button_frame,
                                                                                text="Enlever une année de rotation",
                                                                                state="disabled")
                        enlever_une_annee_a_la_rotation_button.grid(row=0, column=1, pady=3)
                        button_frame.pack()
                    else:
                        scrollable_frame_projection.winfo_children()[
                            len(scrollable_frame_projection.winfo_children()) - 1].destroy()
                        index = len(scrollable_frame_projection.winfo_children()) + 1
                        add_regies_projection(scrollable_frame_projection, index)
                        button_frame = ttk.Frame(scrollable_frame_projection)

                        ajouter_une_annee_a_la_rotation_button = ttk.Button(button_frame,
                                                                            text="Ajouter une année de rotation",
                                                                            command=lambda: ajouter_une_annee_a_la_rotation(
                                                                                scrollable_frame_projection))
                        ajouter_une_annee_a_la_rotation_button.grid(row=0, column=0, pady=3)
                        enlever_une_annee_a_la_rotation_button = ttk.Button(button_frame,
                                                                            text="Enlever une année de rotation",
                                                                            command=lambda: enlever_une_annee_a_la_rotation(
                                                                                scrollable_frame_projection))
                        enlever_une_annee_a_la_rotation_button.grid(row=0, column=1, pady=3)
                        button_frame.pack()

                def enlever_une_annee_a_la_rotation(scrollable_frame_projection):
                    if len(scrollable_frame_projection.winfo_children()) > 1:
                        scrollable_frame_projection.winfo_children()[
                            len(scrollable_frame_projection.winfo_children()) - 2].destroy()
                    if len(scrollable_frame_projection.winfo_children()) - 1 == 1:
                        scrollable_frame_projection.winfo_children()[
                            len(scrollable_frame_projection.winfo_children()) - 1].grid_slaves(row=0, column=1)[
                            0].configure(
                            state="disabled")

                def appliquer_les_regies_aux_autres_zones_rechauffement(champ_label_frame, zone_courante_index):
                    zone_courante = champ_label_frame.nametowidget("zone" + str(zone_courante_index))
                    regies = zone_courante.winfo_children()
                    regies.pop()
                    entree_invalide_liste = []
                    regies_zone_courante = []
                    for regie in regies:
                        culture_principale = regie.grid_slaves(row=0, column=1)[0].get()
                        global cultures_principales_supportees
                        if culture_principale not in cultures_principales_supportees:
                            entree_invalide_liste.append(
                                "\"Culture principale\" doit être parmis les choix disponibles dans la zone que vous souhaitez appliquer")
                        rendement = regie.grid_slaves(row=1, column=1)[0].get()
                        if rendement.isalpha() or (
                                not util.is_decimal_number(rendement) and rendement != "" and float(rendement) < 0):
                            entree_invalide_liste.append(
                                "\"Rendement\" doit être un réel positif ou la case peut être vide pour aller chercher un rendement par défaut dans la zone que vous souhaitez appliquer")
                        if rendement == "":
                            rendement = None
                        if rendement is not None and util.is_decimal_number(rendement) and float(rendement) >= 0:
                            rendement = float(rendement)
                        pourcentage_tige_exporte = regie.grid_slaves(row=2, column=1)[0].get()
                        if (not util.is_decimal_number(
                                pourcentage_tige_exporte) and pourcentage_tige_exporte != "") or (
                                util.is_decimal_number(pourcentage_tige_exporte) and (
                                float(pourcentage_tige_exporte) < 0 or float(pourcentage_tige_exporte) > 100)):
                            entree_invalide_liste.append(
                                "\"Pourcentage paille ou tige exportée\" doit être un réel positif dans l'intervalle [0,1] ou le champs peut être vide pour aller chercher une proportion par défaut dans la zone que vous souhaitez appliquer")
                        if pourcentage_tige_exporte == "":
                            pourcentage_tige_exporte = None
                        if pourcentage_tige_exporte is not None and util.is_decimal_number(
                                pourcentage_tige_exporte):
                            pourcentage_tige_exporte = float(pourcentage_tige_exporte)
                        production_recolte = regie.grid_slaves(row=3, column=1)[0].get()
                        if production_recolte not in ["Oui", "Non"]:
                            entree_invalide_liste.append(
                                "\"Production récolté\" doit être l'une des options de la combobox dans la zone que vous souhaitez appliquer")
                        else:
                            if production_recolte == "Oui":
                                production_recolte = True
                            else:
                                production_recolte = False
                        pourcentage_humidite = regie.grid_slaves(row=4, column=1)[0].get()
                        if (not util.is_decimal_number(
                                pourcentage_humidite) and pourcentage_humidite != "") or (
                                util.is_decimal_number(pourcentage_humidite) and (
                                float(pourcentage_humidite) < 0 or float(pourcentage_humidite) > 100)):
                            entree_invalide_liste.append(
                                "\"Pourcentage d'humidité\" doit être un réel positif dans l'intervalle [0,100] ou le champs peut être vide pour aller chercher une proportion par défaut dans la zone que vous souhaitez appliquer")
                        if pourcentage_humidite == "":
                            pourcentage_humidite = None
                        if pourcentage_humidite is not None and util.is_decimal_number(pourcentage_humidite):
                            pourcentage_humidite = float(pourcentage_humidite)
                        culture_principale_dict = {"culture_principale": culture_principale,
                                                   "rendement": rendement,
                                                   "pourcentage_tige_exporte": pourcentage_tige_exporte,
                                                   "produit_recolte": production_recolte,
                                                   "pourcentage_humidite": pourcentage_humidite}
                        travail_du_sol = regie.grid_slaves(row=5, column=1)[0].get()
                        global types_travail_du_sol_supportes
                        if travail_du_sol not in types_travail_du_sol_supportes:
                            entree_invalide_liste.append(
                                "\"Travail du sol\" doit être parmis les choix disponibles dans la zone que vous souhaitez appliquer")
                        travail_du_sol_dict = {"travail_du_sol": travail_du_sol}
                        culture_secondaire = regie.grid_slaves(row=7, column=1)[0].get()
                        global cultures_secondaires_supportees
                        if culture_secondaire == "":
                            culture_secondaire = None
                        if culture_secondaire is not None and culture_secondaire not in cultures_secondaires_supportees:
                            entree_invalide_liste.append(
                                "\"Culture secondaire\" doit être parmis les choix disponibles ou laissé vide s'il n'y a pas de culture secondaire dans la zone que vous souhaitez appliquer")
                        rendement_culture_secondaire = regie.grid_slaves(row=8, column=1)[0].get()
                        if rendement_culture_secondaire == "":
                            rendement_culture_secondaire = None
                        if rendement_culture_secondaire is not None and not util.is_decimal_number(
                                rendement_culture_secondaire):
                            entree_invalide_liste.append(
                                "\"Rendement culture secondaire\" doit être un réel positif ou laissé vide s'il n'y a pas de culture secondaire")
                        elif (culture_secondaire is None and rendement_culture_secondaire is not None) or (
                                culture_secondaire is not None and rendement_culture_secondaire is None):
                            entree_invalide_liste.append(
                                "\"Rendement culture secondaire\" et \"Culture secondaire\" doivent être tout deux laissé vide s'il n'y a pas de culture secondaire")
                        else:
                            if rendement_culture_secondaire is not None:
                                rendement_culture_secondaire = float(rendement_culture_secondaire)
                        culture_secondaire_dict = {"culture_secondaire": culture_secondaire,
                                                   "rendement": rendement_culture_secondaire}
                        index_composante_amendement = 0
                        composante_amendement_liste = regie.grid_slaves(row=9, column=0)[0]
                        grid_size = composante_amendement_liste.grid_size()
                        amendements = []
                        while index_composante_amendement < grid_size[1] - 1:
                            amendement = \
                                composante_amendement_liste.grid_slaves([index_composante_amendement], column=1)[
                                    0].get()
                            global amendements_supportees
                            if amendement == "":
                                amendement = None
                            if amendement is not None and amendement not in amendements_supportees:
                                entree_invalide_liste.append(
                                    "\"Amendement\" " + str(
                                        index_composante_amendement + 1) + " doit être parmis les choix disponibles dans la zone que vous souhaitez appliquer")
                            apport = \
                                composante_amendement_liste.grid_slaves([index_composante_amendement + 1],
                                                                        column=1)[
                                    0].get()
                            if apport == "":
                                apport = None
                            if apport is not None and not util.is_decimal_number(apport):
                                entree_invalide_liste.append(
                                    "\"Apport\" " + str(index_composante_amendement + 1) +
                                    " est invalide, il doit être un réel positif ou laissé vide s'il n'y a pas d'amendements dans la zone que vous souhaitez appliquer")
                            elif (amendement is None and apport is not None) or (
                                    amendement is not None and apport is None):
                                entree_invalide_liste.append("\"Apport\" " + str(
                                    index_composante_amendement + 1) + " et \"Amendement\" " + str(
                                    index_composante_amendement + 1) +
                                                             " doivent être tout deux laissé vide s'il n'y a pas d'amendements  dans la zone que vous souhaitez appliquer")
                            else:
                                if apport is not None:
                                    apport = float(apport)
                            amendements.append({"amendement": amendement,
                                                "apport": apport})
                            index_composante_amendement += 2
                        regie_dict = {"culture_principale": culture_principale_dict,
                                      "culture_secondaire": culture_secondaire_dict,
                                      "amendements": amendements,
                                      "travail_du_sol": travail_du_sol_dict}
                        regies_zone_courante.append(regie_dict)
                    if len(entree_invalide_liste) == 0:
                        regies_zone_courante = {"regies_sol_et_culture_historique": regies_zone_courante}
                        index_zone = 0
                        index_widget = 0
                        for widget in champ_label_frame.winfo_children():
                            if index_widget % 2 == 0:
                                if index_zone != zone_courante_index:
                                    for zone_widget in widget.winfo_children():
                                        zone_widget.destroy()
                                    ajouter_une_annee_a_la_rotation(widget, zone_rechauffement=regies_zone_courante)
                                index_zone += 1
                            else:
                                pass
                            index_widget += 1
                    else:
                        message = ""
                        for entree_invalide in entree_invalide_liste:
                            message = message + entree_invalide + "\n"
                        messagebox.showwarning("Warning", message)
                        parent_frame_tabs.focus()

                def set_up_regies_rechauffement(rechauffement_frame, simulations):
                    def generer_appliquer_regies_bouton(label_frame, index_champ, index_zone):
                        appliquer_les_regies_aux_autres_zones_rechauffement_button = ttk.Button(
                            label_frame,
                            text="Appliquer la rotation aux autres zones du champ",
                            name="btn" + str(index_champ) + str(index_zone),
                            command=lambda: appliquer_les_regies_aux_autres_zones_rechauffement(
                                champ_label_frame=label_frame,
                                zone_courante_index=index_from_button_name[
                                    appliquer_les_regies_aux_autres_zones_rechauffement_button.winfo_name()]))
                        index_from_button_name[
                            appliquer_les_regies_aux_autres_zones_rechauffement_button.winfo_name()] = index
                        appliquer_les_regies_aux_autres_zones_rechauffement_button.pack()

                    global information_champs
                    canvas_rechauffement_via_rotation = tk.Canvas(rechauffement_frame)
                    scrollbar_rechauffement_via_rotation = ttk.Scrollbar(rechauffement_frame, orient="vertical",
                                                                         command=canvas_rechauffement_via_rotation.yview)
                    scrollable_frame_rechauffement_via_rotation = ttk.Frame(canvas_rechauffement_via_rotation)
                    scrollable_frame_rechauffement_via_rotation.bind("<Configure>",
                                                                     lambda
                                                                         e: canvas_rechauffement_via_rotation.configure(
                                                                         scrollregion=canvas_rechauffement_via_rotation.bbox(
                                                                             "all")))
                    canvas_rechauffement_via_rotation.create_window((0, 0),
                                                                    window=scrollable_frame_rechauffement_via_rotation,
                                                                    anchor="nw")
                    canvas_rechauffement_via_rotation.configure(yscrollcommand=scrollbar_rechauffement_via_rotation.set)
                    canvas_rechauffement_via_rotation.pack(side="left", fill="both", expand=True)
                    scrollbar_rechauffement_via_rotation.pack(side="right", fill="y")
                    index_from_button_name = {}
                    if simulations is not None:
                        entreprise = simulations[0]["entreprise_agricole"]
                        champs_liste = entreprise["champs"]
                    index_champs = 0
                    for champs in information_champs:
                        rechauffement_champs_label_frame = ttk.LabelFrame(scrollable_frame_rechauffement_via_rotation,
                                                                          text=champs["nom_du_champs"],
                                                                          name="champ" + str(index_champs))
                        rechauffement_champs_label_frame.pack()
                        if simulations is not None:
                            champs_courant = champs_liste[index_champs]
                            zone_liste = champs_courant["zones_de_gestion"]
                        index = 0
                        while index < int(champs["nombre_de_zone_de_gestion"]):
                            rechauffement_zone_label_frame = ttk.LabelFrame(rechauffement_champs_label_frame,
                                                                            text="Zone gestion " + str(index + 1),
                                                                            name="zone" + str(index))
                            if simulations is not None:
                                zone_rechauffement = zone_liste[index]
                            rechauffement_zone_label_frame.pack()
                            if simulations is not None:
                                ajouter_une_annee_a_la_rotation(rechauffement_zone_label_frame,
                                                                zone_rechauffement=zone_rechauffement)
                            else:
                                ajouter_une_annee_a_la_rotation(rechauffement_zone_label_frame)
                            generer_appliquer_regies_bouton(rechauffement_champs_label_frame, index_champs, index)
                            index += 1
                        index_champs += 1
                    rechauffement_frame.grid(row=1, column=0, columnspan=2, pady=3)

                def ajouter_des_amendements(amendement_frame, amendements=None):

                    def on_amendment_selected(event):
                        if amendement_combobox.get() != "" and apport_amendement_entry.get() == "":
                            apport_amendement_entry.insert(0, "10.0")

                    def definition_composante_amendements_charges(amendement_courrant):
                        def on_amendment_charge_selected(event):
                            if amendement_charge_combobox.get() != "" and apport_amendement_charge_entry.get() == "":
                                apport_amendement_charge_entry.insert(0, "10.0")

                        amendement_charge_label = ttk.Label(amendement_frame, text="Amendement: ")
                        amendement_charge_combobox = ttk.Combobox(amendement_frame, values=amendements_supportees,
                                                                  postcommand=lambda: filter_combobox_values(
                                                                      amendement_charge_combobox,
                                                                      amendements_supportees))
                        if amendement_courrant["amendement"] is not None:
                            amendement_charge_combobox.set(amendement_courrant["amendement"])
                        apport_amendement_charge_label = ttk.Label(amendement_frame, text="Apport (t/ha):")
                        apport_amendement_charge_entry = ttk.Entry(amendement_frame)
                        if amendement_courrant["apport"] is not None:
                            apport_amendement_charge_entry.insert(0, str(amendement_courrant["apport"]))
                        amendement_charge_label.grid(row=index, column=0, sticky="w", pady=3)
                        amendement_charge_combobox.grid(row=index, column=1, sticky="w", pady=3)
                        apport_amendement_charge_label.grid(row=index + 1, column=0, sticky="w", pady=3)
                        apport_amendement_charge_entry.grid(row=index + 1, column=1, sticky="w", pady=3)
                        amendement_charge_combobox.bind("<<ComboboxSelected>>", on_amendment_charge_selected)

                    global amendements_supportees
                    if amendements is not None:
                        index = 0
                        for amendement in amendements:
                            definition_composante_amendements_charges(amendement)
                            index += 2
                    else:
                        index = 0
                        amendement_label = ttk.Label(amendement_frame, text="Amendement: ")
                        amendement_combobox = ttk.Combobox(amendement_frame, values=amendements_supportees,
                                                           postcommand=lambda: filter_combobox_values(
                                                               amendement_combobox,
                                                               amendements_supportees))
                        apport_amendement_label = ttk.Label(amendement_frame, text="Apport (t/ha):")
                        apport_amendement_entry = ttk.Entry(amendement_frame)
                        amendement_label.grid(row=0, column=0, sticky="w", pady=3)
                        amendement_combobox.grid(row=0, column=1, sticky="w", pady=3)
                        apport_amendement_label.grid(row=1, column=0, sticky="w", pady=3)
                        apport_amendement_entry.grid(row=1, column=1, sticky="w", pady=3)
                        amendement_combobox.bind("<<ComboboxSelected>>", on_amendment_selected)
                        index += 2

                    ajout_a_la_regie_button = ttk.Button(amendement_frame, text="Ajouter à la régie",
                                                         command=lambda: ajouter_amendement_regie(amendement_frame))

                    ajout_a_la_regie_button.grid(row=index, column=0, columnspan=2, pady=3)

                def ajouter_amendement_regie(amendement_frame):

                    def on_amendment_selected(event):
                        if amendement_combobox.get() != "" and apport_amendement_entry.get() == "":
                            apport_amendement_entry.insert(0, "10.0")

                    grid_size = amendement_frame.grid_size()
                    amendement_frame.grid_slaves(grid_size[1] - 1, grid_size[0] - 1)[0].destroy()
                    amendement_label = ttk.Label(amendement_frame, text="Amendement: ")
                    global amendements_supportees
                    amendement_combobox = ttk.Combobox(amendement_frame, values=amendements_supportees,
                                                       postcommand=lambda: filter_combobox_values(amendement_combobox,
                                                                                                  amendements_supportees))
                    apport_amendement_label = ttk.Label(amendement_frame, text="Apport (t/ha):")
                    apport_amendement_entry = ttk.Entry(amendement_frame)
                    ajout_a_la_regie_button = ttk.Button(amendement_frame, text="Ajouter à la régie",
                                                         command=lambda: ajouter_amendement_regie(amendement_frame))
                    amendement_label.grid(row=grid_size[1] - 1, column=grid_size[0] - 2, sticky="w", pady=3)
                    amendement_combobox.grid(row=grid_size[1] - 1, column=grid_size[0] - 1, sticky="w", pady=3)
                    apport_amendement_label.grid(row=grid_size[1], column=grid_size[0] - 2, sticky="w", pady=3)
                    apport_amendement_entry.grid(row=grid_size[1], column=grid_size[0] - 1, sticky="w", pady=3)
                    ajout_a_la_regie_button.grid(row=grid_size[1] + 1, column=grid_size[0] - 2, pady=3, columnspan=2)
                    amendement_combobox.bind("<<ComboboxSelected>>", on_amendment_selected)

                def add_regies_projection(zone_label_frame, index, regie=None):

                    def on_culture_principale_selected(event):
                        widget_parent = culture_principale_combobox.winfo_parent()
                        if "notebook" in widget_parent:
                            widget_parent_split_parts = widget_parent.split(".")
                            champ = widget_parent_split_parts[6]
                            zone = widget_parent_split_parts[8]
                            if champ[len(champ) - 1].isdigit():
                                index_champ = ""
                                for caractere in champ:
                                    if caractere.isdigit():
                                        index_champ = index_champ + caractere
                                index_champ = int(index_champ)
                            else:
                                index_champ = 0
                            if zone[len(zone) - 1].isdigit():
                                index_zone = ""
                                for caractere in zone:
                                    if caractere.isdigit():
                                        index_zone = index_zone + caractere
                                index_zone = int(index_zone)
                            else:
                                index_zone = 0
                        else:
                            widget_parent_split_parts = widget_parent.split(".")
                            champ = widget_parent_split_parts[6]
                            zone = widget_parent_split_parts[7]
                            if champ[len(champ) - 1].isdigit():
                                index_champ = ""
                                for caractere in champ:
                                    if caractere.isdigit():
                                        index_champ = index_champ + caractere
                                index_champ = int(index_champ)
                            else:
                                index_champ = 0
                            if zone[len(zone) - 1].isdigit():
                                index_zone = ""
                                for caractere in zone:
                                    if caractere.isdigit():
                                        index_zone = index_zone + caractere
                                index_zone = int(index_zone)
                            else:
                                index_zone = 0

                        global information_champs
                        municipalite = information_champs[index_champ]["information_zone_de_gestion"][index_zone][
                            "municipalite"]
                        if culture_principale_combobox.get() != "":
                            request_json = {"culture_principale": culture_principale_combobox.get(),
                                            "municipalite": municipalite}
                            response = requests.post(
                                'http://localhost:5000/api/get-parametres-defauts-culture_principale',
                                json=request_json)
                            if rendement_entry.get() == "":
                                rendement_entry.insert(0, response.json()["rendement"])
                            if pourcentage_tige_exporte_entry.get() == "":
                                pourcentage_tige_exporte_entry.insert(0, response.json()["pourcentage_tige_exportee"])
                            if production_recolte_combobox.get() == "":
                                production_recolte_combobox.insert(0, "Oui")
                            if pourcentage_humidite_entry.get() == "":
                                pourcentage_humidite_entry.insert(0, response.json()["pourcentage_humidite"])
                            if travail_du_sol_combobox.get() == "":
                                travail_du_sol_combobox.insert(0, response.json()["travail_du_sol_defaut"])

                    def on_culture_secondaire_selected(event):
                        if culture_secondaire_combobox.get() != "" and rendement_culture_secondaire_entry.get() == "":
                            rendement_culture_secondaire_entry.insert(0, "1.0")

                    annee_courante_frame = ttk.LabelFrame(zone_label_frame, text="Année " + str(index))
                    culture_principale_label = ttk.Label(annee_courante_frame, text="Culture principale: ")
                    global cultures_principales_supportees
                    culture_principale_combobox = ttk.Combobox(annee_courante_frame,
                                                               values=cultures_principales_supportees,
                                                               postcommand=lambda: filter_combobox_values(
                                                                   culture_principale_combobox,
                                                                   cultures_principales_supportees))
                    rendement_label = ttk.Label(annee_courante_frame, text="Rendement (t/ha): ")
                    rendement_entry = ttk.Entry(annee_courante_frame)
                    pourcentage_tige_exporte_label = ttk.Label(annee_courante_frame,
                                                               text="Pourcentage paille ou tige exporté [0-100]: ")
                    pourcentage_tige_exporte_entry = ttk.Entry(annee_courante_frame)
                    production_recolte_label = ttk.Label(annee_courante_frame, text="Production récoltée: ")
                    production_recolte_combobox = ttk.Combobox(annee_courante_frame, values=["Oui", "Non"],
                                                               postcommand=lambda: filter_combobox_values(
                                                                   production_recolte_combobox, ["Oui", "Non"]))
                    pourcentage_humidite_label = ttk.Label(annee_courante_frame,
                                                           text="Pourcentage d'humidité [0-100]: ")
                    pourcentage_humidite_entry = ttk.Entry(annee_courante_frame)
                    travail_du_sol_label = ttk.Label(annee_courante_frame, text="Travail du sol: ")
                    global types_travail_du_sol_supportes
                    travail_du_sol_combobox = ttk.Combobox(annee_courante_frame, values=types_travail_du_sol_supportes,
                                                           postcommand=lambda: filter_combobox_values(
                                                               travail_du_sol_combobox,
                                                               types_travail_du_sol_supportes))
                    culture_secondaire_label = ttk.Label(annee_courante_frame, text="Culture secondaire: ")
                    global cultures_secondaires_supportees
                    culture_secondaire_combobox = ttk.Combobox(annee_courante_frame,
                                                               values=cultures_secondaires_supportees,
                                                               postcommand=lambda: filter_combobox_values(
                                                                   culture_secondaire_combobox,
                                                                   cultures_secondaires_supportees))
                    rendement_culture_secondaire_label = ttk.Label(annee_courante_frame,
                                                                   text="Rendement culture secondaire (t/ha): ")
                    rendement_culture_secondaire_entry = ttk.Entry(annee_courante_frame)

                    if regie is not None:
                        culture_principale_combobox.set(regie["culture_principale"]["culture_principale"])
                        if regie["culture_principale"]["rendement"] is not None:
                            rendement_entry.insert(0, str(regie["culture_principale"]["rendement"]))

                        if regie["culture_principale"]["pourcentage_tige_exporte"] is not None:
                            pourcentage_tige_exporte_entry.insert(0, str(
                                regie["culture_principale"]["pourcentage_tige_exporte"]))
                        if regie["culture_principale"]["produit_recolte"]:
                            production_recolte_combobox.set("Oui")
                        else:
                            production_recolte_combobox.set("Non")
                        if regie["culture_principale"]["pourcentage_humidite"] is not None:
                            pourcentage_humidite_entry.insert(0,
                                                              str(regie["culture_principale"]["pourcentage_humidite"]))
                        travail_du_sol_combobox.set(regie["travail_du_sol"]["travail_du_sol"])
                        if regie["culture_secondaire"]["culture_secondaire"] is not None:
                            culture_secondaire_combobox.set(regie["culture_secondaire"]["culture_secondaire"])
                        if regie["culture_secondaire"]["rendement"]:
                            rendement_culture_secondaire_entry.insert(0, str(regie["culture_secondaire"]["rendement"]))

                    amendement_frame = ttk.LabelFrame(annee_courante_frame, text="Liste des amendements")
                    if regie is not None:
                        ajouter_des_amendements(amendement_frame, regie["amendements"])
                    else:
                        ajouter_des_amendements(amendement_frame)

                    culture_principale_label.grid(row=0, column=0, sticky="w", pady=3)
                    culture_principale_combobox.grid(row=0, column=1, sticky="w", pady=3)
                    rendement_label.grid(row=1, column=0, sticky="w", pady=3)
                    rendement_entry.grid(row=1, column=1, sticky="w", pady=3)
                    pourcentage_tige_exporte_label.grid(row=2, column=0, sticky="w", pady=3)
                    pourcentage_tige_exporte_entry.grid(row=2, column=1, sticky="w", pady=3)
                    production_recolte_label.grid(row=3, column=0, sticky="w", pady=3)
                    production_recolte_combobox.grid(row=3, column=1, sticky="w", pady=3)
                    pourcentage_humidite_label.grid(row=4, column=0, sticky="w", pady=3)
                    pourcentage_humidite_entry.grid(row=4, column=1, sticky="w", pady=3)
                    travail_du_sol_label.grid(row=5, column=0, sticky="w", pady=3)
                    travail_du_sol_combobox.grid(row=5, column=1, sticky="w", pady=3)
                    culture_secondaire_label.grid(row=7, column=0, sticky="w", pady=3)
                    culture_secondaire_combobox.grid(row=7, column=1, sticky="w", pady=3)
                    rendement_culture_secondaire_label.grid(row=8, column=0, sticky="w", pady=3)
                    rendement_culture_secondaire_entry.grid(row=8, column=1, sticky="w", pady=3)
                    amendement_frame.grid(row=9, column=0, columnspan=2, pady=3)
                    culture_principale_combobox.bind("<<ComboboxSelected>>", on_culture_principale_selected)
                    culture_secondaire_combobox.bind("<<ComboboxSelected>>", on_culture_secondaire_selected)
                    annee_courante_frame.pack()

                def appliquer_les_regies_aux_autres_zones(zone_tab):
                    simulation_index = simulation_notebook.index("current")
                    champ_notebook = \
                        simulation_notebook.winfo_children()[simulation_notebook.index("current")].winfo_children()[0]
                    champ_index = champ_notebook.index("current")
                    zone_notebook = champ_notebook.winfo_children()[champ_notebook.index("current")].winfo_children()[0]
                    zone_index = zone_notebook.index("current")
                    regie_frame = zone_tab.winfo_children()[0]
                    regies = get_regies(regie_frame, champ_index, zone_index, simulation_index)
                    entree_invalide_liste = regies[1]
                    if len(entree_invalide_liste) == 0:
                        index = 0
                        for zone in zone_notebook.winfo_children():
                            if index == zone_index or index == len(zone_notebook.winfo_children()) - 1:
                                pass
                            else:
                                zone_scrollable_frame = zone.winfo_children()[0].winfo_children()[0].winfo_children()[0]
                                for zone_label_frame_widget in zone_scrollable_frame.winfo_children():
                                    zone_label_frame_widget.destroy()
                                ajouter_une_annee_a_la_rotation(zone_scrollable_frame,
                                                                zone={"regies_sol_et_culture_projection": regies[0]})

                            index += 1
                    else:
                        message = ""
                        for entree_invalide in entree_invalide_liste:
                            message = message + "Dans la " + entree_invalide[3] + ", le " + entree_invalide[
                                0] + " et la " + \
                                      entree_invalide[
                                          1] + " l'entrée " + entree_invalide[2] + "\n"
                        messagebox.showwarning("Warning", message)
                        parent_frame_tabs.focus()

                def get_information_toutes_les_simulations():
                    simulations = []
                    entree_invalide_liste = []
                    index_simualtion = 0
                    for simulation in simulation_notebook.winfo_children():
                        global nombre_simulations
                        if index_simualtion < nombre_simulations:
                            simulation, entree_invalide_liste_simulation = get_information_simulation(simulation,
                                                                                                      index_simualtion,
                                                                                                      simulation_unique=False)
                            for entree in entree_invalide_liste_simulation:
                                entree_invalide_liste.append(entree)
                            simulations.append(simulation)
                        index_simualtion += 1
                    if len(entree_invalide_liste) == 0:
                        response = requests.post('http://localhost:5000/api/icbm-bilan',
                                                 json={"simulations": simulations})
                        print(response.text)
                        creation_du_rapport(response.json())
                    else:
                        message = ""
                        for entree_invalide in entree_invalide_liste:
                            message = message + "Dans la " + entree_invalide[3] + ", le " + entree_invalide[
                                0] + " et la " + \
                                      entree_invalide[
                                          1] + " l'entrée " + entree_invalide[2] + "\n"
                        messagebox.showwarning("Warning", message)
                        parent_frame_tabs.focus()

                def sauvegarder_toutes_les_simulations():
                    simulations = []
                    entree_invalide_liste = []
                    index_simulation = 0
                    for simulation in simulation_notebook.winfo_children():
                        global duree_simulation
                        global nombre_simulations
                        if index_simulation < nombre_simulations:
                            simulation, entree_invalide_liste_simulation = get_information_simulation(simulation,
                                                                                                      index_simulation,
                                                                                                      simulation_unique=False)
                            for entree in entree_invalide_liste_simulation:
                                entree_invalide_liste.append(entree)
                            simulations.append(simulation)
                        index_simulation += 1
                    if len(entree_invalide_liste) == 0:
                        return {"simulations": simulations}
                    else:
                        return entree_invalide_liste

                def get_information_simulation(simulation_frame, index_simulation, simulation_unique):
                    global information_champs
                    global nombre_de_champs
                    entree_invalide_simulation_liste = [[], []]
                    if simulation_unique:
                        regies_rechauffement, entree_invalide_liste = None, []
                    else:
                        regies_rechauffement, entree_invalide_liste = get_regies_rechauffement()
                    for entree in entree_invalide_liste:
                        entree_invalide_simulation_liste[0].append(entree)
                    champs_notebook = simulation_frame.winfo_children()[0]
                    index_champs = 0
                    champs_list = []
                    for champs_frame in champs_notebook.winfo_children():
                        zone_list = []
                        index_zone = 0
                        if index_champs < nombre_de_champs:
                            zone_notebook = champs_frame.winfo_children()[0]
                            for zone_frames in zone_notebook.winfo_children():
                                if index_zone < int(information_champs[index_champs]["nombre_de_zone_de_gestion"]):
                                    regies_projection_frame = zone_frames.winfo_children()[0]
                                    regies_projection, entree_invalide_liste = get_regies(regies_projection_frame,
                                                                                          index_champs,
                                                                                          index_zone, index_simulation)
                                    if regies_rechauffement is not None:
                                        regie_rechauffement = regies_rechauffement[index_champs][index_zone]
                                    else:
                                        regie_rechauffement = None
                                    zone_list.append({"regies_projection": regies_projection,
                                                      "regies_rechauffement": regie_rechauffement})
                                    for entree in entree_invalide_liste:
                                        entree_invalide_simulation_liste[1].append(entree)
                                index_zone += 1
                            champs_list.append(zone_list)
                        index_champs += 1
                    if len(entree_invalide_simulation_liste[0]) == 0 and len(entree_invalide_simulation_liste[1]) == 0:
                        champs_attributs = []
                        index_champs = 0
                        for champs in information_champs:
                            index_zone = 0
                            zones_de_gestion = []
                            for zone in champs["information_zone_de_gestion"]:
                                zones_de_gestion.append({"taux_matiere_organique": zone["taux_matiere_organique"],
                                                         "municipalite": zone["municipalite"],
                                                         "groupe_textural": zone["groupe_textural"],
                                                         "classe_de_drainage": zone["classe_de_drainage"],
                                                         "masse_volumique_apparente": zone["masse_volumique_apparente"],
                                                         "profondeur": zone["profondeur"],
                                                         "superficie_de_la_zone": zone["superficie_de_la_zone"],
                                                         "regies_sol_et_culture_projection":
                                                             champs_list[index_champs][index_zone]["regies_projection"],
                                                         "regies_sol_et_culture_historique":
                                                             champs_list[index_champs][index_zone][
                                                                 "regies_rechauffement"]})
                                index_zone += 1
                            champs_attributs.append({"nom": champs["nom_du_champs"],
                                                     "zones_de_gestion": zones_de_gestion})
                            index_champs += 1

                        global nom_entreprise
                        entreprise_agricole = {"nom": nom_entreprise,
                                               "champs": champs_attributs}
                        global duree_simulation
                        simulation = {
                            "nom_simulation": duree_simulation[index_simulation]["nom_simulation"],
                            "duree_projection": duree_simulation[index_simulation]["duree_projection"],
                            "annee_initiale_projection": int(
                                duree_simulation[index_simulation]["annee_projection_initiale"]),
                            "annee_finale_projection": int(
                                duree_simulation[index_simulation]["annee_projection_initiale"]) + int(
                                duree_simulation[index_simulation]["duree_projection"]) - 1,
                            "entreprise_agricole": entreprise_agricole}
                        return simulation, []
                    else:
                        entree_invalide_liste = []
                        for sous_liste in entree_invalide_simulation_liste:
                            for entree in sous_liste:
                                entree_invalide_liste.append(entree)
                        return None, entree_invalide_liste

                def get_regies(regies_frame, champs_index, zone_index, simulation_index):
                    entree_invalide_liste = []
                    regies = []
                    regies_canvas = regies_frame.winfo_children()[0]
                    regies_scrollable_frame = regies_canvas.winfo_children()[0]
                    compteur_regie = 0
                    for regie in regies_scrollable_frame.winfo_children():
                        if compteur_regie < len(regies_scrollable_frame.winfo_children()) - 1:
                            culture_principale = regie.grid_slaves(row=0, column=1)[0].get()
                            global cultures_principales_supportees
                            if culture_principale not in cultures_principales_supportees:
                                entree_invalide_liste.append(
                                    (information_champs[champs_index]["nom_du_champs"],
                                     "Zone gestion " + str(zone_index + 1),
                                     "\"Culture principale\" doit être parmis les choix disponibles",
                                     "Régie projection Simulation " + duree_simulation[simulation_index][
                                         "nom_simulation"]))
                            rendement = regie.grid_slaves(row=1, column=1)[0].get()
                            if rendement.isalpha() or (
                                    not util.is_decimal_number(rendement) and rendement != "" and float(rendement) < 0):
                                entree_invalide_liste.append(
                                    (information_champs[champs_index]["nom_du_champs"],
                                     "Zone gestion " + str(zone_index + 1),
                                     "\"Rendement\" doit être un réel positif ou la case peut être vide pour aller chercher un rendement par défaut",
                                     "Régie projection Simulation " + duree_simulation[simulation_index][
                                         "nom_simulation"]))
                            if rendement == "":
                                rendement = None
                            if rendement is not None and util.is_decimal_number(rendement) and float(rendement) >= 0:
                                rendement = float(rendement)
                            pourcentage_tige_exporte = regie.grid_slaves(row=2, column=1)[0].get()
                            if (not util.is_decimal_number(
                                    pourcentage_tige_exporte) and pourcentage_tige_exporte != "") or (
                                    util.is_decimal_number(pourcentage_tige_exporte) and (
                                    float(pourcentage_tige_exporte) < 0 or float(pourcentage_tige_exporte) > 100)):
                                entree_invalide_liste.append(
                                    (information_champs[champs_index]["nom_du_champs"],
                                     "Zone gestion " + str(zone_index + 1),
                                     "\"Pourcentage paille ou tige exportée\" doit être un réel positif dans l'intervalle [0,1] ou le champs peut être vide pour aller chercher une proportion par défaut",
                                     "Régie projection Simulation " + duree_simulation[simulation_index][
                                         "nom_simulation"]))
                            if pourcentage_tige_exporte == "":
                                pourcentage_tige_exporte = None
                            if pourcentage_tige_exporte is not None and util.is_decimal_number(
                                    pourcentage_tige_exporte):
                                pourcentage_tige_exporte = float(pourcentage_tige_exporte)
                            production_recolte = regie.grid_slaves(row=3, column=1)[0].get()
                            if production_recolte not in ["Oui", "Non"]:
                                entree_invalide_liste.append(
                                    (information_champs[champs_index]["nom_du_champs"],
                                     "Zone gestion " + str(zone_index + 1),
                                     "\"Production récolté\" doit être l'une des options de la combobox",
                                     "Régie projection Simulation " + duree_simulation[simulation_index][
                                         "nom_simulation"]))
                            else:
                                if production_recolte == "Oui":
                                    production_recolte = True
                                else:
                                    production_recolte = False
                            pourcentage_humidite = regie.grid_slaves(row=4, column=1)[0].get()
                            if (not util.is_decimal_number(
                                    pourcentage_humidite) and pourcentage_humidite != "") or (
                                    util.is_decimal_number(pourcentage_humidite) and (
                                    float(pourcentage_humidite) < 0 or float(pourcentage_humidite) > 100)):
                                entree_invalide_liste.append(
                                    (information_champs[champs_index]["nom_du_champs"],
                                     "Zone gestion " + str(zone_index + 1),
                                     "\"Pourcentage d'humidité\" doit être un réel positif dans l'intervalle [0,100] ou le champs peut être vide pour aller chercher une proportion par défaut",
                                     "Régie projection Simulation " + duree_simulation[simulation_index][
                                         "nom_simulation"]))
                            if pourcentage_humidite == "":
                                pourcentage_humidite = None
                            if pourcentage_humidite is not None and util.is_decimal_number(pourcentage_humidite):
                                pourcentage_humidite = float(pourcentage_humidite)
                            culture_principale_dict = {"culture_principale": culture_principale,
                                                       "rendement": rendement,
                                                       "pourcentage_tige_exporte": pourcentage_tige_exporte,
                                                       "produit_recolte": production_recolte,
                                                       "pourcentage_humidite": pourcentage_humidite}
                            travail_du_sol = regie.grid_slaves(row=5, column=1)[0].get()
                            global types_travail_du_sol_supportes
                            if travail_du_sol not in types_travail_du_sol_supportes:
                                entree_invalide_liste.append(
                                    (information_champs[champs_index]["nom_du_champs"],
                                     "Zone gestion " + str(zone_index + 1),
                                     "\"Travail du sol\" doit être parmis les choix disponibles",
                                     "Régie projection Simulation " + duree_simulation[simulation_index][
                                         "nom_simulation"]))
                            travail_du_sol_dict = {"travail_du_sol": travail_du_sol}
                            culture_secondaire = regie.grid_slaves(row=7, column=1)[0].get()
                            global cultures_secondaires_supportees
                            if culture_secondaire == "":
                                culture_secondaire = None
                            if culture_secondaire is not None and culture_secondaire not in cultures_secondaires_supportees:
                                entree_invalide_liste.append(
                                    (information_champs[champs_index]["nom_du_champs"],
                                     "Zone gestion " + str(zone_index + 1),
                                     "\"Culture secondaire\" doit être parmis les choix disponibles ou laissé vide s'il n'y a pas de culture secondaire",
                                     "Régie projection Simulation " + duree_simulation[simulation_index][
                                         "nom_simulation"]))
                            rendement_culture_secondaire = regie.grid_slaves(row=8, column=1)[0].get()
                            if rendement_culture_secondaire == "":
                                rendement_culture_secondaire = None
                            if rendement_culture_secondaire is not None and not util.is_decimal_number(
                                    rendement_culture_secondaire):
                                entree_invalide_liste.append(
                                    (information_champs[champs_index]["nom_du_champs"],
                                     "Zone gestion " + str(zone_index + 1),
                                     "\"Rendement culture secondaire\" doit être un réel positif ou laissé vide s'il n'y a pas de culture secondaire",
                                     "Régie projection Simulation " + duree_simulation[simulation_index][
                                         "nom_simulation"]))
                            elif (culture_secondaire is None and rendement_culture_secondaire is not None) or (
                                    culture_secondaire is not None and rendement_culture_secondaire is None):
                                entree_invalide_liste.append(
                                    (information_champs[champs_index]["nom_du_champs"],
                                     "Zone gestion " + str(zone_index + 1),
                                     "\"Rendement culture secondaire\" et \"Culture secondaire\" doivent être tout deux laissé vide s'il n'y a pas de culture secondaire",
                                     "Régie projection Simulation " + duree_simulation[simulation_index][
                                         "nom_simulation"]))
                            else:
                                if rendement_culture_secondaire is not None:
                                    rendement_culture_secondaire = float(rendement_culture_secondaire)
                            culture_secondaire_dict = {"culture_secondaire": culture_secondaire,
                                                       "rendement": rendement_culture_secondaire}
                            index_composante_amendement = 0
                            composante_amendement_liste = regie.grid_slaves(row=9, column=0)[0]
                            grid_size = composante_amendement_liste.grid_size()
                            amendements = []
                            while index_composante_amendement < grid_size[1] - 1:
                                amendement = \
                                    composante_amendement_liste.grid_slaves([index_composante_amendement], column=1)[
                                        0].get()
                                global amendements_supportees
                                if amendement == "":
                                    amendement = None
                                if amendement is not None and amendement not in amendements_supportees:
                                    entree_invalide_liste.append(
                                        (information_champs[champs_index]["nom_du_champs"],
                                         "Zone gestion " + str(zone_index + 1),
                                         "\"Amendement\" " + str(
                                             index_composante_amendement + 1) + " doit être parmis les choix disponibles",
                                         "Régie projection Simulation " + duree_simulation[simulation_index][
                                             "nom_simulation"]))
                                apport = \
                                    composante_amendement_liste.grid_slaves([index_composante_amendement + 1],
                                                                            column=1)[
                                        0].get()
                                if apport == "":
                                    apport = None
                                if apport is not None and not util.is_decimal_number(apport):
                                    entree_invalide_liste.append(
                                        (information_champs[champs_index]["nom_du_champs"],
                                         "Zone gestion " + str(zone_index + 1),
                                         "\"Apport\" " + str(index_composante_amendement + 1) +
                                         " est invalide, il doit être un réel positif ou laissé vide s'il n'y a pas d'amendements",
                                         "Régie projection Simulation " + duree_simulation[simulation_index][
                                             "nom_simulation"]))
                                elif (amendement is None and apport is not None) or (
                                        amendement is not None and apport is None):
                                    entree_invalide_liste.append(
                                        (information_champs[champs_index]["nom_du_champs"],
                                         "Zone gestion " + str(zone_index + 1),
                                         "\"Apport\" " + str(
                                             index_composante_amendement + 1) + " et \"Amendement\" " + str(
                                             index_composante_amendement + 1) +
                                         " doivent être tout deux laissé vide s'il n'y a pas d'amendements",
                                         "Régie projection Simulation " + duree_simulation[simulation_index][
                                             "nom_simulation"]))
                                else:
                                    if apport is not None:
                                        apport = float(apport)
                                amendements.append({"amendement": amendement,
                                                    "apport": apport})
                                index_composante_amendement += 2
                            regie_dict = {"culture_principale": culture_principale_dict,
                                          "culture_secondaire": culture_secondaire_dict,
                                          "amendements": amendements,
                                          "travail_du_sol": travail_du_sol_dict}
                            regies.append(regie_dict)
                        compteur_regie += 1
                    return regies, entree_invalide_liste

                def scroll_right_simulation():
                    global duree_simulation
                    global furthest_left_tab_index_simulation
                    global furthest_right_tab_index_simulation
                    simulation_notebook.tab(furthest_left_tab_index_simulation, state="hidden")
                    furthest_left_tab_index_simulation += 1
                    furthest_right_tab_index_simulation += 1
                    if furthest_right_tab_index_simulation == max_index_simulation:
                        tab_text = "+"
                    else:
                        tab_text = duree_simulation[furthest_right_tab_index_simulation]["nom_simulation"]
                    simulation_notebook.add(simulation_notebook.winfo_children()[furthest_right_tab_index_simulation],
                                            text=tab_text)
                    if furthest_right_tab_index_simulation == max_index_simulation:
                        scroll_right_button_simulation.configure(state="disabled")
                    else:
                        scroll_right_button_simulation.configure(state="normal")
                    if furthest_left_tab_index_simulation == min_index_simulation:
                        scroll_left_button_simulation.configure(state="disabled")
                    else:
                        scroll_left_button_simulation.configure(state="normal")

                def scroll_left_simulation():
                    global duree_simulation
                    global furthest_left_tab_index_simulation
                    global furthest_right_tab_index_simulation
                    simulation_notebook.tab(furthest_right_tab_index_simulation, state="hidden")
                    furthest_left_tab_index_simulation -= 1
                    furthest_right_tab_index_simulation -= 1
                    simulation_notebook.add(simulation_notebook.winfo_children()[furthest_left_tab_index_simulation],
                                            text=duree_simulation[furthest_left_tab_index_simulation]["nom_simulation"])
                    if furthest_right_tab_index_simulation == max_index_simulation:
                        scroll_right_button_simulation.configure(state="disabled")
                    else:
                        scroll_right_button_simulation.configure(state="normal")
                    if furthest_left_tab_index_simulation == min_index_simulation:
                        scroll_left_button_simulation.configure(state="disabled")
                    else:
                        scroll_left_button_simulation.configure(state="normal")

                def scroll_right_champs():
                    global information_champs
                    global furthest_left_tab_index_champs
                    global furthest_right_tab_index_champs
                    index_champs_frame = 0
                    for simulation_frame in simulation_notebook.winfo_children():
                        if index_champs_frame < len(simulation_notebook.winfo_children()) - 1:
                            champs_notebook = simulation_frame.winfo_children()[0]
                            champs_notebook.tab(furthest_left_tab_index_champs, state="hidden")
                        index_champs_frame += 1
                    furthest_left_tab_index_champs += 1
                    furthest_right_tab_index_champs += 1
                    if furthest_right_tab_index_champs == max_index_champs:
                        tab_text = "+"
                    else:
                        tab_text = information_champs[furthest_right_tab_index_champs]["nom_du_champs"]
                    index_champs_frame = 0
                    for simulation_frame in simulation_notebook.winfo_children():
                        if index_champs_frame < len(simulation_notebook.winfo_children()) - 1:
                            champs_notebook = simulation_frame.winfo_children()[0]
                            champs_notebook.add(champs_notebook.winfo_children()[furthest_right_tab_index_champs],
                                                text=tab_text)
                        index_champs_frame += 1
                    if furthest_right_tab_index_champs == max_index_champs:
                        scroll_right_button_champs.configure(state="disabled")
                    else:
                        scroll_right_button_champs.configure(state="normal")
                    if furthest_left_tab_index_champs == min_index_champs:
                        scroll_left_button_champs.configure(state="disabled")
                    else:
                        scroll_left_button_champs.configure(state="normal")

                def scroll_left_champs():
                    global information_champs
                    global furthest_left_tab_index_champs
                    global furthest_right_tab_index_champs
                    index_champs_frame = 0
                    for simulation_frame in simulation_notebook.winfo_children():
                        if index_champs_frame < len(simulation_notebook.winfo_children()) - 1:
                            champs_notebook = simulation_frame.winfo_children()[0]
                            champs_notebook.tab(furthest_right_tab_index_champs, state="hidden")
                        index_champs_frame += 1
                    furthest_left_tab_index_champs -= 1
                    furthest_right_tab_index_champs -= 1
                    index_champs_frame = 0
                    for simulation_frame in simulation_notebook.winfo_children():
                        if index_champs_frame < len(simulation_notebook.winfo_children()) - 1:
                            champs_notebook = simulation_frame.winfo_children()[0]
                            champs_notebook.add(champs_notebook.winfo_children()[furthest_left_tab_index_champs],
                                                text=information_champs[furthest_left_tab_index_champs][
                                                    "nom_du_champs"])
                        index_champs_frame += 1
                    if furthest_right_tab_index_champs == max_index_champs:
                        scroll_right_button_champs.configure(state="disabled")
                    else:
                        scroll_right_button_champs.configure(state="normal")
                    if furthest_left_tab_index_champs == min_index_champs:
                        scroll_left_button_champs.configure(state="disabled")
                    else:
                        scroll_left_button_champs.configure(state="normal")

                def scroll_right_zone():
                    global furthest_left_tab_index_zone
                    global furthest_right_tab_index_zone
                    champs_notebook = \
                        simulation_notebook.winfo_children()[simulation_notebook.index("current")].winfo_children()[0]
                    zone_notebook = champs_notebook.winfo_children()[champs_notebook.index("current")].winfo_children()[
                        0]
                    zone_notebook.tab(furthest_left_tab_index_zone, state="hidden")
                    furthest_left_tab_index_zone += 1
                    furthest_right_tab_index_zone += 1
                    if furthest_right_tab_index_zone == max_index_zone:
                        tab_text = "+"
                    else:
                        tab_text = "Zone gestion " + str(furthest_right_tab_index_zone + 1)
                    zone_notebook.add(zone_notebook.winfo_children()[furthest_right_tab_index_zone], text=tab_text)
                    if furthest_right_tab_index_zone == max_index_zone:
                        scroll_right_button_zone.configure(state="disabled")
                    else:
                        scroll_right_button_zone.configure(state="normal")
                    if furthest_left_tab_index_zone == min_index_zone:
                        scroll_left_button_zone.configure(state="disabled")
                    else:
                        scroll_left_button_zone.configure(state="normal")

                def scroll_left_zone():
                    global furthest_left_tab_index_zone
                    global furthest_right_tab_index_zone
                    champs_notebook = \
                        simulation_notebook.winfo_children()[simulation_notebook.index("current")].winfo_children()[0]
                    zone_notebook = champs_notebook.winfo_children()[champs_notebook.index("current")].winfo_children()[
                        0]
                    zone_notebook.tab(furthest_right_tab_index_zone, state="hidden")
                    furthest_left_tab_index_zone -= 1
                    furthest_right_tab_index_zone -= 1
                    zone_notebook.add(zone_notebook.winfo_children()[furthest_left_tab_index_zone],
                                      text="Zone gestion " + str(furthest_left_tab_index_zone + 1))
                    if furthest_right_tab_index_zone == max_index_zone:
                        scroll_right_button_zone.configure(state="disabled")
                    else:
                        scroll_right_button_zone.configure(state="normal")
                    if furthest_left_tab_index_zone == min_index_zone:
                        scroll_left_button_zone.configure(state="disabled")
                    else:
                        scroll_left_button_zone.configure(state="normal")

                def zone_tab_management(event):
                    global max_index_zone
                    global furthest_left_tab_index_zone
                    global furthest_right_tab_index_zone
                    furthest_left_tab_index_zone = 0
                    furthest_right_tab_index_zone = 0
                    current_simulation_index = simulation_notebook.index("current")
                    champs_notebook = simulation_notebook.winfo_children()[current_simulation_index].winfo_children()[0]
                    if champs_notebook.index("end") - 1 != champs_notebook.index("current"):
                        current_zone_notebook = \
                            champs_notebook.winfo_children()[champs_notebook.index("current")].winfo_children()[0]
                    else:
                        current_zone_notebook = ttk.Notebook()
                    for index in range(len(current_zone_notebook.winfo_children())):
                        current_zone_notebook.tab(index, state="hidden")
                    max_index_zone = len(current_zone_notebook.winfo_children()) - 1
                    if max_index_zone <= 3:
                        furthest_right_tab_index_zone = max_index_zone
                        scroll_left_button_zone.configure(state="disabled")
                        scroll_right_button_zone.configure(state="disabled")
                    else:
                        furthest_right_tab_index_zone = 3
                        scroll_left_button_zone.configure(state="disabled")
                        scroll_right_button_zone.configure(state="normal")
                    for index in range(len(current_zone_notebook.winfo_children())):
                        if index <= furthest_right_tab_index_zone:
                            current_zone_notebook.tab(index, state="normal")
                        else:
                            current_zone_notebook.tab(index, state="hidden")
                    if len(current_zone_notebook.winfo_children()) > 0:
                        current_zone_notebook.select(furthest_left_tab_index_zone)

                def get_regies_rechauffement():
                    entree_invalide_liste = []
                    regies_rechauffement_simulation = []
                    champs_label_frames = donnees_de_rechauffement_label_frame.winfo_children()[0].winfo_children()[
                        0].winfo_children()
                    champs_index = 0
                    for champs_label_frame in champs_label_frames:
                        champs_regies_liste = []
                        zone_label_frames = champs_label_frame.winfo_children()
                        zone_index = 0
                        for zone_label_frame in zone_label_frames:
                            regies_de_rechauffement = zone_label_frame.winfo_children()
                            index_regie = 0
                            zone_regies = []
                            number_of_label_frames = len(regies_de_rechauffement) - 1
                            while index_regie < number_of_label_frames:
                                regie = regies_de_rechauffement[index_regie]
                                culture_principale = regie.grid_slaves(row=0, column=1)[0].get()
                                rendement = regie.grid_slaves(row=1, column=1)[0].get()
                                pourcentage_tige_exporte = regie.grid_slaves(row=2, column=1)[0].get()
                                production_recolte = regie.grid_slaves(row=3, column=1)[0].get()
                                pourcentage_humidite = regie.grid_slaves(row=4, column=1)[0].get()
                                travail_du_sol = regie.grid_slaves(row=5, column=1)[0].get()
                                culture_secondaire = regie.grid_slaves(row=7, column=1)[0].get()
                                rendement_culture_secondaire = regie.grid_slaves(row=8, column=1)[0].get()
                                regie_vide = {"partie_culture": False, "partie_amendement": False}
                                if culture_principale == "" and rendement == "" and pourcentage_tige_exporte == "" and production_recolte == "" and pourcentage_humidite == "" and travail_du_sol == "" and culture_secondaire == "" and rendement_culture_secondaire == "":
                                    regie_vide["partie_culture"] = True
                                else:
                                    global cultures_principales_supportees
                                    if culture_principale not in cultures_principales_supportees:
                                        entree_invalide_liste.append(
                                            (information_champs[champs_index]["nom_du_champs"],
                                             "Zone gestion " + str(zone_index + 1),
                                             "\"Culture principale\" doit être parmis les choix disponibles",
                                             " section Données réchauffement"))
                                    if not util.is_decimal_number(rendement) and rendement != "":
                                        entree_invalide_liste.append(
                                            (information_champs[champs_index]["nom_du_champs"],
                                             "Zone gestion " + str(zone_index + 1),
                                             "\"Rendement\" doit être un réel positif ou la case peut être vide pour aller chercher un rendement par défaut",
                                             " section Données réchauffement"))
                                    if rendement == "":
                                        rendement = None
                                    if rendement is not None and util.is_decimal_number(rendement):
                                        rendement = float(rendement)
                                    if (not util.is_decimal_number(
                                            pourcentage_tige_exporte) and pourcentage_tige_exporte != "") or (
                                            util.is_decimal_number(pourcentage_tige_exporte) and (
                                            float(pourcentage_tige_exporte) < 0 or float(
                                        pourcentage_tige_exporte) > 100)):
                                        entree_invalide_liste.append(
                                            (information_champs[champs_index]["nom_du_champs"],
                                             "Zone gestion " + str(zone_index + 1),
                                             "\"Proportion tige exportée\" doit être un réel positif dans l'intervalle [0,1] ou le champs peut être vide pour aller chercher une proportion par défaut",
                                             "section Données réchauffement"))
                                    if pourcentage_tige_exporte == "":
                                        pourcentage_tige_exporte = None
                                    if pourcentage_tige_exporte is not None and util.is_decimal_number(
                                            pourcentage_tige_exporte):
                                        pourcentage_tige_exporte = float(pourcentage_tige_exporte)
                                    if production_recolte not in ["Oui", "Non"]:
                                        entree_invalide_liste.append(
                                            (information_champs[champs_index]["nom_du_champs"],
                                             "Zone gestion " + str(zone_index + 1),
                                             "\"Production récolté\" doit être parmis les choix disponibles",
                                             "sectionDonnées réchauffement"))
                                    else:
                                        if production_recolte == "Oui":
                                            production_recolte = True
                                        else:
                                            production_recolte = False
                                    if (not util.is_decimal_number(
                                            pourcentage_humidite) and pourcentage_humidite != "") or (
                                            util.is_decimal_number(pourcentage_humidite) and (
                                            float(pourcentage_humidite) < 0 or float(pourcentage_humidite) > 100)):
                                        entree_invalide_liste.append(
                                            (information_champs[champs_index]["nom_du_champs"],
                                             "Zone gestion " + str(zone_index + 1),
                                             "\"Pourcentage d'humidité\" doit être un réel positif dans l'intervalle [0,100] ou le champs peut être vide pour aller chercher une proportion par défaut",
                                             "section Données réchauffement"))
                                    if pourcentage_humidite == "":
                                        pourcentage_humidite = None
                                    if pourcentage_humidite is not None and util.is_decimal_number(
                                            pourcentage_humidite):
                                        pourcentage_humidite = float(pourcentage_humidite)
                                    culture_principale_dict = {"culture_principale": culture_principale,
                                                               "rendement": rendement,
                                                               "pourcentage_tige_exporte": pourcentage_tige_exporte,
                                                               "produit_recolte": production_recolte,
                                                               "pourcentage_humidite": pourcentage_humidite}
                                    global types_travail_du_sol_supportes
                                    if travail_du_sol not in types_travail_du_sol_supportes:
                                        entree_invalide_liste.append(
                                            (information_champs[champs_index]["nom_du_champs"],
                                             "Zone gestion " + str(zone_index + 1),
                                             "\"Travail du sol\" doit être parmis les choix disponibles",
                                             " section Données réchauffement"))
                                    travail_du_sol_dict = {"travail_du_sol": travail_du_sol}
                                    global cultures_secondaires_supportees
                                    if culture_secondaire == "":
                                        culture_secondaire = None
                                    if culture_secondaire is not None and culture_secondaire not in cultures_secondaires_supportees:
                                        entree_invalide_liste.append(
                                            (information_champs[champs_index]["nom_du_champs"],
                                             "Zone gestion " + str(zone_index + 1),
                                             "\"Culture secondaire\" doit être parmis les choix disponibles ou laissé vide s'il n'y a pas de culture secondaire",
                                             " section Données réchauffement"))
                                    if rendement_culture_secondaire == "":
                                        rendement_culture_secondaire = None
                                    if rendement_culture_secondaire is not None and not util.is_decimal_number(
                                            rendement_culture_secondaire):
                                        entree_invalide_liste.append(
                                            (information_champs[champs_index]["nom_du_champs"],
                                             "Zone gestion " + str(zone_index + 1),
                                             "\"Rendement culture secondaire\" doit être un réel positif ou laissé vide s'il n'y a pas de culture secondaire",
                                             "section Données réchauffement"))
                                    elif (culture_secondaire is None and rendement_culture_secondaire is not None) or (
                                            culture_secondaire is not None and rendement_culture_secondaire is None):
                                        entree_invalide_liste.append(
                                            (information_champs[champs_index]["nom_du_champs"],
                                             "Zone gestion " + str(zone_index + 1),
                                             "\"Rendement culture secondaire\" et \"Culture secondaire\" doivent être tout deux laissé vide s'il n'y a pas de culture secondaire",
                                             "section Données réchauffement"))
                                    else:
                                        if rendement_culture_secondaire is not None:
                                            rendement_culture_secondaire = float(rendement_culture_secondaire)
                                    culture_secondaire_dict = {"culture_secondaire": culture_secondaire,
                                                               "rendement": rendement_culture_secondaire}
                                index_composante_amendement = 0
                                composante_amendement_liste = regie.grid_slaves(row=9, column=0)[0]
                                grid_size = composante_amendement_liste.grid_size()
                                amendement_non_vide_compteur = 0
                                amendements = []
                                while index_composante_amendement < grid_size[1] - 1:
                                    amendement = \
                                        composante_amendement_liste.grid_slaves([index_composante_amendement],
                                                                                column=1)[
                                            0].get()
                                    global amendements_supportees
                                    if amendement == "":
                                        amendement = None
                                    if amendement is not None and amendement not in amendements_supportees:
                                        entree_invalide_liste.append(
                                            (information_champs[champs_index]["nom_du_champs"],
                                             "Zone gestion " + str(zone_index + 1),
                                             "\"Amendement\" " + str(
                                                 index_composante_amendement + 1) + " doit être parmis les choix disponibles",
                                             "section Données réchauffement"))
                                    apport = \
                                        composante_amendement_liste.grid_slaves([index_composante_amendement + 1],
                                                                                column=1)[
                                            0].get()
                                    if apport == "":
                                        apport = None
                                    if apport is not None and not util.is_decimal_number(apport):
                                        entree_invalide_liste.append(
                                            (information_champs[champs_index]["nom_du_champs"],
                                             "Zone gestion " + str(zone_index + 1),
                                             "\"Apport\" " + str(index_composante_amendement + 1) +
                                             " est invalide, il doit être un réel positif ou laissé vide s'il n'y a pas d'amendements",
                                             "section Données réchauffement"))
                                    elif (amendement is None and apport is not None) or (
                                            amendement is not None and apport is None):
                                        entree_invalide_liste.append(
                                            (information_champs[champs_index]["nom_du_champs"],
                                             "Zone gestion " + str(zone_index + 1),
                                             "\"Apport\" " + str(
                                                 index_composante_amendement + 1) + " et \"Amendement\" " + str(
                                                 index_composante_amendement + 1) +
                                             " doivent être tout deux laissé vide s'il n'y a pas d'amendements",
                                             "section Données réchauffement"))
                                    else:
                                        if apport is not None:
                                            apport = float(apport)
                                    if amendement is not None or apport is not None:
                                        amendement_non_vide_compteur += 1

                                    amendements.append({"amendement": amendement,
                                                        "apport": apport})
                                    index_composante_amendement += 2
                                if amendement_non_vide_compteur > 0:
                                    regie_vide["partie_amendement"] = False
                                else:
                                    regie_vide["partie_amendement"] = True
                                if regie_vide["partie_culture"] and regie_vide["partie_amendement"]:
                                    pass
                                else:
                                    regie_dict = {"culture_principale": culture_principale_dict,
                                                  "culture_secondaire": culture_secondaire_dict,
                                                  "amendements": amendements,
                                                  "travail_du_sol": travail_du_sol_dict}
                                    zone_regies.append(regie_dict)
                                index_regie += 1
                            champs_regies_liste.append(zone_regies)
                            zone_index += 1
                        regies_rechauffement_simulation.append(champs_regies_liste)
                        champs_index += 1
                    return regies_rechauffement_simulation, entree_invalide_liste

                def editer_caracteristique_physique_entreprise():
                    def fenetre_edition_ferme():
                        root.deiconify()
                        edition_window.destroy()

                    root.withdraw()
                    edition_window = tk.Toplevel()
                    edition_window.resizable(0, 0)
                    edition_window.protocol("WM_DELETE_WINDOW", fenetre_edition_ferme)
                    edition_frame = ttk.Frame(edition_window)
                    canvas = tk.Canvas(edition_frame)
                    scrollbar = ttk.Scrollbar(edition_frame, orient="vertical", command=canvas.yview)
                    scrollable_frame = ttk.Frame(canvas)
                    scrollable_frame.bind("<Configure>", lambda e: canvas.configure(scrollregion=canvas.bbox("all")))
                    canvas.create_window((0, 0), window=scrollable_frame, anchor="nw")
                    canvas.configure(yscrollcommand=scrollbar.set)
                    entreprise_label_frame = ttk.LabelFrame(scrollable_frame, text="Entreprise")
                    nom_entreprise_label = ttk.Label(entreprise_label_frame, text="Nom de l'entreprise: ")
                    nom_entreprise_entry = ttk.Entry(entreprise_label_frame)
                    global nom_entreprise
                    nom_entreprise_entry.insert(0, nom_entreprise)
                    nom_entreprise_label.grid(row=0, column=0, sticky="w", pady=3)
                    nom_entreprise_entry.grid(row=0, column=1, sticky="w", pady=3)
                    champs_row_index = 1
                    global nombre_de_champs
                    for champs_index in range(nombre_de_champs):
                        champs_label_frame = ttk.LabelFrame(entreprise_label_frame,
                                                            text="Champs " + str(champs_index + 1),
                                                            name="champ" + str(champs_index))
                        global information_champs
                        nom_champs_label = ttk.Label(champs_label_frame, text="Nom du champ: ")
                        nom_champs_entry = ttk.Entry(champs_label_frame)
                        if information_champs[champs_index]["nom_du_champs"]:
                            nom_champs_entry.insert(0, information_champs[champs_index]["nom_du_champs"])
                        nom_champs_label.grid(row=0, column=0, sticky="w", pady=3)
                        nom_champs_entry.grid(row=0, column=1, sticky="w", pady=3)
                        zone_row_index = 1
                        information_zones_de_gestion = information_champs[champs_index]["information_zone_de_gestion"]
                        for zone_index in range(int(information_champs[champs_index]["nombre_de_zone_de_gestion"])):
                            information_zone_de_gestion = information_zones_de_gestion[zone_index]
                            zone_label_frame = ttk.LabelFrame(champs_label_frame,
                                                              text="Zone gestion " + str(zone_index + 1),
                                                              name="zonegestion" + str(zone_index))
                            taux_matiere_organique_label = ttk.Label(zone_label_frame,
                                                                     text="Taux matière organique (en %): ")
                            taux_matiere_organique_entry = ttk.Entry(zone_label_frame)
                            if information_zone_de_gestion["taux_matiere_organique"] is not None:
                                taux_matiere_organique_entry.insert(0,
                                                                    information_zone_de_gestion[
                                                                        "taux_matiere_organique"])
                            municipalite_label = ttk.Label(zone_label_frame, text="Municipalité: ")
                            municipalite_combobox = create_municipalite_combobox(information_zone_de_gestion,
                                                                                 zone_label_frame)
                            groupe_textural_label = ttk.Label(zone_label_frame, text="Groupe textural: ")
                            global groupes_texturaux_supportees
                            groupe_textural_combobox = ttk.Combobox(zone_label_frame,
                                                                    values=groupes_texturaux_supportees,
                                                                    postcommand=lambda: filter_combobox_values(
                                                                        groupe_textural_combobox,
                                                                        groupes_texturaux_supportees))
                            groupe_textural_combobox.set(information_zone_de_gestion["groupe_textural"])
                            classe_de_drainage_label = ttk.Label(zone_label_frame, text="Classe de drainage: ")
                            global classes_de_drainage_supportees
                            classe_de_drainage_combobox = ttk.Combobox(zone_label_frame,
                                                                       values=classes_de_drainage_supportees,
                                                                       postcommand=lambda: filter_combobox_values(
                                                                           classe_de_drainage_combobox,
                                                                           classes_de_drainage_supportees))
                            classe_de_drainage_combobox.set(information_zone_de_gestion["classe_de_drainage"])
                            masse_volumique_apparente_label = ttk.Label(zone_label_frame,
                                                                        text="Masse volumique apparente (g/cm3): ")
                            masse_volumique_apparente_entry = ttk.Entry(zone_label_frame)
                            if information_zone_de_gestion["masse_volumique_apparente"] is not None:
                                masse_volumique_apparente_entry.insert(0,
                                                                       information_zone_de_gestion[
                                                                           "masse_volumique_apparente"])
                            profondeur_label = ttk.Label(zone_label_frame, text="Profondeur de la couche arable (cm): ")
                            profondeur_entry = ttk.Entry(zone_label_frame)
                            if information_zone_de_gestion["profondeur"] is not None:
                                profondeur_entry.insert(0, information_zone_de_gestion["profondeur"])
                            superficie_de_la_zone_label = ttk.Label(zone_label_frame,
                                                                    text="Superficie de la zone (ha): ")
                            superficie_de_la_zone_entry = ttk.Entry(zone_label_frame)
                            if information_zone_de_gestion["superficie_de_la_zone"] is not None:
                                superficie_de_la_zone_entry.insert(0,
                                                                   information_zone_de_gestion["superficie_de_la_zone"])
                            taux_matiere_organique_label.grid(row=0, column=0, sticky="w", pady=3)
                            taux_matiere_organique_entry.grid(row=0, column=1, sticky="w", pady=3)
                            municipalite_label.grid(row=1, column=0, sticky="w", pady=3)
                            municipalite_combobox.grid(row=1, column=1, sticky="w", pady=3)
                            groupe_textural_label.grid(row=2, column=0, sticky="w", pady=3)
                            groupe_textural_combobox.grid(row=2, column=1, sticky="w", pady=3)
                            classe_de_drainage_label.grid(row=3, column=0, sticky="w", pady=3)
                            classe_de_drainage_combobox.grid(row=3, column=1, sticky="w", pady=3)
                            masse_volumique_apparente_label.grid(row=4, column=0, sticky="w", pady=3)
                            masse_volumique_apparente_entry.grid(row=4, column=1, sticky="w", pady=3)
                            profondeur_label.grid(row=5, column=0, sticky="w", pady=3)
                            profondeur_entry.grid(row=5, column=1, sticky="w", pady=3)
                            superficie_de_la_zone_label.grid(row=6, column=0, sticky="w", pady=3)
                            superficie_de_la_zone_entry.grid(row=6, column=1, sticky="w", pady=3)
                            zone_label_frame.grid(row=zone_row_index, column=0, columnspan=2, padx=10)
                            zone_row_index += 1
                        champs_label_frame.grid(row=champs_row_index, column=0, columnspan=2, padx=10)
                        champs_row_index += 1
                    entreprise_label_frame.pack(padx=10, pady=10)
                    canvas.pack(side="left", ipadx=5)
                    scrollbar.pack(side="right", fill="y")

                    def effectuer_la_sauvegarde():
                        entree_invalide_liste = []
                        global nom_entreprise
                        nom_entreprise = nom_entreprise_entry.get()
                        global information_champs
                        information_champs = []
                        champs_label_frame_index = 2
                        entreprise_widgets = entreprise_label_frame.winfo_children()
                        while champs_label_frame_index < len(entreprise_widgets):
                            champs_frame = entreprise_widgets[champs_label_frame_index]
                            nom_du_champs = champs_frame.winfo_children()[1].get()
                            if len(nom_du_champs) > 12:
                                entree_invalide_liste.append(("Champs " + str(champs_label_frame_index - 1),
                                                              "Le nom du champ devrait être composé de 12 caractères ou moins."))
                            champs_widgets = champs_frame.winfo_children()
                            zone_label_frame_index = 2
                            info_zones_de_gestion = []
                            while zone_label_frame_index < len(champs_widgets):
                                zone_frame = champs_widgets[zone_label_frame_index]
                                taux_matiere_organique = zone_frame.grid_slaves(row=0, column=1)[0].get()
                                if not util.is_decimal_number(taux_matiere_organique) or float(
                                        taux_matiere_organique) < 0 or float(taux_matiere_organique) > 100:
                                    entree_invalide_liste.append(
                                        ("Champs " + str(champs_label_frame_index - 1),
                                         "Zone gestion " + str(zone_label_frame_index - 1),
                                         "\"Taux de matière organique\" doit être un réel positif dans l'intervalle [0-100]"))
                                else:
                                    taux_matiere_organique = float(taux_matiere_organique)
                                municipalite = zone_frame.grid_slaves(row=1, column=1)[0].get()
                                global municipalites_supportees
                                if municipalite not in municipalites_supportees:
                                    entree_invalide_liste.append(
                                        ("Champs " + str(champs_label_frame_index - 1),
                                         "Zone gestion " + str(zone_label_frame_index - 1),
                                         "\"Municipalité\" doit être parmis les choix disponibles"))
                                groupe_textural = zone_frame.grid_slaves(row=2, column=1)[0].get()
                                global groupes_texturaux_supportees
                                if groupe_textural not in groupes_texturaux_supportees:
                                    entree_invalide_liste.append(
                                        ("Champs " + str(champs_label_frame_index - 1),
                                         "Zone gestion " + str(zone_label_frame_index - 1),
                                         "\"Groupe textural\" doit être parmis les choix disponibles"))
                                classe_de_drainage = zone_frame.grid_slaves(row=3, column=1)[0].get()
                                global classes_de_drainage_supportees
                                if classe_de_drainage not in classes_de_drainage_supportees:
                                    entree_invalide_liste.append(
                                        ("Champs " + str(champs_label_frame_index - 1),
                                         "Zone gestion " + str(zone_label_frame_index - 1),
                                         "\"Classe de drainage\" doit être parmis les choix disponibles"))
                                masse_volumique_apparente = zone_frame.grid_slaves(row=4, column=1)[0].get()
                                if (not util.is_decimal_number(
                                        masse_volumique_apparente) and masse_volumique_apparente != "") or (
                                        util.is_decimal_number(masse_volumique_apparente) and float(
                                    masse_volumique_apparente) < 0):
                                    entree_invalide_liste.append(
                                        ("Champs " + str(champs_label_frame_index - 1),
                                         "Zone gestion " + str(zone_label_frame_index - 1),
                                         "\"Masse volumique apparente\" doit être un réel positif ou laissé vide pour aller chercher la valeur par défaut"))
                                else:
                                    if masse_volumique_apparente == "":
                                        masse_volumique_apparente = None
                                    else:
                                        masse_volumique_apparente = float(masse_volumique_apparente)
                                profondeur = zone_frame.grid_slaves(row=5, column=1)[0].get()
                                if not util.is_decimal_number(profondeur) or float(
                                        profondeur) < 0:
                                    entree_invalide_liste.append(
                                        ("Champs " + str(champs_label_frame_index - 1),
                                         "Zone gestion " + str(zone_label_frame_index - 1),
                                         "\"Profondeur de la couche arable\" doit être un réel positif"))
                                else:
                                    profondeur = float(profondeur)
                                superficie_de_la_zone = zone_frame.grid_slaves(row=6, column=1)[0].get()
                                if not util.is_decimal_number(superficie_de_la_zone) or float(
                                        superficie_de_la_zone) < 0:
                                    entree_invalide_liste.append(
                                        ("Champs " + str(champs_label_frame_index - 1),
                                         "Zone gestion " + str(zone_label_frame_index - 1),
                                         "\"Superficie de la zone\" doit être un réel positif"))
                                else:
                                    superficie_de_la_zone = float(superficie_de_la_zone)
                                info_zones_de_gestion.append(
                                    {"taux_matiere_organique": taux_matiere_organique,
                                     "municipalite": municipalite,
                                     "groupe_textural": groupe_textural,
                                     "classe_de_drainage": classe_de_drainage,
                                     "masse_volumique_apparente": masse_volumique_apparente,
                                     "profondeur": profondeur,
                                     "superficie_de_la_zone": superficie_de_la_zone})
                                zone_label_frame_index += 1
                            information_champs.append({"nom_du_champs": nom_du_champs,
                                                       "nombre_de_zone_de_gestion": str(len(champs_widgets) - 2),
                                                       "information_zone_de_gestion": info_zones_de_gestion})
                            champs_label_frame_index += 1

                        if len(entree_invalide_liste) == 0:
                            root.deiconify()
                            effectuer_les_modifications_a_la_fenetre_principale()
                            sauvegarder_attributs_entreprise_apres_modification()
                            edition_window.destroy()
                        else:
                            message = ""
                            for entree_invalide in entree_invalide_liste:
                                if len(entree_invalide) == 2:
                                    message = message + "Dans le " + entree_invalide[0] + ", " + entree_invalide[
                                        1] + "\n"
                                else:
                                    message = message + "Dans le " + entree_invalide[0] + " et la " + entree_invalide[
                                        1] + " l'entrée " + entree_invalide[2] + "\n"
                            messagebox.showwarning("Warning", message)
                            edition_window.focus()

                    def effectuer_les_modifications_a_la_fenetre_principale():
                        index_simulation_frame = 0
                        simulation_frames = simulation_notebook.winfo_children()
                        global information_champs
                        for simulation_frame in simulation_frames:
                            if index_simulation_frame < len(simulation_frames) - 1:
                                champs_notebook = simulation_frame.winfo_children()[0]
                                index_champs_frame = 0
                                while index_champs_frame < len(champs_notebook.winfo_children()) - 1:
                                    champs_notebook.tab(index_champs_frame,
                                                        text=information_champs[index_champs_frame]["nom_du_champs"])
                                    index_champs_frame += 1

                            index_simulation_frame += 1
                        rechauffement_frame = donnees_de_rechauffement_label_frame.winfo_children()[0].winfo_children()[
                            0]
                        index = 0
                        for champs in rechauffement_frame.winfo_children():
                            champs.configure(text=information_champs[index]["nom_du_champs"])
                            index += 1

                    sauvegarde_des_modifications_button = ttk.Button(scrollable_frame, text="Sauvegarder",
                                                                     command=effectuer_la_sauvegarde)
                    sauvegarde_des_modifications_button.pack()

                    edition_frame.pack()

                def create_municipalite_combobox(information_zone_de_gestion, zone_label_frame):
                    global municipalites_supportees
                    municipalite_combobox = ttk.Combobox(zone_label_frame, values=municipalites_supportees,
                                                         postcommand=lambda: filter_combobox_values(
                                                             municipalite_combobox, municipalites_supportees))
                    municipalite_combobox.set(information_zone_de_gestion["municipalite"])
                    return municipalite_combobox

                for widget in parent_frame_tabs.winfo_children():
                    widget.destroy()

                rechauffement_frame = ttk.Frame(parent_frame_tabs)
                projection_frame = ttk.Frame(parent_frame_tabs, width=575)

                rechauffement_frame.grid(row=0, column=1)
                projection_frame.grid(row=0, column=0)
                projection_frame.grid_propagate(True)

                donnees_de_rechauffement_label_frame = ttk.LabelFrame(rechauffement_frame,
                                                                      text="Régies des sols et cultures historiques (Facultatif)")
                simulation_notebook = ttk.Notebook(projection_frame)

                global nombre_simulations
                if nombre_simulations != 0:
                    pass
                else:
                    nombre_simulations = 0

                global duree_simulation
                if len(duree_simulation) != 0:
                    pass
                else:
                    duree_simulation = []

                tab1 = ttk.Frame(simulation_notebook)

                simulation_notebook.bind("<Button-1>", add_new_simulation_tab)
                simulation_notebook.bind("<Button-3>", delete_simulation_tab)

                simulation_notebook.add(tab1, text="+")

                scroll_right_frame = ttk.Frame(projection_frame)
                scroll_left_frame = ttk.Frame(projection_frame)

                scroll_right_button_simulation = ttk.Button(scroll_right_frame, text="\u25b6")
                scroll_left_button_simulation = ttk.Button(scroll_left_frame, text="\u25c0")

                scroll_right_button_simulation.configure(command=scroll_right_simulation, state="disabled")
                scroll_left_button_simulation.configure(command=scroll_left_simulation, state="disabled")

                scroll_right_button_simulation.grid(row=0, column=0, sticky="n")
                scroll_left_button_simulation.grid(row=0, column=0, sticky="n")

                scroll_right_button_champs = ttk.Button(scroll_right_frame, text="\u25b6")
                scroll_left_button_champs = ttk.Button(scroll_left_frame, text="\u25c0")

                scroll_right_button_champs.configure(command=scroll_right_champs, state="disabled")
                scroll_left_button_champs.configure(command=scroll_left_champs, state="disabled")

                scroll_right_button_champs.grid(row=1, column=0, sticky="n")
                scroll_left_button_champs.grid(row=1, column=0, sticky="n")

                scroll_right_button_zone = ttk.Button(scroll_right_frame, text="\u25b6")
                scroll_left_button_zone = ttk.Button(scroll_left_frame, text="\u25c0")

                scroll_right_button_zone.configure(command=scroll_right_zone, state="disabled")
                scroll_left_button_zone.configure(command=scroll_left_zone, state="disabled")

                scroll_right_button_zone.grid(row=2, column=0, sticky="n")
                scroll_left_button_zone.grid(row=2, column=0, sticky="n")

                scroll_right_frame.grid(row=0, column=2, sticky="n")
                scroll_left_frame.grid(row=0, column=0, sticky="n")

                simulation_notebook.grid(row=0, column=1, sticky="nw")

                set_up_regies_rechauffement(donnees_de_rechauffement_label_frame, simulations)
                donnees_de_rechauffement_label_frame.grid(row=0, column=0, columnspan=2, pady=3)
                sauvegarder_plan_de_gestion_button = ttk.Button(rechauffement_frame,
                                                                text="Sauvegarder le plan de gestion",
                                                                command=lambda: sauvegarder_plan_de_gestion_de_carbone(
                                                                    sauvegarder_toutes_les_simulations()))
                retour_au_menu_principal_button = ttk.Button(rechauffement_frame, text="Retour au menu principal",
                                                             command=retour_au_menu_principal)
                sauvegarder_plan_de_gestion_button.grid(row=1, column=0, pady=3)
                retour_au_menu_principal_button.grid(row=1, column=1, pady=3)

                if simulations is not None:
                    simulation_notebook.winfo_children()[simulation_notebook.index("end") - 1].destroy()
                    index_simulation = 0
                    for simulation in simulations:
                        if index_simulation == 0:
                            set_up_simulation(simulation_notebook, simulation, True)
                        else:
                            set_up_simulation(simulation_notebook, simulation, False)
                        if index_simulation != len(simulations) - 1:
                            simulation_notebook.winfo_children()[simulation_notebook.index("end") - 1].destroy()
                        index_simulation += 1
                    global simulations_chargees
                    simulations_chargees = False

            def menu_initial_ogemos(menu_frame):
                def creation_nouvelle_entreprise():
                    for widget in menu_frame.winfo_children():
                        widget.destroy()
                    show_entreprise_set_up(menu_frame)

                def charger_entreprise():
                    for widget in menu_frame.winfo_children():
                        widget.destroy()
                    root.withdraw()
                    global filename
                    filename = filedialog.askopenfilename(initialdir="/", title="Select file",
                                                          filetypes=(("json files", "*.json"), ("all files", "*.*")))
                    if filename == "":
                        menu_initial_ogemos(menu_frame)
                        root.deiconify()
                        return
                    root.deiconify()

                    no_error = True
                    try:
                        with open(filename) as file:
                            data = json.load(file)
                        global nom_entreprise
                        global nombre_de_champs
                        global information_champs
                        nom_entreprise = data["nom_entreprise"]
                        nombre_de_champs = data["nombre_de_champs"]
                        information_champs = data["information_champs"]
                    except KeyError:
                        root.withdraw()
                        no_error = False
                        messagebox.showwarning("Fichier Invalide", "Le fichier ne correspondant pas au format désiré.")
                        root.deiconify()
                        menu_initial_ogemos(menu_frame)

                    if no_error:
                        show_creation_des_regies(menu_frame)

                def charger_plan_de_gestion_de_carbone():
                    global simulations_chargees
                    simulations_chargees = True
                    for widget in menu_frame.winfo_children():
                        widget.destroy()
                    root.withdraw()
                    global plan_gestion_filename
                    plan_gestion_filename = filedialog.askopenfilename(initialdir="/", title="Select file",
                                                                       filetypes=(
                                                                           ("json files", "*.json"),
                                                                           ("all files", "*.*")))
                    if plan_gestion_filename == "":
                        simulations_chargees = False
                        menu_initial_ogemos(menu_frame)
                        root.deiconify()
                        return
                    root.deiconify()

                    no_error = True
                    try:
                        with open(plan_gestion_filename) as file:
                            data = json.load(file)
                        global nom_entreprise
                        global nombre_de_champs
                        global information_champs
                        global duree_simulation
                        global nombre_simulations
                        information_champs = []
                        duree_simulation = []
                        simulations = data["simulations"]
                        nombre_simulations = len(simulations)
                        for simulation in simulations:
                            duree_simulation.append(
                                {"annee_projection_initiale": simulation["annee_initiale_projection"],
                                 "duree_projection": simulation["duree_projection"],
                                 "nom_simulation": simulation["nom_simulation"]})
                        if len(simulations) > 0:
                            entreprise = simulations[0]["entreprise_agricole"]
                            nom_entreprise = entreprise["nom"]
                            nombre_de_champs = len(entreprise["champs"])
                            if nombre_de_champs > 0:
                                champs_liste = entreprise["champs"]
                                for champs in champs_liste:
                                    if len(champs["zones_de_gestion"]) > 0:
                                        information_zone_de_gestion = []
                                        zone_de_gestion_liste = champs["zones_de_gestion"]
                                        for zone_de_gestion in zone_de_gestion_liste:
                                            information_zone_de_gestion.append(
                                                {"taux_matiere_organique": zone_de_gestion["taux_matiere_organique"],
                                                 "municipalite": zone_de_gestion["municipalite"],
                                                 "groupe_textural": zone_de_gestion["groupe_textural"],
                                                 "classe_de_drainage": zone_de_gestion["classe_de_drainage"],
                                                 "masse_volumique_apparente": zone_de_gestion[
                                                     "masse_volumique_apparente"],
                                                 "profondeur": zone_de_gestion["profondeur"],
                                                 "superficie_de_la_zone": zone_de_gestion["superficie_de_la_zone"]})
                                        information_champs.append({"nom_du_champs": champs["nom"],
                                                                   "nombre_de_zone_de_gestion": len(
                                                                       champs["zones_de_gestion"]),
                                                                   "information_zone_de_gestion": information_zone_de_gestion})
                    except KeyError:
                        root.withdraw()
                        no_error = False
                        messagebox.showwarning("Fichier Invalide", "Le fichier ne correspondant pas au format désiré.")
                        root.deiconify()
                        simulations_chargees = False
                        menu_initial_ogemos(menu_frame)

                    if no_error:
                        show_creation_des_regies(menu_frame, simulations=simulations)

                def ajouter_un_nouvel_amendement(data=None):
                    def fenetre_nouvel_amendement_ferme():
                        root.deiconify()
                        nouvel_amendement_window.destroy()

                    root.withdraw()

                    def get_nouvel_amendement():
                        entree_invalide_liste = []
                        amendement = amendement_entry.get()
                        pourcentage_humidite = pourcentage_humidite_entry.get()
                        carbon_total = carbon_total_entry.get()
                        if not amendement.isalnum():
                            espace_seul_caractere_non_alphanum = True
                            amendement_parties = amendement.split(" ")
                            for partie in amendement_parties:
                                if not partie.isalpha():
                                    espace_seul_caractere_non_alphanum = False
                            if not espace_seul_caractere_non_alphanum:
                                entree_invalide_liste.append(
                                    "L'amendement doit être composé uniquement de caractères alphanumériques et d'espaces.")
                        global amendements_supportees
                        if amendement.isalnum() and amendement in amendements_supportees:
                            entree_invalide_liste.append(
                                "Un amendement avec un tel nom existe déjà.")
                        if not util.is_decimal_number(pourcentage_humidite):
                            entree_invalide_liste.append(
                                "Le pourcentage d'humidité doit être un réel positif dans l'intervalle [0-100].")
                        else:
                            if not (0 <= float(pourcentage_humidite) <= 100):
                                entree_invalide_liste.append(
                                    "Le pourcentage d'humidité doit être un réel positif dans l'intervalle [0-100].")
                        if not util.is_decimal_number(carbon_total):
                            entree_invalide_liste.append("Le total d'azote doit être un réel positif.")
                        if len(entree_invalide_liste) > 0:
                            message = ""
                            for entree in entree_invalide_liste:
                                message = message + entree + "\n"
                            messagebox.showwarning("Entrée invalide", message)
                        else:
                            response = requests.post('http://localhost:5000/api/ajout-amendement',
                                                     json={"amendement": amendement,
                                                           "pourcentage_humidite": pourcentage_humidite,
                                                           "carbon_total": carbon_total})
                            if response.status_code == 200:
                                messagebox.showinfo("Ajout amendement", "L'ajout d'un amendement a été un succès.")
                                amendements_supportees_temp = requests.get("http://localhost:5000/api/get-amendement")
                                amendements_supportees = amendements_supportees_temp.json()["amendements_supportees"]
                                amendements_supportees.sort()
                                nouvel_amendement_window.destroy()
                                root.deiconify()
                            else:
                                messagebox.showinfo("Ajout amendement", "L'ajout d'amendement a échoué")
                                nouvel_amendement_window.destroy()
                                ajouter_un_nouvel_amendement({"amendement": amendement,
                                                              "pourcentage_humidite": pourcentage_humidite,
                                                              "carbon_total": carbon_total})

                    nouvel_amendement_window = tk.Toplevel()
                    nouvel_amendement_window.resizable(0, 0)
                    nouvel_amendement_frame = ttk.Frame(nouvel_amendement_window)
                    nouvel_amendement_window.protocol("WM_DELETE_WINDOW", fenetre_nouvel_amendement_ferme)
                    amendement_label = ttk.Label(nouvel_amendement_frame, text="Amendement: ")
                    amendement_entry = ttk.Entry(nouvel_amendement_frame)
                    pourcentage_humidite_label = ttk.Label(nouvel_amendement_frame,
                                                           text="Pourcentage d'humidité [0-100]: ")
                    pourcentage_humidite_entry = ttk.Entry(nouvel_amendement_frame)
                    carbon_total_label = ttk.Label(nouvel_amendement_frame, text="Carbone total (kg/t): ")
                    carbon_total_entry = ttk.Entry(nouvel_amendement_frame)
                    ajout_amendement_button = ttk.Button(nouvel_amendement_frame, text="Ajouter l'amendement",
                                                         command=get_nouvel_amendement)
                    if data is not None:
                        amendement_entry.insert(0, data["amendement"])
                        pourcentage_humidite_entry.insert(0, data["pourcentage_humidite"])
                        carbon_total_entry.insert(0, data["carbon_total"])
                    amendement_label.grid(row=0, column=0, pady=3, padx=5)
                    amendement_entry.grid(row=0, column=1, pady=3, padx=5)
                    pourcentage_humidite_label.grid(row=1, column=0, pady=3, padx=5)
                    pourcentage_humidite_entry.grid(row=1, column=1, pady=3, padx=5)
                    carbon_total_label.grid(row=2, column=0, pady=3, padx=5)
                    carbon_total_entry.grid(row=2, column=1, pady=3, padx=5)
                    ajout_amendement_button.grid(row=3, column=0, columnspan=2, pady=3, padx=5)
                    nouvel_amendement_frame.pack()

                def menu_transfert_version():

                    def fenetre_menu_transfert_ferme():
                        root.deiconify()
                        menu_transfert_window.destroy()

                    def sauvegarder_amendement_dans_un_fichier():
                        menu_transfert_window.withdraw()
                        amendements_ajoutes_filename = filedialog.asksaveasfilename(initialdir="/",
                                                                                    title="File Explorer",
                                                                                    filetypes=(("json files", "*.json"),
                                                                                               ("all files", "*.*")))
                        if amendements_ajoutes_filename == "":
                            menu_transfert_window.deiconify()
                            return
                        if ".json" not in amendements_ajoutes_filename:
                            amendements_ajoutes_filename = amendements_ajoutes_filename + ".json"
                        amendements_ajoutes = requests.get("http://localhost:5000/api/get-amendement-ajoute").json()
                        with open(amendements_ajoutes_filename, 'w') as json_file:
                            json.dump(amendements_ajoutes, json_file)
                        menu_transfert_window.deiconify()

                    def charger_amendement_a_partir_de_fichier():
                        menu_transfert_window.withdraw()
                        amendements_ajoutes_filename = filedialog.askopenfilename(initialdir="/", title="Select file",
                                                                                  filetypes=(
                                                                                      ("json files", "*.json"),
                                                                                      ("all files", "*.*")))
                        if amendements_ajoutes_filename == "":
                            menu_transfert_window.deiconify()
                            return

                        with open(amendements_ajoutes_filename) as file:
                            data = json.load(file)

                        try:
                            amendements_non_present = []
                            for amendement in data["amendements_ajoutes"]:
                                if amendement["amendement"] not in amendements_supportees:
                                    amendements_non_present.append(amendement)

                            data = {"amendements_ajoutes": amendements_non_present}

                            if len(data["amendements_ajoutes"]) > 0:
                                response = requests.post('http://localhost:5000/api/post-amendement-ajoute', json=data)

                                if response.status_code == 200:
                                    messagebox.showinfo(title="Chargement des amendements",
                                                        message="Le chargement a été un succès!")
                                    menu_transfert_window.deiconify()
                                else:
                                    messagebox.showinfo(title="Chargement des amendements",
                                                        message="Le fichier de sauvegarde a été corrompu!")
                                    menu_transfert_window.deiconify()
                            else:
                                messagebox.showinfo(title="Chargement des amendements",
                                                    message="Les amendements de ce fichier ont déjà été chargé!")
                                menu_transfert_window.deiconify()

                            for amendement in amendements_non_present:
                                amendements_supportees.append(amendement["amendement"])
                        except KeyError:
                            messagebox.showwarning("Chargment amendements", "Le fichier n'a pas le format désiré")
                            menu_transfert_window.deiconify()

                    root.withdraw()
                    menu_transfert_window = tk.Toplevel()
                    menu_transfert_window.resizable(0, 0)
                    menu_transfert_window.protocol("WM_DELETE_WINDOW", fenetre_menu_transfert_ferme)
                    menu_transfert_frame = ttk.Frame(menu_transfert_window)
                    menu_transfert_label = ttk.Label(menu_transfert_frame, text="Menu de transfert des amendements")
                    sauvegarder_fichier_transfert_button = ttk.Button(menu_transfert_frame,
                                                                      text="Sauvegarder amendements ajoutés",
                                                                      command=sauvegarder_amendement_dans_un_fichier)
                    charger_fichier_transfert_button = ttk.Button(menu_transfert_frame,
                                                                  text="Charger amendements ajoutés",
                                                                  command=charger_amendement_a_partir_de_fichier)
                    retour_menu_principal_button = ttk.Button(menu_transfert_frame, text="Retour au menu principal",
                                                              command=fenetre_menu_transfert_ferme)
                    menu_transfert_label.grid(row=0, column=0, pady=3, padx=10)
                    sauvegarder_fichier_transfert_button.grid(row=1, column=0, pady=3, padx=10)
                    charger_fichier_transfert_button.grid(row=2, column=0, pady=3, padx=10)
                    retour_menu_principal_button.grid(row=3, column=0, pady=3, padx=10)
                    menu_transfert_frame.pack()

                bienvenue_label = ttk.Label(menu_frame, text="Bienvenue dans OGEMOS!")
                nouvelle_entreprise_button = ttk.Button(menu_frame, text="Créer une nouvelle entreprise",
                                                        command=creation_nouvelle_entreprise)
                charger_entreprise_button = ttk.Button(menu_frame, text="Charger une entreprise",
                                                       command=charger_entreprise)
                charger_plan_de_gestion_de_carbone_button = ttk.Button(menu_frame,
                                                                       text="Charger un plan de gestion de carbone",
                                                                       command=charger_plan_de_gestion_de_carbone)
                ajouter_nouvel_amendement_button = ttk.Button(menu_frame, text="Ajouter un nouvel amendement",
                                                              command=ajouter_un_nouvel_amendement)
                menu_transfert_amendements_button = ttk.Button(menu_frame, text="Menu de transfert de version",
                                                               command=menu_transfert_version)
                bienvenue_label.pack(pady=5, padx=10)
                nouvelle_entreprise_button.pack(pady=5, padx=10)
                charger_entreprise_button.pack(pady=5, padx=10)
                charger_plan_de_gestion_de_carbone_button.pack(pady=5, padx=10)
                ajouter_nouvel_amendement_button.pack(pady=5, padx=10)
                menu_transfert_amendements_button.pack(pady=5, padx=10)

            def sauvegarder_attributs_entreprise_apres_creation():
                root.withdraw()
                global filename
                filename = filedialog.asksaveasfilename(initialdir="/", title="File Explorer",
                                                        filetypes=(("json files", "*.json"), ("all files", "*.*")))
                if filename == "":
                    root.deiconify()
                    return False
                if ".json" not in filename:
                    filename = filename + ".json"
                global nom_entreprise
                global nombre_de_champs
                global information_champs
                save_dict = {"nom_entreprise": nom_entreprise,
                             "nombre_de_champs": nombre_de_champs,
                             "information_champs": information_champs}

                with open(filename, 'w') as json_file:
                    json.dump(save_dict, json_file)
                root.deiconify()
                return True

            def sauvegarder_attributs_entreprise_apres_modification():
                global nom_entreprise
                global nombre_de_champs
                global information_champs
                save_dict = {"nom_entreprise": nom_entreprise,
                             "nombre_de_champs": nombre_de_champs,
                             "information_champs": information_champs}
                global filename
                if filename is None or filename == "":
                    response = messagebox.askyesno("Sauvegarde entreprise",
                                                   "Souhaitez-vous sauvegarder les modifications apportées à l'entreprise?")
                    if response:
                        filename_valide = sauvegarder_attributs_entreprise_apres_creation()
                        if filename_valide:
                            with open(filename, 'w') as json_file:
                                json.dump(save_dict, json_file)
                else:
                    with open(filename, 'w') as json_file:
                        json.dump(save_dict, json_file)

            def sauvegarder_plan_de_gestion_de_carbone(simulations):
                if isinstance(simulations, list):
                    entree_invalide_liste = simulations
                    message = ""
                    for entree_invalide in entree_invalide_liste:
                        message = message + "Dans la " + entree_invalide[3] + ", le  champ " + entree_invalide[
                            0] + " et la " + \
                                  entree_invalide[
                                      1] + " l'entrée " + entree_invalide[2] + "\n"
                    messagebox.showwarning("Warning", message)
                    root.deiconify()
                    return
                root.withdraw()
                global plan_gestion_filename
                plan_gestion_filename = filedialog.asksaveasfilename(initialdir="/", title="File Explorer",
                                                                     filetypes=(
                                                                         ("json files", "*.json"),
                                                                         ("all files", "*.*")))
                if plan_gestion_filename == "":
                    root.deiconify()
                    return False
                if ".json" not in plan_gestion_filename:
                    plan_gestion_filename = plan_gestion_filename + ".json"

                save_dict = simulations

                with open(plan_gestion_filename, 'w') as json_file:
                    json.dump(save_dict, json_file)
                root.deiconify()

            def retour_au_menu_principal():
                response = messagebox.askyesno(title="Retour menu principal",
                                               message="Lors d'un retour au menu principal toutes les données non sauvegardées seront perdues. Souhaitez-vous retourner au menu principal?")
                if response:
                    root.destroy()
                    initialize_run_gui()

            def creation_du_rapport(bilan_response):
                bilan_workbook = Workbook()

                def sauvegarder_rapport_des_resultats():
                    root.withdraw()
                    filename = filedialog.asksaveasfilename(initialdir="/", title="File Explorer",
                                                            filetypes=(("xlsx files", "*.xlsx"), ("all files", "*.*")))
                    if filename == "":
                        root.deiconify()
                        return 1, ""
                    if ".xlsx" not in filename:
                        filename = filename + ".xlsx"
                    try:
                        bilan_workbook.save(filename)
                        root.deiconify()
                        message = (2, "La sauvegarde a été un succès!")
                        return message
                    except PermissionError:
                        message = (3, "La sauvegarde a échoué car le fichier était ouvert par une autre application!")
                        root.deiconify()
                        return message

                description_champs_worksheet = bilan_workbook.active
                description_champs_worksheet.title = "Zone"
                index_column_cell = 1
                index_row_cell = 1
                global information_champs
                global nom_entreprise
                global duree_simulation
                description_champs_worksheet.cell(row=index_row_cell, column=index_column_cell, value="Producteur")
                index_column_cell += 1
                description_champs_worksheet.cell(row=index_row_cell, column=index_column_cell, value="Champ")
                index_column_cell += 1
                description_champs_worksheet.cell(row=index_row_cell, column=index_column_cell, value="Zone de gestion")
                index_column_cell += 1
                description_champs_worksheet.cell(row=index_row_cell, column=index_column_cell,
                                                  value="Superficie de la zone (ha)")
                index_column_cell += 1
                description_champs_worksheet.cell(row=index_row_cell, column=index_column_cell, value="Municipalité")
                index_column_cell += 1
                description_champs_worksheet.cell(row=index_row_cell, column=index_column_cell,
                                                  value="Taux de matière organique (%)")
                index_column_cell += 1
                description_champs_worksheet.cell(row=index_row_cell, column=index_column_cell,
                                                  value="Masse volumique apparente (g/cm3)")
                index_column_cell += 1
                description_champs_worksheet.cell(row=index_row_cell, column=index_column_cell,
                                                  value="Profondeur de la couche arable (cm)")
                index_column_cell += 1
                description_champs_worksheet.cell(row=index_row_cell, column=index_column_cell, value="Groupe textural")
                index_column_cell += 1
                description_champs_worksheet.cell(row=index_row_cell, column=index_column_cell,
                                                  value="Classe de drainage")
                index_column_cell = 1
                index_row_cell += 1
                font = Font(bold=True)
                alignment = Alignment(wrap_text=True)
                for cell_name in ['A1', 'B1', 'C1', 'D1', 'E1', 'F1', 'G1', 'H1', 'I1', 'J1']:
                    cell = description_champs_worksheet[cell_name]
                    cell.font = font
                    cell.alignment = alignment
                for column_name in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J']:
                    description_champs_worksheet.column_dimensions[column_name].width = 14
                for champ in information_champs:
                    index_zone = 1
                    for zone in champ["information_zone_de_gestion"]:
                        description_champs_worksheet.cell(row=index_row_cell, column=index_column_cell,
                                                          value=nom_entreprise).alignment = alignment
                        index_column_cell += 1
                        description_champs_worksheet.cell(row=index_row_cell, column=index_column_cell,
                                                          value=champ["nom_du_champs"]).alignment = alignment
                        index_column_cell += 1
                        description_champs_worksheet.cell(row=index_row_cell, column=index_column_cell,
                                                          value=str(index_zone)).alignment = alignment
                        index_column_cell += 1
                        description_champs_worksheet.cell(row=index_row_cell, column=index_column_cell,
                                                          value=zone["superficie_de_la_zone"]).alignment = alignment
                        index_column_cell += 1
                        description_champs_worksheet.cell(row=index_row_cell, column=index_column_cell,
                                                          value=zone["municipalite"]).alignment = alignment
                        index_column_cell += 1
                        description_champs_worksheet.cell(row=index_row_cell, column=index_column_cell,
                                                          value=zone["taux_matiere_organique"]).alignment = alignment
                        index_column_cell += 1
                        description_champs_worksheet.cell(row=index_row_cell, column=index_column_cell,
                                                          value=zone["masse_volumique_apparente"]).alignment = alignment
                        index_column_cell += 1
                        description_champs_worksheet.cell(row=index_row_cell, column=index_column_cell,
                                                          value=zone["profondeur"]).alignment = alignment
                        index_column_cell += 1
                        description_champs_worksheet.cell(row=index_row_cell, column=index_column_cell,
                                                          value=zone["groupe_textural"]).alignment = alignment
                        index_column_cell += 1
                        description_champs_worksheet.cell(row=index_row_cell, column=index_column_cell,
                                                          value=zone["classe_de_drainage"]).alignment = alignment
                        index_zone += 1
                        index_column_cell = 1
                        index_row_cell += 1
                description_regies_simulations_worksheet = bilan_workbook.create_sheet("Simulation")
                index_column_cell = 1
                index_row_cell = 1
                description_regies_simulations_worksheet.cell(row=index_row_cell, column=index_column_cell,
                                                              value="Producteur")
                index_column_cell += 1
                description_regies_simulations_worksheet.cell(row=index_row_cell, column=index_column_cell,
                                                              value="Champ")
                index_column_cell += 1
                description_regies_simulations_worksheet.cell(row=index_row_cell, column=index_column_cell,
                                                              value="Zone de gestion")
                index_column_cell += 1
                description_regies_simulations_worksheet.cell(row=index_row_cell, column=index_column_cell,
                                                              value="Simulation")
                index_column_cell += 1
                description_regies_simulations_worksheet.cell(row=index_row_cell, column=index_column_cell,
                                                              value="Année de rotation")
                index_column_cell += 1
                description_regies_simulations_worksheet.cell(row=index_row_cell, column=index_column_cell,
                                                              value="Culture principale")
                index_column_cell += 1
                description_regies_simulations_worksheet.cell(row=index_row_cell, column=index_column_cell,
                                                              value="Culture secondaire")
                index_column_cell += 1
                description_regies_simulations_worksheet.cell(row=index_row_cell, column=index_column_cell,
                                                              value="Amendements")
                index_column_cell += 1
                description_regies_simulations_worksheet.cell(row=index_row_cell, column=index_column_cell,
                                                              value="Apport Cult. Princ. Racinaire (t/ha)")
                index_column_cell += 1
                description_regies_simulations_worksheet.cell(row=index_row_cell, column=index_column_cell,
                                                              value="Apport Cult. Princ. Aérien (t/ha)")
                index_column_cell += 1
                description_regies_simulations_worksheet.cell(row=index_row_cell, column=index_column_cell,
                                                              value="Apport Cult. Sec. Racinaire (t/ha)")
                index_column_cell += 1
                description_regies_simulations_worksheet.cell(row=index_row_cell, column=index_column_cell,
                                                              value="Apport Cult. Sec. Aérien (t/ha)")
                index_column_cell += 1
                description_regies_simulations_worksheet.cell(row=index_row_cell, column=index_column_cell,
                                                              value="Amendements (t/ha)")
                font = Font(bold=True)
                alignment = Alignment(wrap_text=True)
                for cell_name in ['A1', 'B1', 'C1', 'D1', 'E1', 'F1', 'G1', 'H1', 'I1', 'J1', 'K1', 'L1', 'M1']:
                    cell = description_regies_simulations_worksheet[cell_name]
                    cell.font = font
                    cell.alignment = alignment
                for column_name in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M']:
                    description_regies_simulations_worksheet.column_dimensions[column_name].width = 14
                index_column_cell = 1
                index_row_cell += 1
                index_simulation = 0
                bilans_simulations = bilan_response["bilans_des_simulations"]
                for simulation in bilans_simulations:
                    for champ in simulation["bilans_des_champs"]:
                        index_zone = 1
                        for zone in champ["bilans_des_zones"]:
                            index_annee = 0
                            for year in zone["bilan_des_regies_pour_la_duree_de_la_simulation"]:
                                if index_annee == len(zone["bilan_des_regies_projections"]):
                                    break
                                description_regies_simulations_worksheet.cell(row=index_row_cell,
                                                                              column=index_column_cell,
                                                                              value=simulation[
                                                                                  "nom_entreprise"]).alignment = alignment
                                index_column_cell += 1
                                description_regies_simulations_worksheet.cell(row=index_row_cell,
                                                                              column=index_column_cell,
                                                                              value=champ[
                                                                                  "nom_champs"]).alignment = alignment
                                index_column_cell += 1
                                description_regies_simulations_worksheet.cell(row=index_row_cell,
                                                                              column=index_column_cell,
                                                                              value=str(
                                                                                  index_zone)).alignment = alignment
                                index_column_cell += 1
                                description_regies_simulations_worksheet.cell(row=index_row_cell,
                                                                              column=index_column_cell,
                                                                              value=duree_simulation[index_simulation][
                                                                                  "nom_simulation"]).alignment = alignment
                                index_column_cell += 1
                                description_regies_simulations_worksheet.cell(row=index_row_cell,
                                                                              column=index_column_cell,
                                                                              value=str(
                                                                                  index_annee + 1)).alignment = alignment
                                index_column_cell += 1
                                description_regies_simulations_worksheet.cell(row=index_row_cell,
                                                                              column=index_column_cell,
                                                                              value=year["culture_principale"][
                                                                                  "culture_principale"]).alignment = alignment
                                index_column_cell += 1
                                description_regies_simulations_worksheet.cell(row=index_row_cell,
                                                                              column=index_column_cell,
                                                                              value=year["culture_secondaire"][
                                                                                  "culture_secondaire"]).alignment = alignment
                                index_column_cell += 1
                                amendements = ""
                                index_amendements = 0
                                for amendement in year["amendements"]["amendements"]:
                                    if index_amendements < len(year["amendements"]["amendements"]) - 1:
                                        if amendement["amendement"] is None:
                                            pass
                                        else:
                                            amendements = amendements + amendement["amendement"] + " + "
                                    else:
                                        if amendement["amendement"] is None and "+" not in amendements:
                                            amendements = "Aucun"
                                        elif amendement["amendement"] is None and "+" in amendements:
                                            amendements = amendements[0:len(amendements) - 3]
                                        else:
                                            amendements = amendements + amendement["amendement"]
                                    index_amendements += 1
                                description_regies_simulations_worksheet.cell(row=index_row_cell,
                                                                              column=index_column_cell,
                                                                              value=amendements).alignment = alignment
                                index_column_cell += 1
                                bilan_apports_cultures_principales_racinaires = description_regies_simulations_worksheet.cell(
                                    row=index_row_cell,
                                    column=index_column_cell,
                                    value=zone[
                                        "bilan_apports_cultures_principales_racinaires"][
                                        index_annee])
                                bilan_apports_cultures_principales_racinaires.alignment = alignment
                                bilan_apports_cultures_principales_racinaires.number_format = '0.000'
                                index_column_cell += 1
                                bilan_apports_cultures_principales_aeriennes = description_regies_simulations_worksheet.cell(
                                    row=index_row_cell,
                                    column=index_column_cell,
                                    value=zone[
                                        "bilan_apports_cultures_principales_aeriennes"][
                                        index_annee])
                                bilan_apports_cultures_principales_aeriennes.alignment = alignment
                                bilan_apports_cultures_principales_aeriennes.number_format = '0.000'
                                index_column_cell += 1
                                bilan_apports_cultures_secondaires_racinaires = description_regies_simulations_worksheet.cell(
                                    row=index_row_cell,
                                    column=index_column_cell,
                                    value=zone[
                                        "bilan_apports_cultures_secondaires_racinaires"][
                                        index_annee])
                                bilan_apports_cultures_secondaires_racinaires.alignment = alignment
                                bilan_apports_cultures_secondaires_racinaires.number_format = '0.000'
                                index_column_cell += 1
                                bilan_apports_cultures_secondaires_aeriennes = description_regies_simulations_worksheet.cell(
                                    row=index_row_cell,
                                    column=index_column_cell,
                                    value=zone[
                                        "bilan_apports_cultures_secondaires_aeriennes"][
                                        index_annee])
                                bilan_apports_cultures_secondaires_aeriennes.alignment = alignment
                                bilan_apports_cultures_secondaires_aeriennes.number_format = '0.000'
                                index_column_cell += 1
                                bilan_apports_amendements = description_regies_simulations_worksheet.cell(
                                    row=index_row_cell,
                                    column=index_column_cell,
                                    value=zone[
                                        "bilan_apports_amendements"][
                                        index_annee])
                                bilan_apports_amendements.alignment = alignment
                                bilan_apports_amendements.number_format = '0.000'
                                index_column_cell = 1
                                index_row_cell += 1
                                index_annee += 1
                            index_zone += 1
                    index_simulation += 1

                description_resultats_annuels_worksheet = bilan_workbook.create_sheet("Résultats annuels")
                index_column_cell = 1
                index_row_cell = 1
                description_resultats_annuels_worksheet.cell(row=index_row_cell, column=index_column_cell,
                                                             value="Producteur")
                index_column_cell += 1
                description_resultats_annuels_worksheet.cell(row=index_row_cell, column=index_column_cell,
                                                             value="Champ")
                index_column_cell += 1
                description_resultats_annuels_worksheet.cell(row=index_row_cell, column=index_column_cell,
                                                             value="Zone de gestion")
                index_column_cell += 1
                description_resultats_annuels_worksheet.cell(row=index_row_cell, column=index_column_cell,
                                                             value="Simulation")
                index_column_cell += 1
                description_resultats_annuels_worksheet.cell(row=index_row_cell, column=index_column_cell,
                                                             value="Année de projection")
                index_column_cell += 1
                description_resultats_annuels_worksheet.cell(row=index_row_cell, column=index_column_cell,
                                                             value="Y aérien C (t/ha)")
                index_column_cell += 1
                description_resultats_annuels_worksheet.cell(row=index_row_cell, column=index_column_cell,
                                                             value="Y racinaire C (t/ha)")
                index_column_cell += 1
                description_resultats_annuels_worksheet.cell(row=index_row_cell, column=index_column_cell,
                                                             value="Y amendements C (t/ha)")
                index_column_cell += 1
                description_resultats_annuels_worksheet.cell(row=index_row_cell, column=index_column_cell,
                                                             value="C jeune total (t/ha)")
                index_column_cell += 1
                description_resultats_annuels_worksheet.cell(row=index_row_cell, column=index_column_cell,
                                                             value="C stable total (t/ha)")
                index_column_cell += 1
                description_resultats_annuels_worksheet.cell(row=index_row_cell, column=index_column_cell,
                                                             value="C total (t/ha)")
                index_column_cell += 1
                description_resultats_annuels_worksheet.cell(row=index_row_cell, column=index_column_cell,
                                                             value="MO (t/ha)")
                index_column_cell = 1
                index_row_cell += 1

                font = Font(bold=True)
                alignment = Alignment(wrap_text=True)
                for cell_name in ['A1', 'B1', 'C1', 'D1', 'E1', 'F1', 'G1', 'H1', 'I1', 'J1', 'K1', 'L1']:
                    cell = description_resultats_annuels_worksheet[cell_name]
                    cell.font = font
                    cell.alignment = alignment
                for column_name in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L']:
                    description_resultats_annuels_worksheet.column_dimensions[column_name].width = 14
                index_simulation = 0
                for simulation in bilans_simulations:
                    for champ in simulation["bilans_des_champs"]:
                        index_zone = 1
                        for zone in champ["bilans_des_zones"]:
                            index_annee = 0
                            for year in zone["bilan_des_regies_pour_la_duree_de_la_simulation"]:
                                description_resultats_annuels_worksheet.cell(row=index_row_cell,
                                                                             column=index_column_cell,
                                                                             value=simulation[
                                                                                 "nom_entreprise"]).alignment = alignment
                                index_column_cell += 1
                                description_resultats_annuels_worksheet.cell(row=index_row_cell,
                                                                             column=index_column_cell,
                                                                             value=champ[
                                                                                 "nom_champs"]).alignment = alignment
                                index_column_cell += 1
                                description_resultats_annuels_worksheet.cell(row=index_row_cell,
                                                                             column=index_column_cell,
                                                                             value=str(
                                                                                 index_zone)).alignment = alignment
                                index_column_cell += 1
                                description_resultats_annuels_worksheet.cell(row=index_row_cell,
                                                                             column=index_column_cell,
                                                                             value=duree_simulation[index_simulation][
                                                                                 "nom_simulation"]).alignment = alignment
                                index_column_cell += 1
                                description_resultats_annuels_worksheet.cell(row=index_row_cell,
                                                                             column=index_column_cell,
                                                                             value=year[
                                                                                 "annee_culture"]).alignment = alignment
                                index_column_cell += 1
                                bilan_etats_pool_jeune_aerien = description_resultats_annuels_worksheet.cell(
                                    row=index_row_cell,
                                    column=index_column_cell,
                                    value=
                                    zone["bilan_etats_pool_jeune_aerien"][
                                        index_annee])
                                bilan_etats_pool_jeune_aerien.alignment = alignment
                                bilan_etats_pool_jeune_aerien.number_format = '0.000'
                                index_column_cell += 1
                                bilan_etats_pool_jeune_racinaire = description_resultats_annuels_worksheet.cell(
                                    row=index_row_cell,
                                    column=index_column_cell,
                                    value=
                                    zone["bilan_etats_pool_jeune_racinaire"][
                                        index_annee])
                                bilan_etats_pool_jeune_racinaire.alignment = alignment
                                bilan_etats_pool_jeune_racinaire.number_format = '0.000'
                                index_column_cell += 1
                                bilan_etats_pool_jeune_amendements = description_resultats_annuels_worksheet.cell(
                                    row=index_row_cell,
                                    column=index_column_cell,
                                    value=
                                    zone["bilan_etats_pool_jeune_amendements"][
                                        index_annee])
                                bilan_etats_pool_jeune_amendements.alignment = alignment
                                bilan_etats_pool_jeune_amendements.number_format = '0.000'
                                index_column_cell += 1
                                bilan_etats_pool_jeune_total = description_resultats_annuels_worksheet.cell(
                                    row=index_row_cell,
                                    column=index_column_cell,
                                    value=
                                    zone["bilan_etats_pool_jeune_total"][
                                        index_annee])
                                bilan_etats_pool_jeune_total.alignment = alignment
                                bilan_etats_pool_jeune_total.number_format = '0.000'
                                index_column_cell += 1
                                bilan_etats_pool_stable = description_resultats_annuels_worksheet.cell(
                                    row=index_row_cell,
                                    column=index_column_cell,
                                    value=
                                    zone["bilan_etats_pool_stable"][
                                        index_annee])
                                bilan_etats_pool_stable.alignment = alignment
                                bilan_etats_pool_stable.number_format = '0.000'
                                index_column_cell += 1
                                bilan_carbone_de_la_zone_pour_la_simulation = description_resultats_annuels_worksheet.cell(
                                    row=index_row_cell,
                                    column=index_column_cell,
                                    value=
                                    zone[
                                        "bilan_carbone_de_la_zone_pour_la_simulation"][
                                        index_annee])
                                bilan_carbone_de_la_zone_pour_la_simulation.alignment = alignment
                                bilan_carbone_de_la_zone_pour_la_simulation.number_format = '0.000'
                                index_column_cell += 1
                                bilan_matiere_orgagnique_pour_la_simulation = description_resultats_annuels_worksheet.cell(
                                    row=index_row_cell,
                                    column=index_column_cell,
                                    value=
                                    zone[
                                        "bilan_matiere_orgagnique_pour_la_simulation"][
                                        index_annee])
                                bilan_matiere_orgagnique_pour_la_simulation.alignment = alignment
                                bilan_matiere_orgagnique_pour_la_simulation.number_format = '0.000'
                                index_column_cell = 1
                                index_row_cell += 1
                                index_annee += 1
                            index_zone += 1
                    index_simulation += 1

                description_resultats_sommaire_worksheet = bilan_workbook.create_sheet("Résultats sommaires")
                index_column_cell = 1
                index_row_cell = 1
                description_resultats_sommaire_worksheet.cell(row=index_row_cell, column=index_column_cell,
                                                              value="Producteur")
                index_column_cell += 1
                description_resultats_sommaire_worksheet.cell(row=index_row_cell, column=index_column_cell,
                                                              value="Champ")
                index_column_cell += 1
                description_resultats_sommaire_worksheet.cell(row=index_row_cell, column=index_column_cell,
                                                              value="Zone de gestion")
                index_column_cell += 1
                description_resultats_sommaire_worksheet.cell(row=index_row_cell, column=index_column_cell,
                                                              value="Simulation")
                index_column_cell += 1
                description_resultats_sommaire_worksheet.cell(row=index_row_cell, column=index_column_cell,
                                                              value="Nombre d'années projetées")
                index_column_cell += 1
                description_resultats_sommaire_worksheet.cell(row=index_row_cell, column=index_column_cell,
                                                              value="Apport moyen culture principale (t/ha)")
                index_column_cell += 1
                description_resultats_sommaire_worksheet.cell(row=index_row_cell, column=index_column_cell,
                                                              value="Apport moyen culture secondaire (t/ha)")
                index_column_cell += 1
                description_resultats_sommaire_worksheet.cell(row=index_row_cell, column=index_column_cell,
                                                              value="Apport moyen amendement (t/ha)")
                index_column_cell += 1
                description_resultats_sommaire_worksheet.cell(row=index_row_cell, column=index_column_cell,
                                                              value="Teneur initiale en MOS (%)")
                index_column_cell += 1
                description_resultats_sommaire_worksheet.cell(row=index_row_cell, column=index_column_cell,
                                                              value="Teneur finale en MOS (%)")
                index_column_cell += 1
                description_resultats_sommaire_worksheet.cell(row=index_row_cell, column=index_column_cell,
                                                              value="Évolution du taux de MOS (%)")
                index_column_cell += 1
                description_resultats_sommaire_worksheet.cell(row=index_row_cell, column=index_column_cell,
                                                              value="Comparaison 50ème percentile")
                index_column_cell += 1
                description_resultats_sommaire_worksheet.cell(row=index_row_cell, column=index_column_cell,
                                                              value="Comparaison 90ème percentile")
                index_column_cell = 1
                index_row_cell += 1

                font = Font(bold=True)
                alignment = Alignment(wrap_text=True)
                for cell_name in ['A1', 'B1', 'C1', 'D1', 'E1', 'F1', 'G1', 'H1', 'I1', 'J1', 'K1', 'L1', 'M1']:
                    cell = description_resultats_sommaire_worksheet[cell_name]
                    cell.font = font
                    cell.alignment = alignment
                for column_name in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M']:
                    description_resultats_sommaire_worksheet.column_dimensions[column_name].width = 14
                index_simulation = 0
                for simulation in bilans_simulations:
                    for champ in simulation["bilans_des_champs"]:
                        index_zone = 1
                        for zone in champ["bilans_des_zones"]:
                            description_resultats_sommaire_worksheet.cell(row=index_row_cell, column=index_column_cell,
                                                                          value=simulation[
                                                                              "nom_entreprise"])
                            index_column_cell += 1
                            description_resultats_sommaire_worksheet.cell(row=index_row_cell, column=index_column_cell,
                                                                          value=champ[
                                                                              "nom_champs"]).alignment = alignment
                            index_column_cell += 1
                            description_resultats_sommaire_worksheet.cell(row=index_row_cell, column=index_column_cell,
                                                                          value=str(index_zone)).alignment = alignment
                            index_column_cell += 1
                            description_resultats_sommaire_worksheet.cell(row=index_row_cell, column=index_column_cell,
                                                                          value=duree_simulation[index_simulation][
                                                                              "nom_simulation"]).alignment = alignment
                            index_column_cell += 1
                            description_resultats_sommaire_worksheet.cell(row=index_row_cell, column=index_column_cell,
                                                                          value=len(zone[
                                                                                        "bilan_des_regies_pour_la_duree_de_la_simulation"])).alignment = alignment
                            index_column_cell += 1
                            moyenne_apports_cultures_principales = description_resultats_sommaire_worksheet.cell(
                                row=index_row_cell, column=index_column_cell,
                                value=zone[
                                    "moyenne_apports_cultures_principales"])
                            moyenne_apports_cultures_principales.alignment = alignment
                            moyenne_apports_cultures_principales.number_format = '0.000'
                            index_column_cell += 1
                            moyenne_apports_cultures_secondaires = description_resultats_sommaire_worksheet.cell(
                                row=index_row_cell, column=index_column_cell,
                                value=zone[
                                    "moyenne_apports_cultures_secondaires"])
                            moyenne_apports_cultures_secondaires.alignment = alignment
                            moyenne_apports_cultures_secondaires.number_format = '0.000'
                            index_column_cell += 1
                            moyenne_apports_amendement = description_resultats_sommaire_worksheet.cell(
                                row=index_row_cell, column=index_column_cell,
                                value=zone[
                                    "moyenne_apports_amendements"])
                            moyenne_apports_amendement.alignment = alignment
                            moyenne_apports_amendement.number_format = '0.000'
                            index_column_cell += 1
                            taux_matiere_organique_original = description_resultats_sommaire_worksheet.cell(
                                row=index_row_cell, column=index_column_cell,
                                value=zone[
                                    "taux_de_matiere_organique_initial"])
                            taux_matiere_organique_original.alignment = alignment
                            taux_matiere_organique_original.number_format = '0.000'
                            index_column_cell += 1
                            teneur_finale_projetee = description_resultats_sommaire_worksheet.cell(row=index_row_cell,
                                                                                                   column=index_column_cell,
                                                                                                   value=zone[
                                                                                                       "teneur_finale_projetee"])
                            teneur_finale_projetee.alignment = alignment
                            teneur_finale_projetee.number_format = '0.000'
                            index_column_cell += 1
                            difference_entre_la_teneur_finale_et_la_zone = description_resultats_sommaire_worksheet.cell(
                                row=index_row_cell, column=index_column_cell,
                                value=zone[
                                    "difference_entre_la_teneur_finale_et_la_zone"])
                            difference_entre_la_teneur_finale_et_la_zone.alignment = alignment
                            difference_entre_la_teneur_finale_et_la_zone.number_format = '0.000'
                            index_column_cell += 1
                            comparaison_percentile50 = description_resultats_sommaire_worksheet.cell(row=index_row_cell,
                                                                                                     column=index_column_cell,
                                                                                                     value=zone[
                                                                                                         "comparaison_percentile50"])
                            comparaison_percentile50.alignment = alignment
                            comparaison_percentile50.number_format = '0.000'
                            index_column_cell += 1
                            comparaison_percentile90 = description_resultats_sommaire_worksheet.cell(row=index_row_cell,
                                                                                                     column=index_column_cell,
                                                                                                     value=zone[
                                                                                                         "comparaison_percentile90"])
                            comparaison_percentile90.alignment = alignment
                            comparaison_percentile90.number_format = '0.000'

                            index_column_cell = 1
                            index_row_cell += 1
                            index_zone += 1
                    index_simulation += 1
                sauvegarde_succes = sauvegarder_rapport_des_resultats()
                if sauvegarde_succes[0] == 2:
                    messagebox.showinfo("Résultat sauvegarde", sauvegarde_succes[1])
                elif sauvegarde_succes[0] == 1:
                    pass
                else:
                    messagebox.showinfo("Résultat sauvegarde", sauvegarde_succes[1])

            menu_initial_ogemos(frame)

        run_gui(mainframe)

        root.mainloop()

    initialize_run_gui()


if __name__ == "__main__":
    initialize_globals()
