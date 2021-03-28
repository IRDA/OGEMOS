from openpyxl import load_workbook
import copy
from GUI.CustomExceptions import DependanceError


def map_excel_to_json(path):
    wb = load_workbook(filename=path)
    try:
        donnee_entreprises = wb["Entreprises"]
        donnee_simulations = wb["Simulations"]
        donnee_regies_projections = wb["Régies projection"]
        donnee_regies_historiques = wb["Régies historique"]
    except Exception:
        raise TypeError("Format du fichier fourni ne correspond pas au gabarit")

    donnee_entreprises = wb["Entreprises"]
    donnee_simulations = wb["Simulations"]
    donnee_regies_projections = wb["Régies projection"]
    donnee_regies_historiques = wb["Régies historique"]

    entreprises_dict = {}
    simulations_dict = {}

    initial_row = 2
    initial_column = 1

    row = initial_row
    column = initial_column

    while row_check(donnee=donnee_entreprises, row=row, number_of_columns=10, worksheet_name="Entreprises"):
        if str(donnee_entreprises.cell(row=row, column=column).value) in entreprises_dict.keys():
            entreprise = str(donnee_entreprises.cell(row=row, column=column).value)
            if entreprise == "None":
                raise TypeError(
                    "Type du nom de l'entreprise est invalide dans la rangée {} et la colonne {} de la feuille de calcul Entreprises".format(
                        row, column))
            elif not isinstance(entreprise, str):
                raise TypeError(
                    "Type du nom de l'entreprise est invalide dans la rangée {} et la colonne {} de la feuille de calcul Entreprises".format(
                        row, column))
            column += 1
            if str(donnee_entreprises.cell(row=row, column=column).value) in entreprises_dict[entreprise].keys():
                champ = str(donnee_entreprises.cell(row=row, column=column).value)
                if champ == "None":
                    raise TypeError(
                        "Type du nom du champ est invalide dans la rangée {} et la colonne {} de la feuille de calcul Entreprises".format(
                            row, column))
                elif not isinstance(champ, str):
                    raise TypeError(
                        "Type du nom du champ est invalide dans la rangée {} et la colonne {} de la feuille de calcul Entreprises".format(
                            row, column))
                column += 1
                zone_de_gestion = str(donnee_entreprises.cell(row=row, column=column).value)
                if zone_de_gestion in \
                        entreprises_dict[entreprise][champ].keys():
                    raise TypeError(
                        "Duplicat d'une zone de gestion dans la rangée {} et la colonne {} de la feuille de calcul Entreprises".format(
                            row, column))
                elif zone_de_gestion == "None":
                    raise TypeError(
                        "Type du nom de la zone de gestion est invalide dans la rangée {} et la colonne {} de la feuille de calcul Entreprises".format(
                            row, column))
                elif not isinstance(zone_de_gestion, str):
                    raise TypeError(
                        "Type du nom de la zone de gestion est invalide dans la rangée {} et la colonne {} de la feuille de calcul Entreprises".format(
                            row, column))
                else:
                    entreprises_dict[entreprise][champ][zone_de_gestion] = {}
                    zone_de_gestion_data = entreprises_dict[entreprise][champ][zone_de_gestion]
                    column += 1
                    ajout_caracteristiques_physiques(column, donnee_entreprises, row,
                                                     zone_de_gestion_data)
                    column = initial_column

            else:
                entreprises_dict[entreprise][str(donnee_entreprises.cell(row=row, column=column).value)] = {}
                champ = str(donnee_entreprises.cell(row=row, column=column).value)
                if champ == "None":
                    raise TypeError(
                        "Type du nom du champ est invalide dans la rangée {} et la colonne {} de la feuille de calcul Entreprises".format(
                            row, column))
                elif not isinstance(champ, str):
                    raise TypeError(
                        "Type du nom du champ est invalide dans la rangée {} et la colonne {} de la feuille de calcul Entreprises".format(
                            row, column))
                column += 1
                entreprises_dict[entreprise][champ][str(donnee_entreprises.cell(row=row, column=column).value)] = {}
                zone_de_gestion = str(donnee_entreprises.cell(row=row, column=column).value)
                if zone_de_gestion == "None":
                    raise TypeError(
                        "Type du nom de la zone de gestion est invalide dans la rangée {} et la colonne {} de la feuille de calcul Entreprises".format(
                            row, column))
                elif not isinstance(zone_de_gestion, str):
                    raise TypeError(
                        "Type du nom de la zone de gestion est invalide dans la rangée {} et la colonne {} de la feuille de calcul Entreprises".format(
                            row, column))
                zone_de_gestion_data = entreprises_dict[entreprise][champ][zone_de_gestion]
                column += 1
                ajout_caracteristiques_physiques(column, donnee_entreprises, row,
                                                 zone_de_gestion_data)
                column = initial_column
        else:
            entreprises_dict[str(donnee_entreprises.cell(row=row, column=column).value)] = {}
            entreprise = str(donnee_entreprises.cell(row=row, column=column).value)
            if entreprise == "None":
                raise TypeError(
                    "Type du nom de l'entreprise est invalide dans la rangée {} et la colonne {} de la feuille de calcul Entreprises".format(
                        row, column))
            elif not isinstance(entreprise, str):
                raise TypeError(
                    "Type du nom de l'entreprise est invalide dans la rangée {} et la colonne {} de la feuille de calcul Entreprises".format(
                        row, column))
            column += 1
            entreprises_dict[entreprise][str(donnee_entreprises.cell(row=row, column=column).value)] = {}
            champ = str(donnee_entreprises.cell(row=row, column=column).value)
            if champ == "None":
                raise TypeError(
                    "Type du nom du champ est invalide dans la rangée {} et la colonne {} de la feuille de calcul Entreprises".format(
                        row, column))
            elif not isinstance(champ, str):
                raise TypeError(
                    "Type du nom du champ est invalide dans la rangée {} et la colonne {} de la feuille de calcul Entreprises".format(
                        row, column))
            column += 1
            entreprises_dict[entreprise][champ][str(donnee_entreprises.cell(row=row, column=column).value)] = {}
            zone_de_gestion = str(donnee_entreprises.cell(row=row, column=column).value)
            if zone_de_gestion == "None":
                raise TypeError(
                    "Type du nom de la zone de gestion est invalide dans la rangée {} et la colonne {} de la feuille de calcul Entreprises".format(
                        row, column))
            elif not isinstance(zone_de_gestion, str):
                raise TypeError(
                    "Type du nom de la zone de gestion est invalide dans la rangée {} et la colonne {} de la feuille de calcul Entreprises".format(
                        row, column))
            zone_de_gestion_data = entreprises_dict[entreprise][champ][zone_de_gestion]
            column += 1
            ajout_caracteristiques_physiques(column, donnee_entreprises, row,
                                             zone_de_gestion_data)
            column = initial_column
        row += 1

    row = initial_row
    column = initial_column
    while row_check(donnee=donnee_simulations, row=row, number_of_columns=4, worksheet_name="Simulations"):
        simulation = str(donnee_simulations.cell(row=row, column=column).value)
        if simulation in simulations_dict.keys():
            raise TypeError(
                "Duplicat d'une simulation dans la rangée {} et la colonne {} de la feuille de calcul Simulations".format(
                    row, column))
        elif simulation == "None":
            raise TypeError(
                "Type du nom de la simulation est invalide dans la rangée {} et la colonne {} de la feuille de calcul Simulations".format(
                    row, column))
        elif not isinstance(simulation, str):
            raise TypeError(
                "Type du nom de la simulation est invalide dans la rangée {} et la colonne {} de la feuille de calcul Simulations".format(
                    row, column))
        else:
            simulations_dict[simulation] = {"nom_simulation": simulation}
            simulation_data = simulations_dict[simulation]
            column += 1
            try:
                simulation_data["duree_projection"] = int(donnee_simulations.cell(row=row, column=column).value)
            except ValueError:
                raise TypeError(
                    "Durée de la projection invalide dans la rangée {} et la colonne {} de la feuille de calcul Simulations".format(
                        row, column))
            except TypeError:
                raise TypeError(
                    "Durée de la projection invalide dans la rangée {} et la colonne {} de la feuille de calcul Simulations".format(
                        row, column))
            column += 1
            try:
                simulation_data["annee_initiale_projection"] = int(
                    donnee_simulations.cell(row=row, column=column).value)
            except ValueError:
                raise TypeError(
                    "Année initiale de la projection invalide dans la rangée {} et la colonne {} de la feuille de calcul Simulations".format(
                        row, column))
            except TypeError:
                raise TypeError(
                    "Année initiale de la projection invalide dans la rangée {} et la colonne {} de la feuille de calcul Simulations".format(
                        row, column))
            column += 1
            try:
                simulation_data["annee_finale_projection"] = int(
                    donnee_simulations.cell(row=row, column=column).value)
            except ValueError:
                raise TypeError(
                    "Année finale de la projection invalide dans la rangée {} et la colonne {} de la feuille de calcul Simulations".format(
                        row, column))
            except TypeError:
                raise TypeError(
                    "Année finale de la projection invalide dans la rangée {} et la colonne {} de la feuille de calcul Simulations".format(
                        row, column))
            column = initial_column
        row += 1

    simulation_data = {}
    for simulation in simulations_dict.keys():
        simulation_data[simulation] = {}
    row = initial_row
    column = initial_column
    while row_check(donnee=donnee_regies_projections, row=row, number_of_columns=15,
                    worksheet_name="Régies projection"):
        if str(donnee_regies_projections.cell(row=row, column=column).value) in simulations_dict.keys():
            if donnee_regies_projections.cell(row=row, column=column).value is None:
                raise TypeError(
                    "Type du nom de la simulation est invalide dans la rangée {} et la colonne {} de la feuille de calcul Régies projection".format(
                        row, column))
            simulation = str(donnee_regies_projections.cell(row=row, column=column).value)
            column += 1
            entreprise = str(donnee_regies_projections.cell(row=row, column=column).value)
            if entreprise is None:
                raise TypeError(
                    "Type du nom de l'entreprise est invalide dans la rangée {} et la colonne {} de la feuille de calcul Régies projection".format(
                        row, column))
            column += 1
            champ = str(donnee_regies_projections.cell(row=row, column=column).value)
            if champ is None:
                raise TypeError(
                    "Type du nom du champ est invalide dans la rangée {} et la colonne {} de la feuille de calcul Régies projection".format(
                        row, column))
            column += 1
            zone_gestion = str(donnee_regies_projections.cell(row=row, column=column).value)
            if zone_gestion is None:
                raise TypeError(
                    "Type du nom de la zone de gestion est invalide dans la rangée {} et la colonne {} de la feuille de calcul Régies projection".format(
                        row, column))
            column += 1
            try:
                annee_rotation = int(donnee_regies_projections.cell(row=row, column=column).value)
            except ValueError:
                raise TypeError(
                    "Année de rotation invalide dans la rangée {} et la colonne {} de la feuille de calcul Régies projection".format(
                        row, column))
            except TypeError:
                raise TypeError(
                    "Année de rotation invalide dans la rangée {} et la colonne {} de la feuille de calcul Régies projection".format(
                        row, column))
            column += 1
            if entreprise in simulation_data[simulation].keys():
                try:
                    assert champ in entreprises_dict[entreprise].keys()
                except AssertionError:
                    raise ValueError(
                        "Nom du champ est invalide dans la rangée {} et la colonne 3 de la feuille de calcul Régies projection".format(
                            row))
                try:
                    assert zone_gestion in entreprises_dict[entreprise][champ].keys()
                except AssertionError:
                    raise ValueError(
                        "Nom de la zone de gestion est invalide dans la rangée {} et la colonne 4 de la feuille de calcul Régies projection".format(
                            row))
            elif entreprise not in simulation_data[simulation].keys() and len(
                    simulation_data[simulation].keys()) == 1:
                raise ValueError(
                    "Nom de l'entreprise invalide dans la rangée {} de la feuille de calcul Régies projection, les simulations ne peuvent avoir qu'une seule entreprise".format(
                        row))
            else:
                try:
                    assert entreprise in entreprises_dict.keys()
                except AssertionError:
                    raise ValueError(
                        "Nom de l'entreprise est invalide dans la rangée {} et la colonne 2 de la feuille de calcul Régies projection".format(
                            row))
                try:
                    assert champ in entreprises_dict[entreprise].keys()
                except AssertionError:
                    raise ValueError(
                        "Nom du champ est invalide dans la rangée {} et la colonne 3 de la feuille de calcul Régies projection".format(
                            row))
                try:
                    assert zone_gestion in entreprises_dict[entreprise][champ].keys()
                except AssertionError:
                    raise ValueError(
                        "Nom de la zone de gestion est invalide dans la rangée {} et la colonne 4 de la feuille de calcul Régies projection".format(
                            row))
                simulation_data[simulation][entreprise] = copy.deepcopy(entreprises_dict[entreprise])
            if int(annee_rotation) == len(
                    simulation_data[simulation][entreprise][champ][zone_gestion][
                        "regies_sol_et_culture_projection"]) + 1:
                if donnee_regies_projections.cell(row=row, column=column).value is None:
                    raise TypeError(
                        "Type de la culture principale est invalide dans la rangée {} et la colonne {} de la feuille de calcul Régies projection".format(
                            row, column))
                regie = {}
                regie["culture_principale"] = {}
                regie["culture_principale"]["culture_principale"] = str(donnee_regies_projections.cell(row=row,
                                                                                                       column=column).value)
                column += 1
                try:
                    regie["culture_principale"]["rendement"] = float(donnee_regies_projections.cell(row=row,
                                                                                                    column=column).value)
                except ValueError:
                    raise TypeError(
                        "Rendement culture principale invalide dans la rangée {} et la colonne {} de la feuille de calcul Régies projection".format(
                            row, column))
                except TypeError:
                    raise TypeError(
                        "Rendement culture principale invalide dans la rangée {} et la colonne {} de la feuille de calcul Régies projection".format(
                            row, column))
                column += 1
                try:
                    regie["culture_principale"]["pourcentage_tige_exporte"] = float(
                        donnee_regies_projections.cell(row=row,
                                                       column=column).value)
                except ValueError:
                    raise TypeError(
                        "Pourcentage tige exportée invalide dans la rangée {} et la colonne {} de la feuille de calcul Régies projection".format(
                            row, column))
                except TypeError:
                    raise TypeError(
                        "Pourcentage tige exportée invalide dans la rangée {} et la colonne {} de la feuille de calcul Régies projection".format(
                            row, column))
                column += 1
                if donnee_regies_projections.cell(row=row, column=column).value is None:
                    raise TypeError(
                        "Type de produit récolté est invalide dans la rangée {} et la colonne {} de la feuille de calcul Régies projection".format(
                            row, column))
                else:
                    if str(donnee_regies_projections.cell(row=row,
                                                          column=column).value) == "Oui" or str(
                        donnee_regies_projections.cell(
                            row=row, column=column).value) == "Non":
                        if str(donnee_regies_projections.cell(row=row, column=column).value) == "Oui":
                            regie["culture_principale"]["produit_recolte"] = True
                        else:
                            regie["culture_principale"]["produit_recolte"] = False
                    else:
                        raise ValueError(
                            "La valeur de la donnée dans la rangée {} et la colonne {} de la feuille de calcul Régies projection devrait être Oui ou Non".format(
                                row, column))
                column += 1
                try:
                    regie["culture_principale"]["pourcentage_humidite"] = float(
                        donnee_regies_projections.cell(row=row,
                                                       column=column).value)
                except ValueError:
                    raise TypeError(
                        "Pourcentage humidité invalide dans la rangée {} et la colonne {} de la feuille de calcul Régies projection".format(
                            row, column))
                except TypeError:
                    raise TypeError(
                        "Pourcentage humidité invalide dans la rangée {} et la colonne {} de la feuille de calcul Régies projection".format(
                            row, column))
                column += 1
                if donnee_regies_projections.cell(row=row, column=column).value is None:
                    raise TypeError(
                        "Type de culture secondaire est invalide dans la rangée {} et la colonne {} de la feuille de calcul Régies projection".format(
                            row, column))
                regie["culture_secondaire"] = {}
                culture_secondaire = str(donnee_regies_projections.cell(row=row, column=column).value)
                column += 1
                try:
                    rendement_culture_secondaire = float(donnee_regies_projections.cell(row=row,
                                                                                       column=column).value)
                    if culture_secondaire == "-1" and rendement_culture_secondaire == -1:
                        regie["culture_secondaire"]["culture_secondaire"] = None
                        regie["culture_secondaire"]["rendement"] = None
                    elif culture_secondaire == "-1" and not rendement_culture_secondaire == -1:
                        raise DependanceError(
                            "S'il n'y pas de culture secondaire, il ne doit pas y avoir de rendement pour celle-ci dans la feuille de calcul Régies projection à la rangée {}".format(row))
                    elif rendement_culture_secondaire == -1 and not culture_secondaire == "-1":
                        raise DependanceError(
                            "S'il n'y pas de rendement de culture secondaire, il ne doit pas y avoir de culture secondaire dans la feuille de calcul Régies projection à la rangée {}".format(row))
                    else:
                        regie["culture_secondaire"]["culture_secondaire"] = culture_secondaire
                        regie["culture_secondaire"]["rendement"] = rendement_culture_secondaire
                except ValueError:
                    raise TypeError(
                        "Rendement culture secondaire invalide dans la rangée {} et la colonne {} de la feuille de calcul Régies projection".format(
                            row, column))
                except TypeError:
                    raise TypeError(
                        "Rendement culture secondaire invalide dans la rangée {} et la colonne {} de la feuille de calcul Régies projection".format(
                            row, column))
                column += 1
                if donnee_regies_projections.cell(row=row, column=column).value is None:
                    raise TypeError(
                        "Type des amendements est invalide dans la rangée {} et la colonne {} de la feuille de calcul Régies projection".format(
                            row, column))
                regie["amendements"] = []
                amendements = str(donnee_regies_projections.cell(row=row, column=column).value)
                column += 1
                amendements_apports = str(donnee_regies_projections.cell(row=row, column=column).value)
                column += 1
                if amendements == "-1" and amendements_apports == "-1":
                    pass
                elif amendements == "-1" and not amendements_apports == "-1":
                    raise DependanceError("S'il n'y a pas d'amendements, il ne doit pas y avoir d'apports pour ceux-ci dans la feuille de calcul Régies projection à la rangée {}".format(row))
                elif amendements_apports == "-1" and not amendements == "-1":
                    raise DependanceError("S'il n'y a pas d'apports d'amendements, il ne doit pas y avoir d'amendements dans la feuille de calcul Régies projection à la rangée {}".format(row))
                else:
                    amendements_list = amendements.split(",")
                    amendements_apports_list = amendements_apports.split(",")
                    index = 0
                    for amendement in amendements_list:
                        try:
                            float(amendements_apports_list[index])
                        except ValueError:
                            raise TypeError(
                                "Apports d'amendement invalide dans la rangée {} et la colonne 14 de la feuille de calcul Régies projection".format(
                                    row, column))
                        except TypeError:
                            raise TypeError(
                                "Apports d'amendement invalide dans la rangée {} et la colonne 14 de la feuille de calcul Régies projection".format(
                                    row, column))
                        regie["amendements"].append(
                            {"amendement": amendement, "apport": float(amendements_apports_list[index])})
                        index += 1
                if donnee_regies_projections.cell(row=row, column=column).value is None:
                    raise TypeError(
                        "Type de travail du sol est invalide dans la rangée {} et la colonne {} de la feuille de calcul Régies projection".format(
                            row, column))
                regie["travail_du_sol"] = {
                    "travail_du_sol": str(donnee_regies_projections.cell(row=row, column=column).value)}
            else:
                raise TypeError(
                    "Erreur d'ordonancement de la rotation dans le rangée {} de la feuille de calcul Régies projection".format(
                        row))
            simulation_data[simulation][entreprise][champ][zone_gestion]["regies_sol_et_culture_projection"].append(
                regie)
            print(simulation_data)
        else:
            raise TypeError(
                "Nom de simulation invalide dans la rangée {} et la colonne {} de la feuille de calcul Régies projection".format(
                    row, column))
        column = initial_column
        row += 1

    row = initial_row
    column = initial_column
    while row_check(donnee=donnee_regies_historiques, row=row, number_of_columns=15,
                    worksheet_name="Régies historique"):
        if str(donnee_regies_historiques.cell(row=row, column=column).value) in simulations_dict.keys():
            simulation = str(donnee_regies_historiques.cell(row=row, column=column).value)
            if donnee_regies_historiques.cell(row=row, column=column).value is None:
                raise TypeError(
                    "Type du nom de la simulation est invalide dans la rangée {} et la colonne {} de la feuille de calcul Régies historique".format(
                        row, column))

            column += 1
            entreprise = str(donnee_regies_historiques.cell(row=row, column=column).value)
            if entreprise is None:
                raise TypeError(
                    "Type du nom de l'entreprise est invalide dans la rangée {} et la colonne {} de la feuille de calcul Régies historique".format(
                        row, column))
            column += 1
            champ = str(donnee_regies_historiques.cell(row=row, column=column).value)
            if champ is None:
                raise TypeError(
                    "Type du nom du champ est invalide dans la rangée {} et la colonne {} de la feuille de calcul Régies historique".format(
                        row, column))
            column += 1
            zone_gestion = str(donnee_regies_historiques.cell(row=row, column=column).value)
            if zone_gestion is None:
                raise TypeError(
                    "Type du nom de la zone de gestion est invalide dans la rangée {} et la colonne {} de la feuille de calcul Régies historique".format(
                        row, column))
            column += 1
            try:
                annee_rotation = int(donnee_regies_historiques.cell(row=row, column=column).value)
            except ValueError:
                raise TypeError(
                    "Année de rotation invalide dans la rangée {} et la colonne {} de la feuille de calcul Régies historique".format(
                        row, column))
            except TypeError:
                raise TypeError(
                    "Année de rotation invalide dans la rangée {} et la colonne {} de la feuille de calcul Régies historique".format(
                        row, column))
            column += 1
            if entreprise in simulation_data[simulation].keys():
                try:
                    entreprises_dict[entreprise][champ]
                except KeyError:
                    raise ValueError(
                        "Nom du champ est invalide dans la rangée {} et la colonne 3 de la feuille de calcul Régies historique".format(
                            row))
                try:
                    entreprises_dict[entreprise][champ][zone_gestion]
                except KeyError:
                    raise ValueError(
                        "Nom de la zone de gestion est invalide dans la rangée {} et la colonne 4 de la feuille de calcul Régies historique".format(
                            row))
            elif entreprise not in simulation_data[simulation].keys() and len(
                    simulation_data[simulation].keys()) == 1:
                print(entreprise)
                print(simulation_data[simulation].keys())
                raise ValueError(
                    "Nom de l'entreprise invalide dans la rangée {} de la feuille de calcul Régies historique, les simulations ne peuvent avoir qu'une seule entreprise".format(
                        row))
            else:
                try:
                    entreprises_dict[entreprise]
                except KeyError:
                    raise ValueError(
                        "Nom de l'entreprise est invalide dans la rangée {} et la colonne 2 de la feuille de calcul Régies historique".format(
                            row))
                try:
                    entreprises_dict[entreprise][champ]
                except KeyError:
                    raise ValueError(
                        "Nom du champ est invalide dans la rangée {} et la colonne 3 de la feuille de calcul Régies historique".format(
                            row))
                try:
                    entreprises_dict[entreprise][champ][zone_gestion]
                except KeyError:
                    raise ValueError(
                        "Nom de la zone de gestion est invalide dans la rangée {} et la colonne 4 de la feuille de calcul Régies historique".format(
                            row))

                simulation_data[simulation][entreprise] = copy.deepcopy(entreprises_dict[entreprise])
            if int(annee_rotation) == len(
                    simulation_data[simulation][entreprise][champ][zone_gestion][
                        "regies_sol_et_culture_historique"]) + 1:
                if donnee_regies_historiques.cell(row=row, column=column).value is None:
                    raise TypeError(
                        "Type de la culture principale est invalide dans la rangée {} et la colonne {} de la feuille de calcul Régies historique".format(
                            row, column))
                regie = {}
                regie["culture_principale"] = {}
                regie["culture_principale"]["culture_principale"] = str(donnee_regies_historiques.cell(row=row,
                                                                                                       column=column).value)
                column += 1
                try:
                    regie["culture_principale"]["rendement"] = float(donnee_regies_historiques.cell(row=row,
                                                                                                    column=column).value)
                except ValueError:
                    raise TypeError(
                        "Rendement culture principale invalide dans la rangée {} et la colonne {} de la feuille de calcul Régies historique".format(
                            row, column))
                except TypeError:
                    raise TypeError(
                        "Rendement culture principale invalide dans la rangée {} et la colonne {} de la feuille de calcul Régies historique".format(
                            row, column))
                column += 1
                try:
                    regie["culture_principale"]["pourcentage_tige_exporte"] = float(
                        donnee_regies_historiques.cell(row=row,
                                                       column=column).value)
                except ValueError:
                    raise TypeError(
                        "Pourcentage tige exportée invalide dans la rangée {} et la colonne {} de la feuille de calcul Régies historique".format(
                            row, column))
                except TypeError:
                    raise TypeError(
                        "Pourcentage tige exportée invalide dans la rangée {} et la colonne {} de la feuille de calcul Régies historique".format(
                            row, column))
                column += 1
                if donnee_regies_historiques.cell(row=row, column=column).value is None:
                    raise TypeError(
                        "Type de produit récolté est invalide dans la rangée {} et la colonne {} de la feuille de calcul Régies historique".format(
                            row, column))
                else:
                    if str(donnee_regies_historiques.cell(row=row,
                                                          column=column).value) == "Oui" or str(
                        donnee_regies_historiques.cell(
                            row=row, column=column).value) == "Non":
                        if str(donnee_regies_historiques.cell(row=row, column=column).value) == "Oui":
                            regie["culture_principale"]["produit_recolte"] = True
                        else:
                            regie["culture_principale"]["produit_recolte"] = False
                    else:
                        raise ValueError(
                            "La valeur de la donnée dans la rangée {} et la colonne {} de la feuille de calcul Régies historique devrait être Oui ou Non".format(
                                row, column))
                column += 1
                try:
                    regie["culture_principale"]["pourcentage_humidite"] = float(
                        donnee_regies_historiques.cell(row=row,
                                                       column=column).value)
                except ValueError:
                    raise TypeError(
                        "Pourcentage humidité invalide dans la rangée {} et la colonne {} de la feuille de calcul Régies historique".format(
                            row, column))
                except TypeError:
                    raise TypeError(
                        "Pourcentage humidité invalide dans la rangée {} et la colonne {} de la feuille de calcul Régies historique".format(
                            row, column))
                column += 1
                if donnee_regies_historiques.cell(row=row, column=column).value is None:
                    raise TypeError(
                        "Type de culture secondaire est invalide dans la rangée {} et la colonne {} de la feuille de calcul Régies historique".format(
                            row, column))
                regie["culture_secondaire"] = {}
                culture_secondaire = str(donnee_regies_historiques.cell(row=row, column=column).value)
                column += 1
                try:
                    rendement_culture_secondaire = float(donnee_regies_historiques.cell(row=row,
                                                                                       column=column).value)
                    if culture_secondaire == "-1" and rendement_culture_secondaire == -1:
                        regie["culture_secondaire"]["culture_secondaire"] = None
                        regie["culture_secondaire"]["rendement"] = None
                    elif culture_secondaire == "-1" and not rendement_culture_secondaire == -1:
                        raise DependanceError(
                            "S'il n'y pas de culture secondaire, il ne doit pas y avoir de rendement pour celle-ci dans la feuille de calcul Régies historique à la rangée {}".format(row))
                    elif rendement_culture_secondaire == -1 and not culture_secondaire == "-1":
                        raise DependanceError(
                            "S'il n'y pas de rendement de culture secondaire, il ne doit pas y avoir de culture secondaire dans la feuille de calcul Régies historique à la rangée {}".format(row))
                    else:
                        regie["culture_secondaire"]["culture_secondaire"] = culture_secondaire
                        regie["culture_secondaire"]["rendement"] = rendement_culture_secondaire
                except ValueError:
                    raise TypeError(
                        "Rendement culture secondaire invalide dans la rangée {} et la colonne {} de la feuille de calcul Régies historique".format(
                            row, column))
                except TypeError:
                    raise TypeError(
                        "Rendement culture secondaire invalide dans la rangée {} et la colonne {} de la feuille de calcul Régies historique".format(
                            row, column))
                column += 1
                if donnee_regies_historiques.cell(row=row, column=column).value is None:
                    raise TypeError(
                        "Type des amendements est invalide dans la rangée {} et la colonne {} de la feuille de calcul Régies historique".format(
                            row, column))
                regie["amendements"] = []
                amendements = str(donnee_regies_historiques.cell(row=row, column=column).value)
                column += 1
                amendements_apports = str(donnee_regies_historiques.cell(row=row, column=column).value)
                column += 1
                if amendements == "-1" and amendements_apports == "-1":
                    pass
                elif amendements == "-1" and not amendements_apports == "-1":
                    raise DependanceError(
                        "S'il n'y a pas d'amendements, il ne doit pas y avoir d'apports pour ceux-ci dans la feuille de calcul Régies historique à la rangée {}".format(
                            row))
                elif amendements_apports == "-1" and not amendements == "-1":
                    raise DependanceError(
                        "S'il n'y a pas d'apports d'amendements, il ne doit pas y avoir d'amendements dans la feuille de calcul Régies historique à la rangée {}".format(
                            row))
                else:
                    amendements_list = amendements.split(",")
                    amendements_apports_list = amendements_apports.split(",")
                    index = 0
                    for amendement in amendements_list:
                        try:
                            float(amendements_apports_list[index])
                        except ValueError:
                            raise TypeError(
                                "Apports d'amendement invalide dans la rangée {} et la colonne 14 de la feuille de calcul Régies historique".format(
                                    row, column))
                        except TypeError:
                            raise TypeError(
                                "Apports d'amendement invalide dans la rangée {} et la colonne 14 de la feuille de calcul Régies historique".format(
                                    row, column))
                        regie["amendements"].append(
                            {"amendement": amendement, "apport": float(amendements_apports_list[index])})
                        index += 1
                if donnee_regies_historiques.cell(row=row, column=column).value is None:
                    raise TypeError(
                        "Type de travail du sol est invalide dans la rangée {} et la colonne {} de la feuille de calcul Régies historique".format(
                            row, column))
                regie["travail_du_sol"] = {
                    "travail_du_sol": str(donnee_regies_historiques.cell(row=row, column=column).value)}
            else:
                raise TypeError(
                    "Erreur d'ordonancement de la rotation dans le rangée {} de la feuille de calcul Régies historique".format(
                        row))

            simulation_data[simulation][entreprise][champ][zone_gestion]["regies_sol_et_culture_historique"].append(
                regie)
        else:
            raise TypeError(
                "Nom de simulation invalide dans la rangée {} et la colonne {} de la feuille de calcul Régies historique".format(
                    row, column))
        column = initial_column
        row += 1
    json_final = formatter_vers_json_final(simulation_data, simulations_dict)
    return json_final


def ajout_caracteristiques_physiques(column, donnee_entreprises, row, zone_de_gestion_data):
    try:
        zone_de_gestion_data["taux_matiere_organique"] = float(donnee_entreprises.cell(row=row, column=column).value)
    except ValueError:
        raise TypeError(
            "Taux de matière organique invalide dans la rangée {} et la colonne {} de la feuille de calcul Entreprises".format(
                row, column))
    except TypeError:
        raise TypeError(
            "Taux de matière organique invalide dans la rangée {} et la colonne {} de la feuille de calcul Entreprises".format(
                row, column))
    column += 1
    if isinstance(donnee_entreprises.cell(row=row, column=column).value, str):
        zone_de_gestion_data["municipalite"] = donnee_entreprises.cell(row=row, column=column).value
    else:
        raise TypeError(
            "Municipalité invalide dans la rangée {} et la colonne {} de la feuille de calcul Entreprises".format(row,
                                                                                                                  column))
    column += 1
    if isinstance(donnee_entreprises.cell(row=row, column=column).value, str):
        zone_de_gestion_data["groupe_textural"] = donnee_entreprises.cell(row=row, column=column).value
    else:
        raise TypeError(
            "Groupe textural invalide dans la rangée {} et la colonne {} de la feuille de calcul Entreprises".format(
                row,
                column))
    column += 1
    if isinstance(donnee_entreprises.cell(row=row, column=column).value, str):
        zone_de_gestion_data["classe_de_drainage"] = donnee_entreprises.cell(row=row, column=column).value
    else:
        raise TypeError(
            "Classe de drainage invalide dans la rangée {} et la colonne {} de la feuille de calcul Entreprises".format(
                row,
                column))
    column += 1
    try:
        if float(donnee_entreprises.cell(row=row, column=column).value) == -1:
            zone_de_gestion_data["masse_volumique_apparente"] = None
        else:
            zone_de_gestion_data["masse_volumique_apparente"] = float(
                donnee_entreprises.cell(row=row, column=column).value)
    except ValueError:
        raise TypeError(
            "Masse volumique apparente invalide dans la rangée {} et la colonne {} de la feuille de calcul Entreprises".format(
                row, column))
    except TypeError:
        raise TypeError(
            "Masse volumique apparente invalide dans la rangée {} et la colonne {} de la feuille de calcul Entreprises".format(
                row, column))
    column += 1
    try:
        zone_de_gestion_data["profondeur"] = float(donnee_entreprises.cell(row=row, column=column).value)
    except ValueError:
        raise TypeError(
            "Profondeur invalide dans la rangée {} et la colonne {} de la feuille de calcul Entreprises".format(row,
                                                                                                                column))
    except TypeError:
        raise TypeError(
            "Profondeur invalide dans la rangée {} et la colonne {} de la feuille de calcul Entreprises".format(row,
                                                                                                                column))
    column += 1
    try:
        zone_de_gestion_data["superficie_de_la_zone"] = float(donnee_entreprises.cell(row=row, column=column).value)
    except ValueError:
        raise TypeError(
            "Superficie de la zone invalide dans la rangée {} et la colonne {} de la feuille de calcul Entreprises".format(
                row, column))
    except TypeError:
        raise TypeError(
            "Superficie de la zone invalide dans la rangée {} et la colonne {} de la feuille de calcul Entreprises".format(
                row, column))
    zone_de_gestion_data["regies_sol_et_culture_projection"] = []
    zone_de_gestion_data["regies_sol_et_culture_historique"] = []


def formatter_vers_json_final(simulation_data, simualtion_dict):
    json_final = {"simulations": []}
    for simulation in simualtion_dict.keys():
        simulation_final = {}
        simulation_final["nom_simulation"] = simualtion_dict[simulation]["nom_simulation"]
        simulation_final["duree_projection"] = simualtion_dict[simulation]["duree_projection"]
        simulation_final["annee_initiale_projection"] = simualtion_dict[simulation]["annee_initiale_projection"]
        simulation_final["annee_finale_projection"] = simualtion_dict[simulation]["annee_finale_projection"]
        if len(list(simulation_data[simulation].keys())) == 0:
            raise ValueError("Il n'y pas d'entreprise associé à la simulation {}".format(simulation))
        entreprise = list(simulation_data[simulation].keys())[0]
        simulation_final["entreprise_agricole"] = {"nom": entreprise}
        simulation_final["entreprise_agricole"]["champs"] = []
        for champ in simulation_data[simulation][entreprise].keys():
            champ_final = {"nom": champ,
                           "zones_de_gestion": []}
            for zone_de_gestion in simulation_data[simulation][entreprise][champ].keys():
                champ_final["zones_de_gestion"].append(simulation_data[simulation][entreprise][champ][zone_de_gestion])
            simulation_final["entreprise_agricole"]["champs"].append(champ_final)
        json_final["simulations"].append(simulation_final)
    return json_final


def row_check(donnee, row, number_of_columns, worksheet_name):
    if donnee.cell(row=row, column=1).value is not None:
        return True
    else:
        column = 2
        while column <= number_of_columns:
            if donnee.cell(row=row, column=column).value is None:
                pass
            else:
                raise TypeError(
                    "La rangée {} de la feuille de calcul {} à la colonne 1 n'a pas de valeur mais la rangée contient des données. La rangée devrait être soit complètement vide ou complètement pleine.".format(
                        row, worksheet_name))
            column += 1
        return False


if __name__ == "__main__":
    path = 'C:\\Users\\Samuel\\Documents\\Stage IRDA\\Test_de_sensibilite.xlsx'
    print(map_excel_to_json(path))
